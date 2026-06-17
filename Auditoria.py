import calendar
import hashlib
import io
import os
import re
import time
import unicodedata
from datetime import datetime, timedelta

import gspread
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import pytz
import streamlit as st
from gspread.exceptions import APIError
from oauth2client.service_account import ServiceAccountCredentials

# --- 1. CONFIGURACIÓN TÉCNICA ---
JSON_FILE = "creds_nsg.json"
USUARIOS_FILE = "usuarios.json"
ID_LIBRO = "13ZF5TXwgEZSlrODQFF43Rvs4JmB19s6V0KNV1l72RHA"
ID_LIBRO_RRHH = "1hXxm3yOx7lwzbDuUAUauKl-VoW7L-5UdfkFbjYg-s7g"
SCOPE = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive",
]
LOGO_FILENAME = "LOGO NSG SFONDO.png"
TIMEZONE = "America/Mexico_City"
CORTES_DICT = {"11:00 AM (3h)": 3, "14:00 PM (6h)": 6, "17:00 PM (9h)": 9}
AREAS_DEFAULT = ["MOLDEO", "CORAZONES", "CORTE", "ENSAMBLE"]
PIEZA_TERMINADA = "PIEZA TERMINADA"
DIAS_ES = {
    0: "Lunes",
    1: "Martes",
    2: "Miércoles",
    3: "Jueves",
    4: "Viernes",
    5: "Sábado",
    6: "Domingo",
}
# Semana operativa NSG: jueves a miércoles (sin sábado ni domingo)
SEMANA_NSG = ["Jueves", "Viernes", "Lunes", "Martes", "Miércoles"]

LIDERES_EXCLUIDOS = [
    "JOSÉ ANTONIO REYES RUBIO",
    "LUIS DAVID ESPINOSA TORRES",
    "ARELI PALOMA FLORES GARFIAS",
]

LIDERES_AREAS = {
    "JOSÉ ANTONIO REYES RUBIO": ["MOLDEO / CORAZONES"],
    "LUIS DAVID ESPINOSA TORRES": ["CORTE"],
    "ARELI PALOMA FLORES GARFIAS": ["ENSAMBLE"],
}

MOTIVOS_PARO = [
    "SIN PARO",
    "FALLA MECANICA",
    "FALLA ELECTRICA",
    "FALTA DE MATERIAL",
    "ESPERA DE PROCESO ANTERIOR",  # ← Flujo de cadena
    "FALLA DE HERRAMENTAL / MOLDES",  # ← Ingeniería de moldes
    "FALLA DE SERVICIOS (AIRE/GAS)",  # ← Infraestructura
    "ESPERA DE LIBERACION / CALIDAD",  # ← Control administrativo
    "AJUSTE OPERATIVO / PUESTA A PUNTO",  # ← Proceso térmico
    "CAMBIO DE MODELO / SET-UP",
    "AUSENCIA DE OPERADOR",
    "JUNTA DE CALIDAD / SEGURIDAD",
    "LIMPIEZA / 5S",
    "OTRO (ESPECIFICAR EN NOTAS)",
]


def ahora_local():
    return datetime.now(pytz.timezone(TIMEZONE))


def normalizar_clave(texto):
    texto = "" if texto is None else str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9]", "", texto)


def obtener_version_hoja(nombre_hoja):
    return st.session_state.get(f"version_hoja_{nombre_hoja}", 0)


def invalidar_cache_hoja(nombre_hoja):
    clave = f"version_hoja_{nombre_hoja}"
    st.session_state[clave] = st.session_state.get(clave, 0) + 1
    # Limpia el caché global para que TODOS los usuarios vean datos frescos
    leer_datos_seguro.clear()


def encontrar_columna(df, aliases, contiene_todos=None):
    if df.empty:
        return None

    aliases_norm = {normalizar_clave(alias) for alias in aliases}
    for col in df.columns:
        if normalizar_clave(col) in aliases_norm:
            return col

    if contiene_todos:
        tokens = [normalizar_clave(token) for token in contiene_todos]
        for col in df.columns:
            clave = normalizar_clave(col)
            if all(token in clave for token in tokens):
                return col

    return None


def preparar_dataframe(nombre_hoja, fila_encabezado=0):
    try:
        df = leer_datos_seguro(
            nombre_hoja, fila_encabezado, obtener_version_hoja(nombre_hoja)
        )
    except RuntimeError as exc:
        st.warning(f"⚠️ {exc}")
        df = pd.DataFrame()
    columnas = {
        "pieza": encontrar_columna(df, ["PIEZA"]),
        "area": encontrar_columna(df, ["AREA", "PROCESO", "ÁREA"]),
        "proceso": encontrar_columna(df, ["PROCESO", "AREA", "ÁREA"]),
        "subproceso": encontrar_columna(
            df,
            ["SUBPROCESO", "SUB PROCESO", "SUB_PROCESO", "SUB PRO CESO"],
            contiene_todos=["SUB", "CESO"],
        ),
        "activo": encontrar_columna(df, ["ACTIVO", "ESTATUS", "COL_4"]),
        "total": encontrar_columna(df, ["TOTAL"]),
        "real": encontrar_columna(df, ["REAL"]),
        "fecha": encontrar_columna(df, ["FECHA"]),
        "corte": encontrar_columna(df, ["CORTE"]),
        "pzxh": encontrar_columna(df, ["PZ X H", "PZXH", "PZS X H", "PZS/H"]),
    }
    return df, columnas


def validar_columnas(columnas, requeridas):
    faltantes = [nombre for nombre in requeridas if not columnas.get(nombre)]
    return faltantes


def convertir_serie_numerica(serie):
    serie_limpia = (
        serie.astype(str)
        .str.strip()
        .str.replace(",", "", regex=False)
        .str.replace(r"[^\d\.\-]", "", regex=True)
    )
    return pd.to_numeric(serie_limpia, errors="coerce")


@st.cache_resource
def obtener_cliente():
    try:
        if "gspread_creds" in st.secrets:
            _creds_dict = dict(st.secrets["gspread_creds"])
            _creds = ServiceAccountCredentials.from_json_keyfile_dict(
                _creds_dict, SCOPE
            )
            return gspread.authorize(_creds)
    except Exception:
        pass
    # Entorno local: usa el archivo (en .gitignore, nunca al repo)
    creds = ServiceAccountCredentials.from_json_keyfile_name(JSON_FILE, SCOPE)
    return gspread.authorize(creds)


def conectar_libro():
    try:
        cliente = obtener_cliente()
        return cliente.open_by_key(ID_LIBRO)
    except Exception as e:
        if "429" in str(e):
            st.warning(
                "⚠️ Google Sheets alcanzó su límite de lectura. La App usará datos en caché."
            )
        return None


@st.cache_data(ttl=600)
def leer_datos_seguro(nombre_hoja, fila_encabezado=0, version=0):
    libro = conectar_libro()
    if not libro:
        raise RuntimeError(
            f"Sin conexión a Google Sheets — '{nombre_hoja}' no pudo cargarse."
        )
    hoja = libro.worksheet(nombre_hoja)
    datos = leer_hoja_con_reintentos(hoja)
    if len(datos) <= fila_encabezado:
        return pd.DataFrame()
    nombres = datos[fila_encabezado]
    df = pd.DataFrame(datos[fila_encabezado + 1 :])
    df.columns = [
        str(nombre).strip().upper() if nombre else f"COL_{indice}"
        for indice, nombre in enumerate(nombres)
    ]
    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()
    return df


@st.cache_data(ttl=600)
def leer_datos_rrhh(version=0):
    _ = version
    try:
        cliente = obtener_cliente()
    except Exception as exc:
        raise RuntimeError(
            f"Sin conexión a Google Sheets — RRHH no pudo cargarse: {exc}"
        ) from exc
    libro_rrhh = cliente.open_by_key(ID_LIBRO_RRHH)
    hoja = libro_rrhh.worksheet("REGISTRO")
    datos = leer_hoja_con_reintentos(hoja)
    if not datos or len(datos) < 2:
        return pd.DataFrame()
    nombres = datos[0]
    df = pd.DataFrame(datos[1:])
    seen: dict = {}
    cols_unicos = []
    for i, n in enumerate(nombres):
        base = str(n).strip().upper() if n else f"COL_{i}"
        if base in seen:
            seen[base] += 1
            cols_unicos.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 0
            cols_unicos.append(base)
    df.columns = cols_unicos
    df = df.apply(lambda s: s.astype(str).str.strip())
    return df


def es_error_cuota(exc):
    mensaje = str(exc)
    return "429" in mensaje or "Read requests per minute per user" in mensaje


def leer_hoja_con_reintentos(hoja, max_intentos=4, pausa_inicial=1.5):
    ultimo_error = None
    for intento in range(max_intentos):
        try:
            return hoja.get_all_values()
        except APIError as exc:
            ultimo_error = exc
            if not es_error_cuota(exc) or intento == max_intentos - 1:
                raise
            time.sleep(pausa_inicial * (2**intento))
        except Exception as exc:
            ultimo_error = exc
            if not es_error_cuota(exc) or intento == max_intentos - 1:
                raise
            time.sleep(pausa_inicial * (2**intento))
    if ultimo_error:
        raise ultimo_error
    return []


def _ejecutar_con_reintentos(fn, max_intentos=3, pausa_inicial=1.5):
    """Retry con backoff exponencial para operaciones de lectura/escritura en Sheets."""
    ultimo_error = None
    for _intento in range(max_intentos):
        try:
            return fn()
        except Exception as _exc:
            ultimo_error = _exc
            if not es_error_cuota(_exc) or _intento == max_intentos - 1:
                raise
            time.sleep(pausa_inicial * (2**_intento))
    if ultimo_error:
        raise ultimo_error


def obtener_color_nsg(valor):
    if valor >= 85:
        return "#2ecc71"
    if valor >= 80:
        return "#f39c12"
    if valor >= 70:
        return "#e67e22"
    return "#E32B13"


def filtrar_bdd_activa(df_bdd, columnas_bdd):
    if df_bdd.empty:
        return df_bdd
    col_sub = columnas_bdd["subproceso"]
    col_activo = columnas_bdd["activo"]
    if not col_sub:
        return pd.DataFrame()
    df_filtrado = df_bdd.copy()
    if col_activo:
        df_filtrado = df_filtrado[
            df_filtrado[col_activo].astype(str).str.upper().str.strip() == "TRUE"
        ].copy()
    df_filtrado = df_filtrado[df_filtrado[col_sub].str.len() > 1].copy()
    df_filtrado = df_filtrado[df_filtrado[col_sub] != "0"].copy()
    return df_filtrado


def obtener_lista_areas(df_programa, columnas_programa):
    col_area = columnas_programa["area"]
    if not df_programa.empty and col_area:
        areas = [
            area
            for area in df_programa[col_area].unique().tolist()
            if area and normalizar_clave(area) != "AREA"
        ]
        if areas:
            return areas
    return AREAS_DEFAULT


def obtener_areas_con_programa(df_programa, col_prog, fecha_sel):
    """Áreas que tienen al menos una pieza programada para fecha_sel."""
    if df_programa.empty or not col_prog.get("area") or not col_prog.get("fecha"):
        return obtener_lista_areas(df_programa, col_prog)
    _df_f = df_programa[df_programa[col_prog["fecha"]] == fecha_sel]
    if _df_f.empty:
        return obtener_lista_areas(df_programa, col_prog)
    _areas = [
        a
        for a in _df_f[col_prog["area"]].unique().tolist()
        if a and normalizar_clave(a) != "AREA"
    ]
    return _areas if _areas else obtener_lista_areas(df_programa, col_prog)


def sugerir_corte_actual():
    hora_actual = ahora_local().time()
    if hora_actual < datetime.strptime("11:30", "%H:%M").time():
        return 0
    if hora_actual < datetime.strptime("14:30", "%H:%M").time():
        return 1
    return 2


def calcular_bono(p):
    """Algoritmo NSG-RH-AC-002: doble tramo lineal, máx 40 pts."""
    if p <= 0:
        return 0.0
    elif p < 70:
        return (p / 70) * 15
    elif p < 100:
        return 15 + ((p - 70) / 30) * 25
    else:
        return 40.0


def obtener_plan_del_dia(df_programa, columnas_programa, fecha_sel, area_sel):
    requeridas = validar_columnas(
        columnas_programa, ["fecha", "area", "pieza", "total"]
    )
    if df_programa.empty or requeridas:
        return pd.DataFrame()
    df_plan = df_programa[
        (df_programa[columnas_programa["fecha"]] == fecha_sel)
        & (df_programa[columnas_programa["area"]] == area_sel)
    ].copy()
    if area_sel.upper() == "MOLDEO" and not df_plan.empty:
        df_plan = df_plan[
            df_plan[columnas_programa["pieza"]].str.contains(
                "GENERAL|VACIADO|ADOBES", case=False, na=False
            )
        ]
    df_plan[columnas_programa["total"]] = convertir_serie_numerica(
        df_plan[columnas_programa["total"]]
    )
    df_plan = df_plan[df_plan[columnas_programa["total"]] > 0].copy()
    return df_plan


def calcular_resumen(
    df_plan_dia, col_prog, df_bdd, col_bdd, df_auditorias, fecha_sel, area_sel
):
    avance_global = 0.0
    df_resumen = pd.DataFrame()
    requeridas_plan = validar_columnas(col_prog, ["pieza", "total"])
    requeridas_bdd = validar_columnas(col_bdd, ["pieza", "proceso", "subproceso"])
    if df_plan_dia.empty or df_bdd.empty or requeridas_plan or requeridas_bdd:
        return avance_global, df_resumen
    df_sub_base_total = df_bdd[df_bdd[col_bdd["proceso"]] == area_sel][
        [col_bdd["pieza"], col_bdd["subproceso"]]
    ].copy()
    piezas_en_plan = df_plan_dia[col_prog["pieza"]].unique()
    df_base = df_sub_base_total[
        df_sub_base_total[col_bdd["pieza"]].isin(piezas_en_plan)
    ].copy()
    if df_base.empty:
        return avance_global, df_resumen
    df_base = pd.merge(
        df_base,
        df_plan_dia[[col_prog["pieza"], col_prog["total"]]],
        left_on=col_bdd["pieza"],
        right_on=col_prog["pieza"],
        how="left",
    )
    df_base[col_prog["total"]] = convertir_serie_numerica(
        df_base[col_prog["total"]]
    ).fillna(0)
    if not df_auditorias.empty:
        col_aud = {
            "fecha": encontrar_columna(df_auditorias, ["FECHA"]),
            "area": encontrar_columna(df_auditorias, ["AREA", "ÁREA"]),
            "pieza": encontrar_columna(df_auditorias, ["PIEZA"]),
            "subproceso": encontrar_columna(
                df_auditorias,
                ["SUBPROCESO", "SUB PROCESO", "SUB_PROCESO"],
                contiene_todos=["SUB", "CESO"],
            ),
            "real": encontrar_columna(df_auditorias, ["REAL"]),
        }
        if not validar_columnas(
            col_aud, ["fecha", "area", "pieza", "subproceso", "real"]
        ):
            df_aud_hoy = df_auditorias[
                (df_auditorias[col_aud["fecha"]] == fecha_sel)
                & (df_auditorias[col_aud["area"]] == area_sel)
            ].copy()
            df_aud_hoy[col_aud["real"]] = pd.to_numeric(
                df_aud_hoy[col_aud["real"]], errors="coerce"
            ).fillna(0)
            df_max_real = (
                df_aud_hoy.groupby([col_aud["pieza"], col_aud["subproceso"]])[
                    col_aud["real"]
                ]
                .max()
                .reset_index()
            )
            df_final = pd.merge(
                df_base,
                df_max_real,
                left_on=[col_bdd["pieza"], col_bdd["subproceso"]],
                right_on=[col_aud["pieza"], col_aud["subproceso"]],
                how="left",
            ).fillna(0)
        else:
            df_final = df_base.assign(REAL=0)
            col_aud = {"real": "REAL"}
    else:
        df_final = df_base.assign(REAL=0)
        col_aud = {"real": "REAL"}
    total_seguro = convertir_serie_numerica(df_final[col_prog["total"]]).fillna(0)
    real_seguro = convertir_serie_numerica(df_final[col_aud["real"]]).fillna(0)
    df_final["% REAL"] = 0.0
    mask_total_valido = total_seguro > 0
    df_final.loc[mask_total_valido, "% REAL"] = (
        real_seguro[mask_total_valido] / total_seguro[mask_total_valido] * 100
    )
    df_resumen = df_final[
        [
            col_bdd["pieza"],
            col_bdd["subproceso"],
            col_prog["total"],
            col_aud["real"],
            "% REAL",
        ]
    ].copy()
    df_resumen.columns = ["PIEZA", "SUBPROCESO", "PROGRAMADO", "AVANCE", "% REAL"]
    avance_global = round(df_final["% REAL"].mean(), 1)
    return avance_global, df_resumen


def obtener_auditorias_hoy(df_auditorias, fecha_sel, area_sel):
    if df_auditorias.empty:
        return pd.DataFrame(), {}
    columnas = {
        "fecha": encontrar_columna(df_auditorias, ["FECHA"]),
        "area": encontrar_columna(df_auditorias, ["AREA", "ÁREA"]),
        "corte": encontrar_columna(df_auditorias, ["CORTE"]),
        "pieza": encontrar_columna(df_auditorias, ["PIEZA"]),
        "subproceso": encontrar_columna(
            df_auditorias,
            ["SUBPROCESO", "SUB PROCESO", "SUB_PROCESO"],
            contiene_todos=["SUB", "CESO"],
        ),
        "real": encontrar_columna(df_auditorias, ["REAL"]),
    }
    faltantes = validar_columnas(columnas, ["fecha", "area", "pieza", "subproceso"])
    if faltantes:
        return pd.DataFrame(), columnas
    df_hoy = df_auditorias[
        (df_auditorias[columnas["fecha"]] == fecha_sel)
        & (df_auditorias[columnas["area"]] == area_sel)
    ].copy()
    return df_hoy, columnas


def obtener_piezas_pendientes(
    df_plan_dia, col_prog, df_bdd, col_bdd, df_aud_hoy, col_aud, area_sel, corte_sel
):
    if df_plan_dia.empty or df_bdd.empty:
        return []
    faltantes_plan = validar_columnas(col_prog, ["pieza"])
    faltantes_bdd = validar_columnas(col_bdd, ["pieza", "proceso", "subproceso"])
    if faltantes_plan or faltantes_bdd:
        return []
    piezas_validas = df_plan_dia[col_prog["pieza"]].unique()
    piezas_pendientes = []
    for pieza in piezas_validas:
        mask = (df_bdd[col_bdd["pieza"]] == pieza) & (
            df_bdd[col_bdd["proceso"]] == area_sel
        )
        subs_totales = df_bdd[mask][col_bdd["subproceso"]].unique()
        reps_pieza = []
        if not df_aud_hoy.empty and col_aud.get("corte") and col_aud.get("pieza"):
            reps_pieza = df_aud_hoy[
                (df_aud_hoy[col_aud["corte"]] == corte_sel)
                & (df_aud_hoy[col_aud["pieza"]] == pieza)
            ][col_aud["subproceso"]].tolist()
        if any(sub for sub in subs_totales if sub not in reps_pieza):
            piezas_pendientes.append(pieza)
    return piezas_pendientes


def guardar_registro(
    fecha_sel,
    area_sel,
    corte_sel,
    pieza_sel,
    subproceso_sel,
    real,
    meta,
    ops,
    mot,
    mins,
    notas,
    usuario,
):
    hora = ahora_local().strftime("%H:%M:%S")
    _nota_completa = f"[{mot}-{int(mins)}min]"
    if notas.strip():
        _nota_completa += f" {notas.strip()}"
    fila = [
        fecha_sel,
        area_sel,
        corte_sel,
        pieza_sel,
        subproceso_sel,
        int(real),
        int(meta),
        int(real - meta),
        int(ops),
        _nota_completa,
        hora,
        usuario,
    ]
    try:
        def _escribir():
            _libro = conectar_libro()
            if not _libro:
                raise RuntimeError("Sin conexión al libro")
            _libro.worksheet("AUDITAR").append_row(fila)
        _ejecutar_con_reintentos(_escribir, max_intentos=4, pausa_inicial=3)
        return True
    except Exception as exc:
        if es_error_cuota(exc):
            st.error(
                "⚠️ Google Sheets alcanzó su límite momentáneo. Espera 30 segundos e intenta de nuevo."
            )
        else:
            st.error(f"Error al guardar el registro: {exc}")
        return False


def render_kpis(avance_global, df_resumen_final):
    k1, k2, k3 = st.columns(3)
    with k1:
        st.markdown(
            f"<div class='metric-card'><small>EFICIENCIA GLOBAL</small><h2>{avance_global}%</h2></div>",
            unsafe_allow_html=True,
        )
    with k2:
        meta_turno = (
            int(df_resumen_final["PROGRAMADO"].sum())
            if not df_resumen_final.empty
            else 0
        )
        st.markdown(
            f"<div class='metric-card'><small>META TURNO</small><h2>{meta_turno} PZS</h2></div>",
            unsafe_allow_html=True,
        )
    with k3:
        real_turno = (
            int(df_resumen_final["AVANCE"].sum()) if not df_resumen_final.empty else 0
        )
        st.markdown(
            f"<div class='metric-card'><small>REAL TURNO</small><h2>{real_turno} PZS</h2></div>",
            unsafe_allow_html=True,
        )


def render_graficos(avance_global, df_resumen_final):
    c_g, c_b = st.columns([1, 2])
    with c_g:
        color_g = obtener_color_nsg(avance_global)
        fig = go.Figure(
            go.Indicator(
                mode="gauge+number",
                value=avance_global,
                gauge={
                    "bar": {"color": color_g},
                    "axis": {"range": [0, 100]},
                    "bgcolor": "#f0f2f6",
                    "steps": [
                        {"range": [0, 80], "color": "rgba(227, 43, 19, 0.05)"},
                        {"range": [80, 85], "color": "rgba(241, 196, 15, 0.1)"},
                        {"range": [85, 100], "color": "rgba(46, 204, 113, 0.1)"},
                    ],
                },
            )
        )
        fig.update_layout(height=280, margin=dict(l=30, r=30, t=30, b=10))
        st.plotly_chart(fig, use_container_width=True)
    with c_b:
        if not df_resumen_final.empty:
            fig_bar = go.Figure()
            fig_bar.add_trace(
                go.Bar(
                    name="Meta",
                    x=df_resumen_final["SUBPROCESO"],
                    y=df_resumen_final["PROGRAMADO"],
                    marker_color="#610B0B",
                )
            )
            fig_bar.add_trace(
                go.Bar(
                    name="Real",
                    x=df_resumen_final["SUBPROCESO"],
                    y=df_resumen_final["AVANCE"],
                    marker_color="#E32B13",
                )
            )
            fig_bar.update_layout(
                barmode="group",
                height=280,
                margin=dict(l=0, r=0, t=20, b=0),
                legend=dict(orientation="h", y=1.1),
            )
            st.plotly_chart(fig_bar, use_container_width=True)


def render_estatus_detallado(df_resumen_final):
    st.markdown("### ESTATUS DETALLADO")
    if df_resumen_final.empty:
        st.info("No hay datos suficientes para mostrar el estatus detallado.")
        return
    for _, row in df_resumen_final.iterrows():
        val_n = round(row["% REAL"], 1)
        color_n = obtener_color_nsg(val_n)
        ancho_barra = min(val_n, 100)
        st.markdown(
            f"""
            <div style='background:white; padding:10px 15px; border-radius:8px; border-left:5px solid {color_n}; margin-bottom:8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);'>
                <div style='display:flex; justify-content:space-between; align-items:center;'>
                    <div style='font-size:14px;'><b>{row['PIEZA']}</b> - <span style='color:#555;'>{row['SUBPROCESO']}</span></div>
                    <div style='color:{color_n}; font-weight:bold; font-size:16px;'>{val_n}%</div>
                </div>
                <div style='background:#eee; height:6px; border-radius:3px; margin: 6px 0;'>
                    <div style='background:{color_n}; width:{ancho_barra}%; height:100%; border-radius:3px;'></div>
                </div>
                <div style='display:flex; justify-content:space-between; font-size:11px; color:gray;'>
                    <span>Prog: <b>{row['PROGRAMADO']}</b></span>
                    <span>Real: <b>{row['AVANCE']}</b></span>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_capacidades(df_s, col_bdd, pieza_sel):
    st.divider()
    titulo = f"CONSULTAR CAPACIDADES (PZ/H) - PIEZA: {pieza_sel if pieza_sel else 'NO SELECCIONADA'}"
    with st.expander(titulo):
        if not df_s.empty and col_bdd.get("subproceso") and col_bdd.get("pzxh"):
            st.write(f"Capacidades estándar para la pieza: **{pieza_sel}**")
            st.table(df_s[[col_bdd["subproceso"], col_bdd["pzxh"]]])
        else:
            st.warning("Selecciona una pieza válida para ver sus capacidades.")


def obtener_datos_unificados(
    df_auditorias, df_programa, df_bdd, col_prog, col_bdd, f_ini, f_fin
):
    col_aud = {
        "fecha": encontrar_columna(df_auditorias, ["FECHA"]),
        "pieza": encontrar_columna(df_auditorias, ["PIEZA"]),
        "subproceso": encontrar_columna(
            df_auditorias,
            ["SUBPROCESO", "SUB PROCESO", "SUB_PROCESO"],
            contiene_todos=["SUB", "CESO"],
        ),
        "real": encontrar_columna(df_auditorias, ["REAL"]),
    }
    if df_auditorias.empty or df_programa.empty or df_bdd.empty:
        return pd.DataFrame(), col_aud
    df_a_v = df_auditorias.copy()
    df_a_v[col_aud["real"]] = pd.to_numeric(
        df_a_v[col_aud["real"]], errors="coerce"
    ).fillna(0)
    df_p_v = df_programa.copy()
    df_p_v["FECHA_DT"] = pd.to_datetime(
        df_p_v[col_prog["fecha"]], format="%d/%m/%Y", errors="coerce"
    )
    df_p_v = df_p_v[
        (df_p_v["FECHA_DT"].dt.date >= f_ini) & (df_p_v["FECHA_DT"].dt.date <= f_fin)
    ]
    if df_p_v.empty:
        return pd.DataFrame(), col_aud
    mask_m = df_p_v[col_prog["area"]].str.upper() == "MOLDEO"
    df_p_m = df_p_v[
        mask_m
        & df_p_v[col_prog["pieza"]].str.contains(
            "GENERAL|VACIADO|ADOBES", case=False, na=False
        )
    ]
    df_p_o = df_p_v[~mask_m]
    df_p_final = pd.concat([df_p_m, df_p_o])
    df_p_final[col_prog["total"]] = convertir_serie_numerica(
        df_p_final[col_prog["total"]]
    )
    df_p_final = df_p_final[df_p_final[col_prog["total"]] > 0].copy()
    df_max_a = (
        df_a_v.groupby([col_aud["fecha"], col_aud["pieza"], col_aud["subproceso"]])[
            col_aud["real"]
        ]
        .max()
        .reset_index()
    )
    df_base_v = pd.merge(
        df_p_final[
            [
                col_prog["fecha"],
                col_prog["area"],
                col_prog["pieza"],
                col_prog["total"],
                "FECHA_DT",
            ]
        ],
        df_bdd[[col_bdd["pieza"], col_bdd["subproceso"], col_bdd["proceso"]]],
        left_on=col_prog["pieza"],
        right_on=col_bdd["pieza"],
        how="inner",
    )
    df_base_v = df_base_v[df_base_v[col_prog["area"]] == df_base_v[col_bdd["proceso"]]]
    df_uni = pd.merge(
        df_base_v,
        df_max_a,
        left_on=[col_prog["fecha"], col_prog["pieza"], col_bdd["subproceso"]],
        right_on=[col_aud["fecha"], col_aud["pieza"], col_aud["subproceso"]],
        how="left",
    ).fillna(0)
    total_seguro = convertir_serie_numerica(df_uni[col_prog["total"]]).fillna(0)
    real_seguro = convertir_serie_numerica(df_uni[col_aud["real"]]).fillna(0)
    df_uni["% REAL"] = 0.0
    mask_total = total_seguro > 0
    df_uni.loc[mask_total, "% REAL"] = (
        real_seguro[mask_total] / total_seguro[mask_total] * 100
    )
    return df_uni, col_aud


def render_dashboard_direccion(df_auditorias, df_programa, df_bdd, col_prog, col_bdd):
    st.markdown("## 🏭 Resultados del Programa de Producción NSG")
    _hoy_dash = ahora_local().date()
    c1, c2 = st.columns(2)
    with c1:
        f_ini = st.date_input(
            "Analizar desde:", _hoy_dash.replace(day=1), key="dash_f1"
        )
    with c2:
        f_fin = st.date_input("Hasta:", _hoy_dash, key="dash_f2")

    hojas_vacias = []
    if df_auditorias.empty:
        hojas_vacias.append("AUDITAR")
    if df_programa.empty:
        hojas_vacias.append("PROGRAMA")
    if df_bdd.empty:
        hojas_vacias.append("BDD")
    if hojas_vacias:
        st.warning(
            f"⚠️ Esperando datos... No se detectaron registros en: {', '.join(hojas_vacias)}"
        )
        return

    df_uni, col_aud = obtener_datos_unificados(
        df_auditorias, df_programa, df_bdd, col_prog, col_bdd, f_ini, f_fin
    )
    if df_uni.empty:
        st.info("No hay registros que coincidan con el rango de fechas.")
        return

    # Filtro por área
    _areas_disp = ["TODAS LAS ÁREAS"] + sorted(
        df_uni[col_prog["area"]].astype(str).unique().tolist()
    )
    _area_sel = st.selectbox(
        "🏭 Filtrar por área:", _areas_disp, key="dash_area_filtro"
    )
    if _area_sel != "TODAS LAS ÁREAS":
        df_uni = df_uni[df_uni[col_prog["area"]] == _area_sel].copy()
        if df_uni.empty:
            st.info(f"No hay datos para el área **{_area_sel}** en este periodo.")
            return

    # Periodo anterior (misma duración, inmediatamente antes)
    _n_dias_periodo = (f_fin - f_ini).days + 1
    _f_ini_ant = f_ini - timedelta(days=_n_dias_periodo)
    _f_fin_ant = f_ini - timedelta(days=1)
    df_ant, _ = obtener_datos_unificados(
        df_auditorias, df_programa, df_bdd, col_prog, col_bdd, _f_ini_ant, _f_fin_ant
    )
    _cum_ant = df_ant["% REAL"].mean() if not df_ant.empty else None
    if _area_sel != "TODAS LAS ÁREAS" and not df_ant.empty:
        df_ant = df_ant[df_ant[col_prog["area"]] == _area_sel].copy()
        _cum_ant = df_ant["% REAL"].mean() if not df_ant.empty else None

    # ── KPIs base ────────────────────────────────────────────────────────────
    df_dias = df_uni.groupby("FECHA_DT")["% REAL"].mean().reset_index()
    total_dias = len(df_dias)
    dias_ganados = len(df_dias[df_dias["% REAL"] >= 80])
    dias_riesgo = len(df_dias[(df_dias["% REAL"] >= 70) & (df_dias["% REAL"] < 80)])
    dias_perdidos = len(df_dias[df_dias["% REAL"] < 70])
    cumplimiento_total = df_uni["% REAL"].mean()

    if not df_ant.empty:
        _df_dias_ant = df_ant.groupby("FECHA_DT")["% REAL"].mean().reset_index()
        _dias_ganados_ant = len(_df_dias_ant[_df_dias_ant["% REAL"] >= 80])
        _dias_riesgo_ant = len(
            _df_dias_ant[(_df_dias_ant["% REAL"] >= 70) & (_df_dias_ant["% REAL"] < 80)]
        )
        _dias_perdidos_ant = len(_df_dias_ant[_df_dias_ant["% REAL"] < 70])
    else:
        _dias_ganados_ant = _dias_riesgo_ant = _dias_perdidos_ant = None

    def _delta_html(cur, prev):
        if prev is None:
            return "<span style='font-size:11px;color:#95a5a6;'>sin periodo anterior</span>"
        d = cur - prev
        col = "#27ae60" if d >= 0 else "#e74c3c"
        arr = "▲" if d >= 0 else "▼"
        return f"<span style='font-size:13px;color:{col};font-weight:700;'>{arr} {abs(d):.1f}pp vs anterior</span>"

    def _delta_d(cur, prev, bueno_si_sube=True):
        if prev is None:
            return "<span style='font-size:11px;color:#95a5a6;'>—</span>"
        d = cur - prev
        if d == 0:
            return "<span style='font-size:11px;color:#95a5a6;'>= igual que anterior</span>"
        col = (
            ("#27ae60" if d > 0 else "#e74c3c")
            if bueno_si_sube
            else ("#e74c3c" if d > 0 else "#27ae60")
        )
        return f"<span style='font-size:13px;color:{col};font-weight:700;'>{'▲' if d > 0 else '▼'} {abs(d)} día(s)</span>"

    # ── Diagnóstico compartido ────────────────────────────────────────────────
    area_estrella = df_uni.groupby(col_prog["area"])["% REAL"].mean().idxmax()
    area_estrella_val = df_uni.groupby(col_prog["area"])["% REAL"].mean().max()
    sub_critico = df_uni.groupby(col_bdd["subproceso"])["% REAL"].mean().idxmin()
    mask_pt = ~(
        df_uni[col_prog["area"]].str.upper().isin(["CORAZONES", "MOLDEO"])
        | df_uni[col_prog["pieza"]].str.contains("GENERAL", case=False, na=False)
    )
    df_pt_only = df_uni[mask_pt]
    if not df_pt_only.empty:
        pieza_estrella = df_pt_only.groupby(col_prog["pieza"])["% REAL"].mean().idxmax()
        p_est_val = df_pt_only.groupby(col_prog["pieza"])["% REAL"].mean().max()
        texto_pieza = f"La pieza líder en <b>Producto Terminado</b> es <b>{pieza_estrella}</b> ({p_est_val:.1f}%)."
    else:
        texto_pieza = "No hay datos suficientes de piezas terminadas (Ensamble/Corte) en este rango."

    df_a_raw = df_auditorias.copy()
    c_f_raw = encontrar_columna(df_a_raw, ["FECHA"])
    c_n_raw = encontrar_columna(df_a_raw, ["NOTAS", "NOTA"])
    df_a_raw["FECHA_DT"] = pd.to_datetime(
        df_a_raw[c_f_raw], format="%d/%m/%Y", errors="coerce"
    )
    df_a_raw = df_a_raw[
        (df_a_raw["FECHA_DT"].dt.date >= f_ini)
        & (df_a_raw["FECHA_DT"].dt.date <= f_fin)
    ]
    df_a_raw["MOTIVO"] = (
        df_a_raw[c_n_raw].astype(str).str.extract(r"^\[(.*?)\]")[0] if c_n_raw else None
    )
    df_paros = df_a_raw[df_a_raw["MOTIVO"].notna() & (df_a_raw["MOTIVO"] != "SIN PARO")]
    paro_principal = "No hay paros registrados."
    if not df_paros.empty:
        paro_top = df_paros["MOTIVO"].value_counts().idxmax()
        paro_principal = f"El motivo de paro más frecuente fue <b>{paro_top}</b>."

    if cumplimiento_total >= 80:
        estado_planta = (
            "<span style='color:#27ae60;'>🟢 OPERACIÓN RENTABLE (&ge; 80%)</span>"
        )
        borde_color = "#27ae60"
    elif cumplimiento_total >= 70:
        estado_planta = (
            "<span style='color:#f39c12;'>🟡 ZONA DE RIESGO (70% - 79%)</span>"
        )
        borde_color = "#f39c12"
    else:
        estado_planta = (
            "<span style='color:#c0392b;'>🔴 ESTADO CRÍTICO (&lt; 70%)</span>"
        )
        borde_color = "#c0392b"

    # ── Proyección de cierre del mes en curso (días hábiles) ─────────────────
    _inicio_mes = _hoy_dash.replace(day=1)
    _dias_mes_total = calendar.monthrange(_hoy_dash.year, _hoy_dash.month)[1]
    _fin_mes = _hoy_dash.replace(day=_dias_mes_total)
    # Contar solo días hábiles (lunes–viernes)
    _dias_lab_trans = sum(
        1
        for _d in range((_hoy_dash - _inicio_mes).days + 1)
        if (_inicio_mes + timedelta(days=_d)).weekday() < 5
    )
    _dias_lab_rest = sum(
        1
        for _d in range(1, (_fin_mes - _hoy_dash).days + 1)
        if (_hoy_dash + timedelta(days=_d)).weekday() < 5
    )
    _dias_lab_total = _dias_lab_trans + _dias_lab_rest
    # Promedio necesario en días hábiles restantes para cerrar el mes en 80%:
    _prom_necesario = None
    if _dias_lab_rest > 0 and cumplimiento_total < 80:
        _num = 80 * _dias_lab_total - cumplimiento_total * _dias_lab_trans
        _prom_necesario = min(max(_num / _dias_lab_rest, 0), 100)

    # ── Rachas críticas por área ──────────────────────────────────────────────
    _df_dias_area = (
        df_uni.groupby(["FECHA_DT", col_prog["area"]])["% REAL"]
        .mean()
        .reset_index()
        .sort_values("FECHA_DT")
    )
    _rachas_criticas = []
    for _ar in _df_dias_area[col_prog["area"]].unique():
        _sub_ar = _df_dias_area[_df_dias_area[col_prog["area"]] == _ar].sort_values(
            "FECHA_DT"
        )
        _racha_actual = 0
        for _, _fila in _sub_ar.iterrows():
            _racha_actual = _racha_actual + 1 if _fila["% REAL"] < 70 else 0
        if _racha_actual >= 3:
            _rachas_criticas.append({"ÁREA": _ar, "DÍAS": _racha_actual})
    _df_rachas = pd.DataFrame(_rachas_criticas)

    # ════════════════════════════════════════════════════════════════════════
    # TABS
    # ════════════════════════════════════════════════════════════════════════
    tab_exec, tab_analisis, tab_fallas = st.tabs(
        ["📋 Resumen Ejecutivo", "📈 Análisis del Periodo", "📝 Detalle de Fallas"]
    )

    # ── TAB 1: RESUMEN EJECUTIVO ─────────────────────────────────────────────
    with tab_exec:
        with st.expander("📖 ¿Cómo leer este resumen?"):
            st.markdown("""
**Esta vista está diseñada para dirección — lectura en menos de 60 segundos.**

**KPIs de días:**
| Tarjeta | Qué mide | ▲▼ |
|---------|----------|----|
| 🎯 Cumplimiento | % promedio diario de avance vs. programa en el periodo | ▲ pp = mejora · ▼ pp = caída |
| 🟢 Meta lograda | Días donde se alcanzó ≥ 80% del programa | ▲ días = bueno |
| 🟡 Casi llegamos | Días en zona 70–79% — cerca pero sin llegar | ▲ días = señal de alerta |
| 🔴 Días críticos | Días donde se produjo < 70% de lo programado | ▲ días = urgente |

> Los **▲▼ vs anterior** comparan contra el periodo inmediatamente anterior de igual duración.

**Proyección de cierre del mes:**
- *Al ritmo del periodo analizado, el mes cierra en X%* — asume que los días restantes del mes tendrán el mismo promedio diario que el periodo seleccionado
- *Necesitas Y% en los días restantes* — qué promedio diario requerirían los días que faltan del mes para que el acumulado mensual llegue al 80%

**Tendencia reciente:** Los últimos días del periodo seleccionado. Verde = meta alcanzada, amarillo = zona de riesgo, rojo = día crítico.

**Motivos de paro:** Causas capturadas en bitácora que afectaron el programa. El tamaño de cada rebanada es proporcional a su frecuencia.
            """)
        if not _df_rachas.empty:
            for _, _r in _df_rachas.iterrows():
                st.error(
                    f"🚨 **{_r['ÁREA']}** lleva **{int(_r['DÍAS'])} días consecutivos** "
                    f"por debajo del 70% — intervención urgente recomendada."
                )
        st.markdown(
            f"""
<div style='display:flex;justify-content:space-between;gap:15px;margin-bottom:20px;flex-wrap:wrap;'>
    <div style='background:white;border-top:5px solid #8B1A1A;padding:15px;border-radius:8px;flex:1;text-align:center;min-width:150px;box-shadow:0 4px 6px rgba(0,0,0,0.05);'>
        <div style='font-size:12px;color:#7f8c8d;font-weight:bold;text-transform:uppercase;'>🎯 CUMPLIMIENTO DEL PROGRAMA</div>
        <div style='font-size:28px;color:#8B1A1A;font-weight:900;margin-top:5px;'>{cumplimiento_total:.1f}%</div>
        <div style='margin-top:4px;'>{_delta_html(cumplimiento_total, _cum_ant)}</div>
    </div>
    <div style='background:white;border-top:5px solid #95a5a6;padding:15px;border-radius:8px;flex:1;text-align:center;min-width:150px;box-shadow:0 4px 6px rgba(0,0,0,0.05);'>
        <div style='font-size:12px;color:#7f8c8d;font-weight:bold;text-transform:uppercase;'>🗓️ DÍAS TRABAJADOS (PERIODO)</div>
        <div style='font-size:28px;color:#7f8c8d;font-weight:900;margin-top:5px;'>{total_dias}</div>
    </div>
    <div style='background:white;border-top:5px solid #2ecc71;padding:15px;border-radius:8px;flex:1;text-align:center;min-width:150px;box-shadow:0 4px 6px rgba(0,0,0,0.05);'>
        <div style='font-size:12px;color:#7f8c8d;font-weight:bold;text-transform:uppercase;'>🟢 META LOGRADA (80% o más)</div>
        <div style='font-size:28px;color:#27ae60;font-weight:900;margin-top:5px;'>{dias_ganados}</div>
        <div style='margin-top:4px;'>{_delta_d(dias_ganados, _dias_ganados_ant, bueno_si_sube=True)}</div>
    </div>
    <div style='background:white;border-top:5px solid #f39c12;padding:15px;border-radius:8px;flex:1;text-align:center;min-width:150px;box-shadow:0 4px 6px rgba(0,0,0,0.05);'>
        <div style='font-size:12px;color:#7f8c8d;font-weight:bold;text-transform:uppercase;'>🟡 CASI LLEGAMOS (70% - 79%)</div>
        <div style='font-size:28px;color:#f39c12;font-weight:900;margin-top:5px;'>{dias_riesgo}</div>
        <div style='margin-top:4px;'>{_delta_d(dias_riesgo, _dias_riesgo_ant, bueno_si_sube=False)}</div>
    </div>
    <div style='background:white;border-top:5px solid #e74c3c;padding:15px;border-radius:8px;flex:1;text-align:center;min-width:150px;box-shadow:0 4px 6px rgba(0,0,0,0.05);'>
        <div style='font-size:12px;color:#7f8c8d;font-weight:bold;text-transform:uppercase;'>🔴 DÍAS CRÍTICOS (Menos de 70%)</div>
        <div style='font-size:28px;color:#c0392b;font-weight:900;margin-top:5px;'>{dias_perdidos}</div>
        <div style='margin-top:4px;'>{_delta_d(dias_perdidos, _dias_perdidos_ant, bueno_si_sube=False)}</div>
    </div>
</div>
""",
            unsafe_allow_html=True,
        )
        st.divider()
        st.markdown(
            f"""
    <div style='background-color:#f8f9fa;border-left:6px solid {borde_color};padding:15px;border-radius:5px;margin-bottom:20px;box-shadow:0 2px 4px rgba(0,0,0,0.05);'>
        <h4 style='margin-top:0;color:#2c3e50;'>🤖 Diagnóstico del Sistema: {estado_planta}</h4>
        <ul style='margin-bottom:0;font-size:15px;color:#34495e;'>
            <li><b>Líder de Área:</b> <b>{area_estrella}</b> lidera con un <b>{area_estrella_val:.1f}%</b> de cumplimiento.</li>
            <li><b>Análisis de Salida:</b> {texto_pieza}</li>
            <li><b>Alerta de Proceso:</b> El subproceso <b>{sub_critico}</b> es el mayor cuello de botella actual.</li>
            <li><b>Causa Raíz Operativa:</b> {paro_principal}</li>
        </ul>
    </div>
    """,
            unsafe_allow_html=True,
        )
        st.divider()
        st.markdown(
            f"#### 🔮 Proyección de Cierre — {_hoy_dash.strftime('%B %Y').capitalize()}"
        )
        _cp1, _cp2 = st.columns(2)
        with _cp1:
            _pc = (
                "#27ae60"
                if cumplimiento_total >= 80
                else ("#f39c12" if cumplimiento_total >= 70 else "#e74c3c")
            )
            st.markdown(
                f"<div style='background:white;border-left:6px solid {_pc};padding:16px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,0.05);'>"
                f"<div style='font-size:12px;color:#7f8c8d;font-weight:bold;text-transform:uppercase;'>AL RITMO DEL PERIODO ANALIZADO, EL MES CIERRA EN</div>"
                f"<div style='font-size:32px;font-weight:900;color:{_pc};margin-top:4px;'>{cumplimiento_total:.1f}%</div>"
                f"<div style='font-size:12px;color:#95a5a6;margin-top:4px;'>{_dias_lab_rest} día(s) restantes en el mes · basado en ritmo del periodo seleccionado</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
        with _cp2:
            if _prom_necesario is not None:
                _rq = (
                    "#27ae60"
                    if _prom_necesario <= 85
                    else ("#f39c12" if _prom_necesario <= 95 else "#e74c3c")
                )
                st.markdown(
                    f"<div style='background:white;border-left:6px solid {_rq};padding:16px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,0.05);'>"
                    f"<div style='font-size:12px;color:#7f8c8d;font-weight:bold;text-transform:uppercase;'>PROMEDIO NECESARIO PARA CERRAR EL MES EN 80%</div>"
                    f"<div style='font-size:32px;font-weight:900;color:{_rq};margin-top:4px;'>{_prom_necesario:.1f}%</div>"
                    f"<div style='font-size:12px;color:#95a5a6;margin-top:4px;'>en los {_dias_lab_rest} día(s) restantes del mes</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            elif _dias_lab_rest == 0:
                st.markdown(
                    f"<div style='background:white;border-left:6px solid #7f8c8d;padding:16px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,0.05);'>"
                    f"<div style='font-size:12px;color:#7f8c8d;font-weight:bold;text-transform:uppercase;'>CIERRE DE MES</div>"
                    f"<div style='font-size:32px;font-weight:900;color:#7f8c8d;margin-top:4px;'>Hoy es el último día</div>"
                    f"<div style='font-size:12px;color:#95a5a6;margin-top:4px;'>El mes cierra con el cumplimiento actual</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f"<div style='background:white;border-left:6px solid #27ae60;padding:16px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,0.05);'>"
                    f"<div style='font-size:12px;color:#7f8c8d;font-weight:bold;text-transform:uppercase;'>ESTADO DE META</div>"
                    f"<div style='font-size:32px;font-weight:900;color:#27ae60;margin-top:4px;'>✅ En meta</div>"
                    f"<div style='font-size:12px;color:#95a5a6;margin-top:4px;'>Ya superas el 80% — mantén el ritmo</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

        st.divider()
        st.markdown("#### 📊 Tendencia Reciente (últimos días del periodo)")
        _df_mini = df_dias.tail(10).copy()
        _df_mini["%_V"] = _df_mini["% REAL"].clip(upper=100)
        _fig_mini = go.Figure(
            go.Bar(
                x=_df_mini["FECHA_DT"].dt.strftime("%d %b"),
                y=_df_mini["%_V"],
                marker_color=[
                    "#2ecc71" if x >= 80 else "#f39c12" if x >= 70 else "#e74c3c"
                    for x in _df_mini["% REAL"]
                ],
                text=[f"{x:.0f}%" for x in _df_mini["% REAL"]],
                textposition="auto",
                textfont=dict(size=14, color="white", weight="bold"),
            )
        )
        _fig_mini.add_hline(
            y=80,
            line_dash="dash",
            line_color="#27ae60",
            line_width=2,
            annotation_text="🎯 META 80%",
            annotation_font=dict(color="#27ae60", size=12),
        )
        _fig_mini.update_layout(
            height=260,
            showlegend=False,
            yaxis=dict(range=[0, 110], showgrid=False, title="Cumplimiento (%)"),
            xaxis=dict(showgrid=False, tickangle=0),
            margin=dict(t=10, b=10, l=10, r=10),
        )
        st.plotly_chart(_fig_mini, use_container_width=True)

        st.divider()
        st.markdown("#### 🛑 Motivos de Paro del Periodo")
        if not df_paros.empty:
            _cp_e = df_paros["MOTIVO"].value_counts().reset_index()
            _cp_e.columns = ["MOTIVO", "FRECUENCIA"]
            _fig_pe = go.Figure(
                go.Pie(
                    labels=_cp_e["MOTIVO"],
                    values=_cp_e["FRECUENCIA"],
                    hole=0.38,
                    marker=dict(
                        colors=[
                            "#2c3e50",
                            "#8B1A1A",
                            "#117A65",
                            "#6E2F8B",
                            "#A04000",
                            "#1A5276",
                        ]
                    ),
                    textfont=dict(size=14, color="white"),
                )
            )
            _fig_pe.update_traces(
                textinfo="percent+label",
                textposition="inside",
                insidetextorientation="radial",
            )
            _fig_pe.update_layout(
                height=420,
                showlegend=True,
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=-0.25,
                    xanchor="center",
                    x=0.5,
                    font=dict(size=13),
                ),
                margin=dict(t=20, b=20, l=20, r=20),
            )
            st.plotly_chart(_fig_pe, use_container_width=True)
        else:
            st.success("🎉 ¡Cero paros registrados en el periodo!")

    # ── TAB 2: ANÁLISIS DEL PERIODO ──────────────────────────────────────────
    with tab_analisis:
        with st.expander("📖 ¿Cómo leer este análisis?"):
            st.markdown("""
**Vista operativa — para líderes de área y coordinación de producción.**

**Tendencia diaria:** Avance real vs. lo programado cada día. Un día verde (≥ 80%) significa que la planta produjo suficiente para cumplir el plan.

**Cumplimiento por Área (actual vs. anterior):** Barras de color = periodo actual. Barras grises = periodo anterior equivalente. Si la barra gris supera a la de color, el área retrocedió.

**Ranking de avance:**
- 🏆 *Equipos estrella* — piezas y operaciones que más avanzaron. Reconoce y replica sus condiciones.
- 🚨 *Focos rojos* — piezas y operaciones con mayor atraso. Son los puntos de intervención prioritaria.

**Cumplimiento por día de la semana:** Promedio de todos los días del periodo agrupado por día. Identifica si hay un patrón estructural (p. ej. los lunes siempre arrancan mal).

**Mapa Pieza × Área:** Muestra el % de avance promedio para cada combinación específica de pieza y área. Permite ubicar exactamente *qué pieza* está fallando *en qué área*, no solo en general.
- 🟢 ≥ 80% — en meta
- 🟡 70–79% — zona de riesgo
- 🔴 < 70% — crítico
- ⬜ Sin dato — esa pieza no fue programada en esa área durante el periodo
            """)
        col_izq, col_der = st.columns(2)
        with col_izq:
            st.subheader("📈 Tendencia Diaria del periodo")
            df_dias["%_VISUAL"] = df_dias["% REAL"].clip(upper=100)
            posiciones_texto = [
                "bottom center" if val >= 92 else "top center"
                for val in df_dias["%_VISUAL"]
            ]
            fig_t = go.Figure()
            fig_t.add_hrect(
                y0=0,
                y1=70,
                fillcolor="#e74c3c",
                opacity=0.1,
                line_width=0,
                annotation_text="ZONA CRÍTICA",
                annotation_position="bottom right",
                annotation_font_color="#c0392b",
            )
            fig_t.add_hrect(
                y0=70, y1=80, fillcolor="#f39c12", opacity=0.15, line_width=0
            )
            fig_t.add_hrect(
                y0=80,
                y1=115,
                fillcolor="#2ecc71",
                opacity=0.1,
                line_width=0,
                annotation_text="ZONA SEGURA",
                annotation_position="top right",
                annotation_font_color="#27ae60",
            )
            fig_t.add_trace(
                go.Scatter(
                    x=df_dias["FECHA_DT"],
                    y=df_dias["%_VISUAL"],
                    mode="lines+markers+text",
                    text=[f"{x:.0f}%" for x in df_dias["% REAL"]],
                    textposition=posiciones_texto,
                    textfont=dict(size=13, color="black", family="Arial Black"),
                    line=dict(color="#2c3e50", width=3),
                    marker=dict(
                        size=12,
                        color=[
                            (
                                "#2ecc71"
                                if x >= 80
                                else "#f39c12" if x >= 70 else "#e74c3c"
                            )
                            for x in df_dias["% REAL"]
                        ],
                        line=dict(width=2, color="white"),
                    ),
                )
            )
            fig_t.add_hline(
                y=80,
                line_dash="dash",
                line_color="#27ae60",
                line_width=3,
                annotation_text="🎯 META (80%)",
                annotation_position="top left",
                annotation_font=dict(size=13, color="#27ae60", weight="bold"),
            )
            fig_t.update_layout(
                height=380,
                yaxis=dict(range=[0, 115], title="Cumplimiento (%)", showgrid=False),
                xaxis=dict(showgrid=False),
                margin=dict(t=30, b=10, l=10, r=10),
            )
            st.plotly_chart(fig_t, use_container_width=True)
        with col_der:
            st.subheader("🏢 Cumplimiento por Área — actual vs. anterior")
            df_area = df_uni.groupby(col_prog["area"])["% REAL"].mean().reset_index()
            df_area["%_VISUAL"] = df_area["% REAL"].clip(upper=100)
            fig_a = go.Figure()
            fig_a.add_hrect(
                y0=0, y1=70, fillcolor="#e74c3c", opacity=0.07, line_width=0
            )
            fig_a.add_hrect(
                y0=70, y1=80, fillcolor="#f39c12", opacity=0.10, line_width=0
            )
            fig_a.add_hrect(
                y0=80, y1=105, fillcolor="#2ecc71", opacity=0.07, line_width=0
            )
            fig_a.add_trace(
                go.Bar(
                    name="Periodo actual",
                    x=df_area[col_prog["area"]],
                    y=df_area["%_VISUAL"],
                    marker_color=[
                        "#2ecc71" if x >= 80 else "#f39c12" if x >= 70 else "#e74c3c"
                        for x in df_area["% REAL"]
                    ],
                    text=[f"{x:.1f}%" for x in df_area["% REAL"]],
                    textposition="auto",
                    textfont=dict(size=13, color="white", weight="bold"),
                )
            )
            if not df_ant.empty:
                _df_area_ant = (
                    df_ant.groupby(col_prog["area"])["% REAL"].mean().reset_index()
                )
                _df_area_ant["%_VISUAL"] = _df_area_ant["% REAL"].clip(upper=100)
                fig_a.add_trace(
                    go.Bar(
                        name="Periodo anterior",
                        x=_df_area_ant[col_prog["area"]],
                        y=_df_area_ant["%_VISUAL"],
                        marker_color="rgba(127,140,141,0.55)",
                        text=[f"{x:.1f}%" for x in _df_area_ant["% REAL"]],
                        textposition="auto",
                        textfont=dict(size=12, color="white"),
                    )
                )
            fig_a.add_hline(
                y=80,
                line_dash="dash",
                line_color="#27ae60",
                line_width=2,
                annotation_text="🎯 80%",
                annotation_font=dict(size=12, color="#27ae60", weight="bold"),
            )
            fig_a.update_layout(
                barmode="group",
                yaxis=dict(range=[0, 105], title="Cumplimiento (%)", showgrid=False),
                legend=dict(
                    orientation="h", yanchor="bottom", y=1.02, font=dict(size=11)
                ),
                margin=dict(t=40, b=10, l=10, r=10),
            )
            st.plotly_chart(fig_a, use_container_width=True)
        st.divider()
        st.markdown("### 🥇 Ranking de Avance del Programa de Producción")
        df_p_rank = df_pt_only.groupby(col_prog["pieza"])["% REAL"].mean().reset_index()
        df_s_rank = df_uni.groupby(col_bdd["subproceso"])["% REAL"].mean().reset_index()
        cfg_progreso = st.column_config.ProgressColumn(
            "NIVEL DE AVANCE",
            help="Barra de cumplimiento",
            format="%.1f%%",
            min_value=0,
            max_value=100,
        )
        c_best, c_worst = st.columns(2)
        with c_best:
            st.success("🏆 EQUIPOS ESTRELLA (Aplaude a estos procesos)")
            st.write("📦 **TOP 5: Piezas con mayor avance:**")
            st.dataframe(
                df_p_rank.nlargest(5, "% REAL"),
                column_config={col_prog["pieza"]: "PIEZA", "% REAL": cfg_progreso},
                hide_index=True,
                use_container_width=True,
            )
            st.write("⭐ **TOP 5: Operaciones con mejor avance:**")
            st.dataframe(
                df_s_rank.nlargest(5, "% REAL"),
                column_config={
                    col_bdd["subproceso"]: "OPERACIÓN",
                    "% REAL": cfg_progreso,
                },
                hide_index=True,
                use_container_width=True,
            )
        with c_worst:
            st.error("🚨 FOCOS ROJOS (Aquí hay que meter las manos)")
            st.write("⚠️ **ALERTA: Piezas con mayor atraso:**")
            st.dataframe(
                df_p_rank.nsmallest(5, "% REAL"),
                column_config={col_prog["pieza"]: "PIEZA", "% REAL": cfg_progreso},
                hide_index=True,
                use_container_width=True,
            )
            st.write("🛑 **ALERTA: Operaciones con mayor rezago:**")
            st.dataframe(
                df_s_rank.nsmallest(5, "% REAL"),
                column_config={
                    col_bdd["subproceso"]: "OPERACIÓN",
                    "% REAL": cfg_progreso,
                },
                hide_index=True,
                use_container_width=True,
            )
        _buf_rank = io.BytesIO()
        with pd.ExcelWriter(_buf_rank, engine="openpyxl") as _wr_rank:
            df_p_rank.sort_values("% REAL", ascending=False).rename(
                columns={col_prog["pieza"]: "PIEZA", "% REAL": "AVANCE (%)"}
            ).to_excel(_wr_rank, sheet_name="RANKING PIEZAS", index=False)
            df_s_rank.sort_values("% REAL", ascending=False).rename(
                columns={col_bdd["subproceso"]: "OPERACIÓN", "% REAL": "AVANCE (%)"}
            ).to_excel(_wr_rank, sheet_name="RANKING OPERACIONES", index=False)
        st.download_button(
            label="📥 Descargar ranking completo en Excel",
            data=_buf_rank.getvalue(),
            file_name=f"Ranking_{f_ini.strftime('%Y%m%d')}_{f_fin.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_ranking",
        )
        st.divider()
        st.subheader("📅 Cumplimiento por Día de la Semana")
        st.caption("Comparativo del periodo contra la meta del 80%")
        df_dias["DIA"] = df_dias["FECHA_DT"].dt.dayofweek.map(DIAS_ES)
        ord_d = pd.CategoricalDtype(categories=SEMANA_NSG, ordered=True)
        df_dias["DIA"] = df_dias["DIA"].astype(ord_d)
        san_lunes = df_dias.groupby("DIA")["% REAL"].mean().dropna().reset_index()
        san_lunes["%_VISUAL"] = san_lunes["% REAL"].clip(upper=100)
        fig_san = go.Figure()
        fig_san.add_hrect(y0=0, y1=70, fillcolor="#e74c3c", opacity=0.1, line_width=0)
        fig_san.add_hrect(y0=70, y1=80, fillcolor="#f39c12", opacity=0.15, line_width=0)
        fig_san.add_hrect(y0=80, y1=100, fillcolor="#2ecc71", opacity=0.1, line_width=0)
        fig_san.add_trace(
            go.Bar(
                x=san_lunes["DIA"],
                y=san_lunes["%_VISUAL"],
                marker_color=[
                    "#2ecc71" if x >= 80 else "#f39c12" if x >= 70 else "#e74c3c"
                    for x in san_lunes["% REAL"]
                ],
                text=[f"{x:.0f}%" for x in san_lunes["% REAL"]],
                textposition="auto",
                textfont=dict(size=14, color="white", weight="bold"),
            )
        )
        fig_san.add_hline(
            y=80,
            line_dash="dash",
            line_color="#27ae60",
            line_width=3,
            annotation_text="🎯 META (80%)",
            annotation_font=dict(size=13, color="#27ae60", weight="bold"),
        )
        fig_san.update_layout(
            height=320,
            yaxis=dict(range=[0, 105], title="Cumplimiento (%)", showgrid=False),
            margin=dict(t=30, b=10, l=10, r=10),
        )
        st.plotly_chart(fig_san, use_container_width=True)

        st.divider()
        st.subheader("🗺️ Mapa de Cumplimiento Pieza × Área")
        st.caption(
            "% de avance promedio por cada combinación de Pieza y Área en el periodo."
        )
        _heat_grp = (
            df_uni.groupby([col_prog["pieza"], col_prog["area"]])["% REAL"]
            .mean()
            .reset_index()
        )
        _heat_piv = _heat_grp.pivot(
            index=col_prog["pieza"], columns=col_prog["area"], values="% REAL"
        )
        if not _heat_piv.empty:
            _hz = _heat_piv.values
            _ht = [
                [
                    (
                        f"{_heat_piv.iloc[_ri, _ci]:.0f}%"
                        if not pd.isna(_heat_piv.iloc[_ri, _ci])
                        else "—"
                    )
                    for _ci in range(len(_heat_piv.columns))
                ]
                for _ri in range(len(_heat_piv.index))
            ]
            _fig_heat = go.Figure(
                go.Heatmap(
                    z=_hz,
                    x=_heat_piv.columns.tolist(),
                    y=_heat_piv.index.tolist(),
                    colorscale=[
                        [0.0, "#e74c3c"],
                        [0.7, "#e74c3c"],
                        [0.7, "#f39c12"],
                        [0.8, "#f39c12"],
                        [0.8, "#27ae60"],
                        [1.0, "#27ae60"],
                    ],
                    zmin=0,
                    zmax=100,
                    text=_ht,
                    texttemplate="%{text}",
                    textfont=dict(size=12, color="white"),
                    hovertemplate="<b>%{y}</b> · %{x}: %{z:.1f}%<extra></extra>",
                    colorbar=dict(
                        title="Avance %",
                        tickvals=[0, 70, 80, 100],
                        ticktext=["0%", "70%", "80% meta", "100%"],
                    ),
                )
            )
            _fig_heat.update_layout(
                height=max(300, len(_heat_piv) * 30 + 100),
                margin=dict(t=20, b=20, l=10, r=10),
                xaxis=dict(side="top"),
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(_fig_heat, use_container_width=True)
        else:
            st.info("No hay suficientes combinaciones Pieza+Área para mostrar el mapa.")

    # ── TAB 3: DETALLE DE FALLAS ─────────────────────────────────────────────
    with tab_fallas:
        with st.expander("📖 ¿Cómo leer este detalle?"):
            st.markdown("""
**Vista de intervención — para líderes de área al cierre del turno o al día siguiente.**

**Reporte de fallas (< 80%):** Cada fila es una combinación Fecha + Pieza + Subproceso que no alcanzó el 80% del programa.
| Columna | Qué significa |
|---------|--------------|
| PROGRAMADO (pzs) | Cantidad que se planeó producir ese día en ese subproceso |
| PRODUCIDO (pzs) | Cantidad que realmente se capturó en auditoría |
| AVANCE % | PRODUCIDO ÷ PROGRAMADO × 100 |
| NOTAS (MOTIVO) | Causa registrada por el auditor en bitácora |

**Paros por Área:** Frecuencia de cada motivo de paro desglosada por área. Permite saber si un problema es generalizado o focalizado en una sola área.

**Análisis de Ritmo (Takt Time):** Divide el turno en bloques horarios y mide qué % del programa se avanzó en cada bloque. El 26.6% es la referencia si el turno tiene 4 cortes. Un bloque muy bajo indica dónde ocurrió la pérdida de tiempo real.
            """)
        st.subheader("📝 Reporte Detallado de Fallas (< 80%)")
        df_desv = df_uni[df_uni["% REAL"] < 80].copy()
        if not df_desv.empty:
            df_a_notes = df_auditorias.copy()
            c_f_a = encontrar_columna(df_a_notes, ["FECHA"])
            c_p_a = encontrar_columna(df_a_notes, ["PIEZA"])
            c_s_a = encontrar_columna(df_a_notes, ["SUBPROCESO", "SUB PRO CESO"])
            c_n_a = encontrar_columna(df_a_notes, ["NOTAS", "NOTA"])
            df_bit = pd.merge(
                df_desv,
                df_a_notes[[c_f_a, c_p_a, c_s_a, c_n_a]],
                left_on=[col_prog["fecha"], col_prog["pieza"], col_bdd["subproceso"]],
                right_on=[c_f_a, c_p_a, c_s_a],
                how="left",
            ).drop_duplicates(
                subset=[col_prog["fecha"], col_prog["pieza"], col_bdd["subproceso"]]
            )
            df_bit["_PROG"] = (
                convertir_serie_numerica(df_bit[col_prog["total"]])
                .fillna(0)
                .astype(int)
            )
            df_bit["_REAL"] = (
                convertir_serie_numerica(df_bit[col_aud["real"]]).fillna(0).astype(int)
            )
            df_bit_show = df_bit[
                [
                    col_prog["fecha"],
                    col_prog["area"],
                    col_prog["pieza"],
                    col_bdd["subproceso"],
                    "_PROG",
                    "_REAL",
                    "% REAL",
                    c_n_a,
                ]
            ].copy()
            df_bit_show.columns = [
                "FECHA",
                "ÁREA",
                "PIEZA",
                "SUBPROCESO",
                "PROGRAMADO",
                "PRODUCIDO",
                "AVANCE %",
                "NOTAS (MOTIVO)",
            ]
            st.dataframe(
                df_bit_show.sort_values("FECHA", ascending=False),
                column_config={
                    "AVANCE %": st.column_config.ProgressColumn(
                        "AVANCE (%)", format="%.1f%%", min_value=0, max_value=100
                    ),
                    "PROGRAMADO": st.column_config.NumberColumn(
                        "PROGRAMADO (pzs)", format="%d"
                    ),
                    "PRODUCIDO": st.column_config.NumberColumn(
                        "PRODUCIDO (pzs)", format="%d"
                    ),
                },
                use_container_width=True,
                hide_index=True,
            )
            _buf_fallas = io.BytesIO()
            with pd.ExcelWriter(_buf_fallas, engine="openpyxl") as _wr_fallas:
                df_bit_show.sort_values("FECHA", ascending=False).to_excel(
                    _wr_fallas, sheet_name="FALLAS", index=False
                )
            st.download_button(
                label="📥 Descargar reporte de fallas en Excel",
                data=_buf_fallas.getvalue(),
                file_name=f"Fallas_{f_ini.strftime('%Y%m%d')}_{f_fin.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_fallas",
            )
        else:
            st.success(
                "¡Excelente! No hay desviaciones menores al 80% reportadas en este periodo."
            )

        if not df_paros.empty:
            st.divider()
            st.subheader("🛑 Paros por Área")
            st.caption(
                "Frecuencia de cada motivo de paro desglosada por área de producción."
            )
            _pieza_area_map = (
                df_uni[[col_prog["pieza"], col_prog["area"]]]
                .assign(
                    _PZ=df_uni[col_prog["pieza"]].astype(str).str.strip().str.upper()
                )
                .groupby("_PZ")[col_prog["area"]]
                .first()
            )
            _paros_x_area = df_paros.copy()
            _paros_x_area["_AREA"] = (
                _paros_x_area[col_aud["pieza"]]
                .astype(str)
                .str.strip()
                .str.upper()
                .map(_pieza_area_map)
                .fillna("Sin área")
            )
            _par_grp = (
                _paros_x_area.groupby(["_AREA", "MOTIVO"])
                .size()
                .reset_index(name="FRECUENCIA")
                .sort_values("FRECUENCIA", ascending=False)
            )
            _par_grp.columns = ["ÁREA", "MOTIVO", "FRECUENCIA"]
            _pal = [
                "#2c3e50",
                "#8B1A1A",
                "#117A65",
                "#6E2F8B",
                "#A04000",
                "#1A5276",
                "#5D6D7E",
            ]
            _fig_par = go.Figure()
            for _idx, _ar in enumerate(_par_grp["ÁREA"].unique().tolist()):
                _sub = _par_grp[_par_grp["ÁREA"] == _ar]
                _fig_par.add_trace(
                    go.Bar(
                        name=_ar,
                        x=_sub["MOTIVO"],
                        y=_sub["FRECUENCIA"],
                        marker_color=_pal[_idx % len(_pal)],
                        text=_sub["FRECUENCIA"],
                        textposition="auto",
                    )
                )
            _fig_par.update_layout(
                barmode="group",
                height=320,
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
                yaxis=dict(title="Frecuencia", showgrid=False),
                xaxis=dict(title="Motivo de Paro"),
                margin=dict(t=50, b=10, l=10, r=10),
            )
            st.plotly_chart(_fig_par, use_container_width=True)

        st.divider()
        st.markdown("### ⏱️ Análisis de Ritmo de Producción (Aportación al Turno)")
        st.info(
            "💡 **Tip de Planta:** Si un bloque horario está muy por debajo del 26.6%, revisa en la "
            "bitácora si hubo falta de material, falla de máquinas o cambios de modelo específicos en esas horas."
        )
        try:
            c_ritmo_global, c_ritmo_area = st.columns(2)
            with c_ritmo_global:
                st.selectbox(
                    "Vista global de la planta:",
                    ["TODAS LAS ÁREAS"],
                    disabled=True,
                    key="sel_dummy",
                )
                render_ritmo_por_bloque(
                    df_auditorias,
                    df_programa,
                    df_bdd,
                    col_prog,
                    col_bdd,
                    f_ini,
                    f_fin,
                    "TODAS",
                    "izq",
                )
            with c_ritmo_area:
                lista_areas_dash = obtener_lista_areas(df_programa, col_prog)
                area_ritmo = st.selectbox(
                    "Selecciona un área específica para ver su detalle:",
                    lista_areas_dash,
                    key="sel_ritmo_area_dashboard",
                )
                render_ritmo_por_bloque(
                    df_auditorias,
                    df_programa,
                    df_bdd,
                    col_prog,
                    col_bdd,
                    f_ini,
                    f_fin,
                    area_ritmo,
                    "der",
                )
        except Exception as _e_ritmo:
            st.warning(f"El análisis de ritmo no pudo calcularse: {_e_ritmo}")


def render_estadistica_rango(df_auditorias, df_programa, df_bdd, col_prog, col_bdd):
    st.divider()
    st.markdown("### 📊 DESEMPEÑO POR RANGO DE FECHAS (CUMPLIMIENTO REAL)")
    _hoy_stat = ahora_local().date()
    c_r1, c_r2 = st.columns(2)
    with c_r1:
        f_ini_stat = st.date_input("Desde:", _hoy_stat.replace(day=1), key="vfinal_ini")
    with c_r2:
        f_fin_stat = st.date_input("Hasta:", _hoy_stat, key="vfinal_fin")
    df_uni, col_aud = obtener_datos_unificados(
        df_auditorias, df_programa, df_bdd, col_prog, col_bdd, f_ini_stat, f_fin_stat
    )
    if df_uni.empty:
        st.warning(
            "Faltan datos en las hojas para calcular la estadística o no hay programación."
        )
        return
    res_final = df_uni.groupby(col_prog["area"])["% REAL"].mean().reset_index()
    for _, row in res_final.iterrows():
        area_n = row[col_prog["area"]]
        val_n = round(row["% REAL"], 1)
        color_n = obtener_color_nsg(val_n)
        ancho_barra = min(val_n, 100)
        st.markdown(
            f"""
            <div style='background:white; padding:18px; border-radius:15px; border-left:8px solid {color_n}; margin-bottom:15px; box-shadow: 0 4px 10px rgba(0,0,0,0.06);'>
                <div style='display:flex; justify-content:space-between; align-items:center; margin-bottom: 8px;'>
                    <span style='font-weight:bold; font-size:16px;'>{area_n}</span>
                    <span style='color:{color_n}; font-weight:800; font-size:24px;'>{val_n}%</span>
                </div>
                <div style='background:#e9ecef; height:14px; border-radius:10px; overflow:hidden;'>
                    <div style='background:{color_n}; width:{ancho_barra}%; height:100%; border-radius:10px;'></div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_ritmo_por_bloque(
    df_auditorias,
    df_programa,
    df_bdd,
    col_prog,
    col_bdd,
    f_ini,
    f_fin,
    area_sel,
    sufijo="1",
):
    es_global = area_sel == "TODAS"
    titulo_str = "🏭 PLANTA COMPLETA" if es_global else f"ÁREA: {area_sel}"
    st.caption(f"Aportación a la meta del día (**{titulo_str}**)")

    # 1. Base programada (Programa × BDD)
    df_p = df_programa.copy()
    df_p["FECHA_DT"] = pd.to_datetime(
        df_p[col_prog["fecha"]], format="%d/%m/%Y", errors="coerce"
    )
    df_p = df_p[
        (df_p["FECHA_DT"].dt.date >= f_ini) & (df_p["FECHA_DT"].dt.date <= f_fin)
    ]
    if not es_global:
        df_p = df_p[df_p[col_prog["area"]] == area_sel]
    if df_p.empty:
        st.info("No hay programación para estas fechas.")
        return

    mask_m = df_p[col_prog["area"]].str.upper() == "MOLDEO"
    df_p_m = df_p[
        mask_m
        & df_p[col_prog["pieza"]].str.contains(
            "GENERAL|VACIADO|ADOBES", case=False, na=False
        )
    ]
    df_p_o = df_p[~mask_m]
    df_p_final = pd.concat([df_p_m, df_p_o])
    df_p_final[col_prog["total"]] = convertir_serie_numerica(
        df_p_final[col_prog["total"]]
    ).fillna(0)

    df_base = pd.merge(
        df_p_final[
            [
                col_prog["fecha"],
                col_prog["area"],
                col_prog["pieza"],
                col_prog["total"],
                "FECHA_DT",
            ]
        ],
        df_bdd[[col_bdd["pieza"], col_bdd["subproceso"], col_bdd["proceso"]]],
        left_on=col_prog["pieza"],
        right_on=col_bdd["pieza"],
        how="inner",
    )
    df_base = df_base[df_base[col_bdd["proceso"]] == df_base[col_prog["area"]]]
    if not es_global:
        df_base = df_base[df_base[col_bdd["proceso"]] == area_sel]
    if df_base.empty:
        st.info("No se pudieron cruzar las piezas programadas.")
        return

    prog_por_dia = df_base.groupby("FECHA_DT")[col_prog["total"]].sum().to_dict()

    # 2. Auditorías por corte
    col_aud = {
        "fecha": encontrar_columna(df_auditorias, ["FECHA"]),
        "pieza": encontrar_columna(df_auditorias, ["PIEZA"]),
        "subproceso": encontrar_columna(
            df_auditorias,
            ["SUBPROCESO", "SUB PROCESO", "SUB_PROCESO"],
            contiene_todos=["SUB", "CESO"],
        ),
        "real": encontrar_columna(df_auditorias, ["REAL"]),
        "corte": encontrar_columna(df_auditorias, ["CORTE"]),
        "area": encontrar_columna(df_auditorias, ["AREA", "ÁREA"]),
    }
    cols_req = [k for k, v in col_aud.items() if v is None]
    if cols_req:
        st.info(
            f"Análisis de ritmo no disponible — columnas faltantes en AUDITAR: {cols_req}"
        )
        return

    df_a = df_auditorias.copy()
    df_a["FECHA_DT"] = pd.to_datetime(
        df_a[col_aud["fecha"]], format="%d/%m/%Y", errors="coerce"
    )
    df_a = df_a[
        (df_a["FECHA_DT"].dt.date >= f_ini) & (df_a["FECHA_DT"].dt.date <= f_fin)
    ]
    if not es_global:
        df_a = df_a[df_a[col_aud["area"]] == area_sel]
    df_a[col_aud["real"]] = pd.to_numeric(
        df_a[col_aud["real"]], errors="coerce"
    ).fillna(0)

    def calcular_real_hasta_corte(df_aud, cortes_permitidos):
        df_corte = df_aud[df_aud[col_aud["corte"]].isin(cortes_permitidos)]
        if df_corte.empty:
            return {}
        df_max = (
            df_corte.groupby(
                ["FECHA_DT", col_aud["area"], col_aud["pieza"], col_aud["subproceso"]]
            )[col_aud["real"]]
            .max()
            .reset_index()
        )
        df_cruzado = pd.merge(
            df_base,
            df_max,
            left_on=[
                "FECHA_DT",
                col_prog["area"],
                col_prog["pieza"],
                col_bdd["subproceso"],
            ],
            right_on=[
                "FECHA_DT",
                col_aud["area"],
                col_aud["pieza"],
                col_aud["subproceso"],
            ],
            how="left",
        ).fillna({col_aud["real"]: 0})
        return df_cruzado.groupby("FECHA_DT")[col_aud["real"]].sum().to_dict()

    real_11 = calcular_real_hasta_corte(df_a, ["11:00 AM (3h)"])
    real_14 = calcular_real_hasta_corte(df_a, ["11:00 AM (3h)", "14:00 PM (6h)"])
    real_17 = calcular_real_hasta_corte(
        df_a, ["11:00 AM (3h)", "14:00 PM (6h)", "17:00 PM (9h)"]
    )

    # 3. Aportaciones por bloque
    bloques = []
    for fecha, total_prog in prog_por_dia.items():
        if total_prog == 0:
            continue
        r11 = real_11.get(fecha, 0)
        r14 = real_14.get(fecha, 0)
        r17 = real_17.get(fecha, 0)
        bloques.append(
            {"Bloque": "08:00 - 11:00", "Aportacion": (r11 / total_prog) * 100}
        )
        bloques.append(
            {
                "Bloque": "11:00 - 14:00",
                "Aportacion": (max(0, r14 - r11) / total_prog) * 100,
            }
        )
        bloques.append(
            {
                "Bloque": "14:00 - 17:00",
                "Aportacion": (max(0, r17 - r14) / total_prog) * 100,
            }
        )

    if not bloques:
        st.info("No hay auditorías registradas.")
        return

    df_bloques = (
        pd.DataFrame(bloques).groupby("Bloque")["Aportacion"].mean().reset_index()
    )

    etiquetas = {
        "08:00 - 11:00": "☕ Arranque<br>(08:00-11:00)",
        "11:00 - 14:00": "🔥 Medio Día<br>(11:00-14:00)",
        "14:00 - 17:00": "🏁 Cierre<br>(14:00-17:00)",
    }
    df_bloques["Bloque_Visual"] = df_bloques["Bloque"].map(etiquetas)

    colores_semaforo = [
        "#2ecc71" if v >= 26.6 else "#f39c12" if v >= 20.0 else "#e74c3c"
        for v in df_bloques["Aportacion"]
    ]

    fig_b = go.Figure()
    fig_b.add_trace(
        go.Bar(
            x=df_bloques["Bloque_Visual"],
            y=df_bloques["Aportacion"],
            marker_color=colores_semaforo,
            text=[f"{x:.1f}%" for x in df_bloques["Aportacion"]],
            textposition="auto",
            textfont=dict(color="white", weight="bold", size=15),
        )
    )
    fig_b.add_hline(
        y=26.6,
        line_dash="dash",
        line_color="#2c3e50",
        line_width=2,
        annotation_text="🎯 CUOTA DEL BLOQUE (26.6%)",
        annotation_font=dict(size=12, color="#2c3e50", weight="bold"),
    )
    fig_b.update_layout(
        height=350,
        margin=dict(t=20, b=20, l=10, r=10),
        yaxis=dict(range=[0, 45], showticklabels=False),
        xaxis=dict(tickfont=dict(size=13, weight="bold")),
    )
    st.plotly_chart(
        fig_b, use_container_width=True, key=f"grafica_ritmo_{area_sel}_{sufijo}"
    )


# ============================================================
# MÓDULO PRODUCTIVIDAD — Acuerdo NSG-RH-AC-002
# ============================================================


def _construir_pivot_rrhh(df_periodo, col_nombre, col_eficiencia):
    if df_periodo.empty:
        return pd.DataFrame(columns=["OPERADOR", "FECHA_DT", "EFICIENCIA_DIARIA"])
    dias_activos = sorted(df_periodo["FECHA_DT"].dt.normalize().unique())
    operadores = sorted(df_periodo[col_nombre].str.upper().str.strip().unique())
    idx = pd.MultiIndex.from_product(
        [operadores, dias_activos], names=["OPERADOR", "FECHA_DT"]
    )
    serie = (
        df_periodo.assign(
            _OP=df_periodo[col_nombre].str.upper().str.strip(),
            _DT=df_periodo["FECHA_DT"].dt.normalize(),
        )
        .groupby(["_OP", "_DT"])[col_eficiencia]
        .mean()
    )
    serie.index.names = ["OPERADOR", "FECHA_DT"]
    resultado = serie.reindex(idx, fill_value=0.0).reset_index()
    resultado.rename(columns={col_eficiencia: "EFICIENCIA_DIARIA"}, inplace=True)
    return resultado


def _semaforo_prod(v):
    if v >= 90.1:
        return "#27ae60", "🟢", "VERDE — Rendimiento Óptimo"
    if v >= 70.0:
        return "#f39c12", "🟡", "AMARILLO — Zona de Mejora"
    return "#e74c3c", "🔴", "ROJO — Requiere Atención"


def render_productividad(df_programa, col_prog, df_bdd, col_bdd):
    st.markdown("## 📈 Dashboard de Productividad del Personal")
    st.caption("Acuerdo NSG-RH-AC-002 — Análisis forense de eficiencia operativa")

    hoy = ahora_local().date()
    # Semana NSG: jueves a miércoles. Calcular el jueves de inicio de la semana en curso.
    _dias_desde_jueves = (hoy.weekday() - 3) % 7  # Thu=0, Fri=1, Mon=4, Tue=5, Wed=6
    semana_nsg_ini = hoy - timedelta(days=_dias_desde_jueves)

    c1, c2 = st.columns(2)
    with c1:
        f_ini = st.date_input("Periodo Inicio:", value=semana_nsg_ini, key="prod_f_ini")
    with c2:
        f_fin = st.date_input("Periodo Fin:", value=hoy, key="prod_f_fin")

    if f_ini > f_fin:
        st.error("La fecha de inicio no puede ser posterior a la fecha fin.")
        return

    n_dias = (f_fin - f_ini).days + 1
    f_ini_ant = f_ini - timedelta(days=n_dias)
    f_fin_ant = f_ini - timedelta(days=1)

    with st.spinner("Cargando datos de RRHH..."):
        df_rrhh = leer_datos_rrhh(version=obtener_version_hoja("REGISTRO"))

    if df_rrhh.empty:
        st.warning(
            f"⚠️ No se pudieron cargar los datos del libro de RRHH (ID: {ID_LIBRO_RRHH}). "
            "Verifica que la cuenta de servicio tenga acceso y que la hoja se llame REGISTRO."
        )
        return

    col_fecha = encontrar_columna(df_rrhh, ["FECHA"])
    col_nombre = encontrar_columna(
        df_rrhh, ["NOMBRE", "OPERADOR", "EMPLEADO", "TRABAJADOR", "COLABORADOR"]
    )
    col_area = encontrar_columna(
        df_rrhh,
        ["AREA", "ÁREA", "DEPARTAMENTO", "DEPTO", "PROCESO", "SECCION", "SECCIÓN"],
    )
    col_actividad = encontrar_columna(
        df_rrhh,
        ["ACTIVIDAD", "TAREA", "DESCRIPCION", "DESCRIPCIÓN", "OPERACION", "OPERACIÓN"],
    )
    col_eficiencia = (
        encontrar_columna(df_rrhh, ["PRODUCTIVIDADR"])
        or encontrar_columna(
            df_rrhh,
            [
                "EFICIENCIA",
                "PRODUCTIVIDAD",
                "RENDIMIENTO",
                "% EFICIENCIA",
                "%EFICIENCIA",
                "PORCENTAJE",
                "% PRODUCTIVIDAD",
                "%PRODUCTIVIDAD",
            ],
        )
    )
    col_pieza = encontrar_columna(
        df_rrhh, ["PIEZA", "MODELO", "PRODUCTO", "REFERENCIA", "REF", "PARTE"]
    )
    col_cantidad = encontrar_columna(
        df_rrhh,
        [
            "CANTIDAD",
            "PIEZAS",
            "REAL",
            "PRODUCCION",
            "PRODUCCIÓN",
            "PZAS",
            "PZS",
            "CANT",
        ],
    )
    col_tiempo = encontrar_columna(
        df_rrhh, ["TIEMPO USADO", "TIEMPOUSADO", "TIEMPO", "DURACION", "DURACIÓN"]
    )
    col_capxh = encontrar_columna(
        df_rrhh, ["CAP PXH", "CAPPXH", "CAP_PXH", "CAPXH", "CAPACIDAD PXH"]
    )
    col_hora_ini = encontrar_columna(
        df_rrhh, ["HORA INICIO", "HORAINICIO", "INICIO", "HORA INI", "ENTRADA"]
    )
    col_hora_fin = encontrar_columna(
        df_rrhh,
        ["HORA FINAL", "HORAFINAL", "FINAL", "HORA FIN", "SALIDA", "HORA SALIDA"],
    )

    faltantes = [
        k
        for k, v in {
            "FECHA": col_fecha,
            "NOMBRE/OPERADOR": col_nombre,
            "EFICIENCIA": col_eficiencia,
        }.items()
        if not v
    ]
    if faltantes:
        st.error(
            f"No se encontraron columnas requeridas en la hoja REGISTRO: **{', '.join(faltantes)}**"
        )
        with st.expander("Columnas detectadas en la hoja"):
            st.write(list(df_rrhh.columns))
        return

    df = df_rrhh.copy()
    df["FECHA_DT"] = pd.to_datetime(df[col_fecha], format="%d/%m/%Y", errors="coerce")
    df = df[df["FECHA_DT"].notna()].copy()

    # Unificación inmediata de área (solo afecta la copia local de RRHH, nunca df_programa)
    if col_area:
        df[col_area] = df[col_area].astype(str).str.upper().str.strip()
        df[col_area] = df[col_area].replace(
            {"MOLDEO": "MOLDEO / CORAZONES", "CORAZONES": "MOLDEO / CORAZONES"}
        )

    lideres_norm = {normalizar_clave(l) for l in LIDERES_EXCLUIDOS}
    df = df[
        ~df[col_nombre].apply(lambda x: normalizar_clave(str(x))).isin(lideres_norm)
    ].copy()
    df["EFIC_NUM"] = convertir_serie_numerica(df[col_eficiencia]).clip(upper=100.0)
    df = df[df["FECHA_DT"].dt.weekday < 5].copy()

    def _filtrar(fi, ff):
        return df[
            (df["FECHA_DT"].dt.date >= fi) & (df["FECHA_DT"].dt.date <= ff)
        ].copy()

    df_periodo = _filtrar(f_ini, f_fin)
    df_anterior = _filtrar(f_ini_ant, f_fin_ant)

    # Historial completo para métricas estructurales (Polivalencia, Proceso vs Persona).
    # Solo operadores activos: al menos una captura en los últimos 30 días.
    _corte_activos = ahora_local().date() - timedelta(days=30)
    _ops_activos = set(
        df[df["FECHA_DT"].dt.date >= _corte_activos][col_nombre]
        .astype(str)
        .str.upper()
        .str.strip()
        .unique()
    ) - {"", "NAN", "NONE", "N/A"}
    df_historico = df[
        df[col_nombre].astype(str).str.upper().str.strip().isin(_ops_activos)
    ].copy()

    # Historial para planeación de turno: incluye líderes (son candidatos de cobertura).
    # Misma lógica pero parte de df_rrhh antes del filtro de líderes excluidos.
    _df_rrhh_full = df_rrhh.copy()
    _df_rrhh_full["FECHA_DT"] = pd.to_datetime(
        _df_rrhh_full[col_fecha], format="%d/%m/%Y", errors="coerce"
    )
    _df_rrhh_full = _df_rrhh_full[_df_rrhh_full["FECHA_DT"].notna()].copy()
    if col_area:
        _df_rrhh_full[col_area] = (
            _df_rrhh_full[col_area].astype(str).str.upper().str.strip()
        )
        _df_rrhh_full[col_area] = _df_rrhh_full[col_area].replace(
            {"MOLDEO": "MOLDEO / CORAZONES", "CORAZONES": "MOLDEO / CORAZONES"}
        )
    _df_rrhh_full["EFIC_NUM"] = convertir_serie_numerica(
        _df_rrhh_full[col_eficiencia]
    ).clip(upper=100.0)
    _df_rrhh_full = _df_rrhh_full[_df_rrhh_full["FECHA_DT"].dt.weekday < 5].copy()
    _ops_activos_full = set(
        _df_rrhh_full[_df_rrhh_full["FECHA_DT"].dt.date >= _corte_activos][col_nombre]
        .astype(str)
        .str.upper()
        .str.strip()
        .unique()
    ) - {"", "NAN", "NONE", "N/A"}
    df_historico_all = _df_rrhh_full[
        _df_rrhh_full[col_nombre]
        .astype(str)
        .str.upper()
        .str.strip()
        .isin(_ops_activos_full)
    ].copy()

    if df_periodo.empty:
        st.info("No hay registros de productividad en el periodo seleccionado.")
        return

    df_pivot = _construir_pivot_rrhh(df_periodo, col_nombre, "EFIC_NUM")

    prod_anterior = None
    if not df_anterior.empty:
        df_pivot_ant = _construir_pivot_rrhh(df_anterior, col_nombre, "EFIC_NUM")
        if not df_pivot_ant.empty:
            prod_anterior = df_pivot_ant["EFICIENCIA_DIARIA"].mean()

    prod_planta = df_pivot["EFICIENCIA_DIARIA"].mean()
    bono_promedio = calcular_bono(prod_planta)
    total_op_dias = len(df_pivot)
    op_dias_presentes = (df_pivot["EFICIENCIA_DIARIA"] > 0).sum()
    asistencia_pct = (
        (op_dias_presentes / total_op_dias * 100) if total_op_dias > 0 else 0.0
    )

    if prod_anterior is not None:
        delta_val = prod_planta - prod_anterior
        delta_flecha = "▲" if delta_val >= 0 else "▼"
        delta_color = "#27ae60" if delta_val >= 0 else "#e74c3c"
        delta_txt = f"{delta_flecha} {abs(delta_val):.1f}% vs periodo anterior"
    else:
        delta_color, delta_txt = "#95a5a6", "Sin datos del periodo anterior"

    (
        tab_vg,
        tab_alertas,
        tab_profundo,
        tab_lideres,
        tab_colaboradores,
        tab_planeacion,
    ) = st.tabs(
        [
            "📊 Vista General",
            "🚨 Panel de Alertas",
            "🔬 Análisis Profundo",
            "👔 Líderes",
            "👥 Colaboradores",
            "📋 Planeación de Turno",
        ]
    )

    with tab_vg:
        with st.expander("📖 Guía de métricas — ¿qué mide cada indicador?"):
            st.markdown("""
| Indicador | Qué mide | Meta NSG |
|-----------|----------|----------|
| 🏭 **Productividad Promedio** | Eficiencia promedio de todos los operadores del área en el periodo seleccionado | ≥ 90.1% |
| 🎁 **Puntos de Bono** | Puntos calculados con el Acuerdo NSG-RH-AC-002 sobre la productividad promedio del área (máximo 40 pts) | 40 pts |
| ✅ **Asistencia** | Días-persona con registro activo ÷ días laborables disponibles × 100 | 100% |

**Semáforo de eficiencia (aplica a todos los gráficos):**

| Color | Rango | Significado |
|-------|-------|-------------|
| 🟢 Verde | ≥ 90.1% | Rendimiento óptimo — califica para bono máximo |
| 🟡 Amarillo | 70–90% | Zona de mejora — bono parcial, se recomienda acompañamiento |
| 🔴 Rojo | < 70% | Zona crítica — bono mínimo o nulo, requiere intervención |

**Cálculo del bono (doble tramo lineal):**
- 0% → 70%: escala de 0 a 15 puntos
- 70% → 100%: escala de 15 a 40 puntos
- ≥ 100%: 40 puntos (tope máximo)

> 💡 La flecha junto al porcentaje muestra la variación vs. el periodo anterior del mismo tamaño.
            """)
        st.markdown(
            f"""
        <div style='display:flex; gap:15px; margin-bottom:25px; flex-wrap:wrap;'>
          <div style='background:white; border-top:5px solid #8B1A1A; padding:20px 25px; border-radius:10px; flex:1; min-width:200px; box-shadow:0 4px 12px rgba(0,0,0,0.07); text-align:center;'>
            <div style='font-size:11px; color:#7f8c8d; font-weight:700; text-transform:uppercase; letter-spacing:1px;'>🏭 PRODUCTIVIDAD PROMEDIO DE PLANTA</div>
            <div style='font-size:40px; color:#8B1A1A; font-weight:900; margin:8px 0;'>{prod_planta:.1f}%</div>
            <div style='font-size:13px; color:{delta_color}; font-weight:700;'>{delta_txt}</div>
          </div>
          <div style='background:white; border-top:5px solid #e67e22; padding:20px 25px; border-radius:10px; flex:1; min-width:200px; box-shadow:0 4px 12px rgba(0,0,0,0.07); text-align:center;'>
            <div style='font-size:11px; color:#7f8c8d; font-weight:700; text-transform:uppercase; letter-spacing:1px;'>🎁 PUNTOS PROMEDIO DE BONO</div>
            <div style='font-size:40px; color:#e67e22; font-weight:900; margin:8px 0;'>{bono_promedio:.1f} pts</div>
            <div style='font-size:13px; color:#95a5a6;'>Máximo 40 pts — Acuerdo NSG-RH-AC-002</div>
          </div>
          <div style='background:white; border-top:5px solid #27ae60; padding:20px 25px; border-radius:10px; flex:1; min-width:200px; box-shadow:0 4px 12px rgba(0,0,0,0.07); text-align:center;'>
            <div style='font-size:11px; color:#7f8c8d; font-weight:700; text-transform:uppercase; letter-spacing:1px;'>✅ ASISTENCIA GENERAL DE PLANTA</div>
            <div style='font-size:40px; color:#27ae60; font-weight:900; margin:8px 0;'>{asistencia_pct:.1f}%</div>
            <div style='font-size:13px; color:#95a5a6;'>Días-persona registrados vs días activos</div>
          </div>
        </div>
        """,
            unsafe_allow_html=True,
        )

        st.divider()

        # Tendencia diaria
        st.subheader("📈 Tendencia Diaria de Productividad — Toda la Planta")
        tend = (
            df_pivot.groupby("FECHA_DT")["EFICIENCIA_DIARIA"]
            .mean()
            .reset_index()
            .sort_values("FECHA_DT")
        )
        tend["_VIS"] = tend["EFICIENCIA_DIARIA"].clip(upper=100)
        pos_txt = ["bottom center" if v >= 93 else "top center" for v in tend["_VIS"]]
        fig_tend = go.Figure()
        fig_tend.add_hrect(
            y0=0,
            y1=70,
            fillcolor="#e74c3c",
            opacity=0.08,
            line_width=0,
            annotation_text="ZONA CRÍTICA",
            annotation_position="bottom right",
            annotation_font_color="#c0392b",
        )
        fig_tend.add_hrect(
            y0=70, y1=90.1, fillcolor="#f39c12", opacity=0.08, line_width=0
        )
        fig_tend.add_hrect(
            y0=90.1,
            y1=115,
            fillcolor="#2ecc71",
            opacity=0.08,
            line_width=0,
            annotation_text="ZONA ÓPTIMA",
            annotation_position="top right",
            annotation_font_color="#27ae60",
        )
        fig_tend.add_trace(
            go.Scatter(
                x=tend["FECHA_DT"],
                y=tend["_VIS"],
                mode="lines+markers+text",
                text=[f"{v:.0f}%" for v in tend["EFICIENCIA_DIARIA"]],
                textposition=pos_txt,
                textfont=dict(size=12, color="black", family="Arial Black"),
                line=dict(color="#2c3e50", width=3),
                marker=dict(
                    size=12,
                    color=[
                        "#27ae60" if v >= 90.1 else "#f39c12" if v >= 70 else "#e74c3c"
                        for v in tend["EFICIENCIA_DIARIA"]
                    ],
                    line=dict(width=2, color="white"),
                ),
            )
        )
        fig_tend.add_hline(
            y=70,
            line_dash="dash",
            line_color="#e74c3c",
            line_width=2,
            annotation_text="MÍNIMO ACEPTABLE (70%)",
            annotation_position="bottom left",
            annotation_font=dict(size=11, color="#e74c3c"),
        )
        fig_tend.add_hline(
            y=90.1,
            line_dash="dash",
            line_color="#27ae60",
            line_width=2,
            annotation_text="🎯 META ÓPTIMA (90.1%)",
            annotation_position="top left",
            annotation_font=dict(size=11, color="#27ae60"),
        )
        fig_tend.update_layout(
            height=390,
            yaxis=dict(range=[0, 115], title="Productividad (%)", showgrid=False),
            xaxis=dict(showgrid=False),
            margin=dict(t=30, b=10, l=10, r=10),
        )
        st.plotly_chart(fig_tend, use_container_width=True)
        st.divider()

        # Reporte por área
        if col_area:
            st.subheader("🏢 Productividad Promedio por Área / Departamento")
            df_area_ref = (
                df_periodo[[col_nombre, col_area]]
                .assign(OPERADOR=df_periodo[col_nombre].str.upper().str.strip())
                .drop_duplicates(subset=["OPERADOR"])
            )
            df_con_area = pd.merge(
                df_pivot, df_area_ref[["OPERADOR", col_area]], on="OPERADOR", how="left"
            )
            prod_area = (
                df_con_area.groupby(col_area)["EFICIENCIA_DIARIA"]
                .mean()
                .reset_index()
                .sort_values("EFICIENCIA_DIARIA", ascending=True)
            )
            prod_area.columns = ["ÁREA", "PRODUCTIVIDAD"]
            fig_area = go.Figure()
            fig_area.add_vline(
                x=70,
                line_dash="dash",
                line_color="#e74c3c",
                line_width=2,
                annotation_text="Mínimo (70%)",
                annotation_font=dict(color="#e74c3c", size=11),
            )
            fig_area.add_vline(
                x=90.1,
                line_dash="dash",
                line_color="#27ae60",
                line_width=2,
                annotation_text="Meta (90.1%)",
                annotation_font=dict(color="#27ae60", size=11),
            )
            fig_area.add_trace(
                go.Bar(
                    x=prod_area["PRODUCTIVIDAD"],
                    y=prod_area["ÁREA"],
                    orientation="h",
                    marker_color=[
                        "#27ae60" if v >= 90.1 else "#f39c12" if v >= 70 else "#e74c3c"
                        for v in prod_area["PRODUCTIVIDAD"]
                    ],
                    text=[f"{v:.1f}%" for v in prod_area["PRODUCTIVIDAD"]],
                    textposition="auto",
                    textfont=dict(size=13, color="white", weight="bold"),
                )
            )
            fig_area.update_layout(
                height=max(260, len(prod_area) * 60),
                xaxis=dict(range=[0, 110], title="Productividad (%)", showgrid=False),
                yaxis=dict(showgrid=False),
                margin=dict(t=20, b=10, l=10, r=10),
            )
            st.plotly_chart(fig_area, use_container_width=True)
            st.divider()

        # Top 5 por actividad y pieza
        if col_actividad or col_pieza:
            st.markdown("### 📊 Análisis de Rendimiento por Actividad y Pieza")
            _cfg_top = st.column_config.ProgressColumn(
                "PRODUCTIVIDAD (%)", format="%.1f%%", min_value=0, max_value=100
            )
            if col_actividad and not df_periodo.empty:
                prod_act = (
                    df_periodo.groupby(col_actividad)["EFIC_NUM"].mean().reset_index()
                )
                prod_act.columns = ["ACTIVIDAD", "PRODUCTIVIDAD"]
                st.markdown("#### Por Actividad")
                _ca1, _ca2 = st.columns(2)
                with _ca1:
                    st.success("⭐ TOP 5 — Mayor Rendimiento")
                    st.dataframe(
                        prod_act.nlargest(5, "PRODUCTIVIDAD"),
                        column_config={"PRODUCTIVIDAD": _cfg_top},
                        hide_index=True,
                        use_container_width=True,
                    )
                with _ca2:
                    st.error("🚨 TOP 5 — Mayor Rezago")
                    st.dataframe(
                        prod_act.nsmallest(5, "PRODUCTIVIDAD"),
                        column_config={"PRODUCTIVIDAD": _cfg_top},
                        hide_index=True,
                        use_container_width=True,
                    )
            if col_pieza and not df_periodo.empty:
                prod_pza = (
                    df_periodo.groupby(col_pieza)["EFIC_NUM"].mean().reset_index()
                )
                prod_pza.columns = ["PIEZA", "PRODUCTIVIDAD"]
                st.markdown("#### Por Pieza")
                _cp1, _cp2 = st.columns(2)
                with _cp1:
                    st.success("⭐ TOP 5 — Mayor Rendimiento")
                    st.dataframe(
                        prod_pza.nlargest(5, "PRODUCTIVIDAD"),
                        column_config={"PRODUCTIVIDAD": _cfg_top},
                        hide_index=True,
                        use_container_width=True,
                    )
                with _cp2:
                    st.error("🚨 TOP 5 — Mayor Rezago")
                    st.dataframe(
                        prod_pza.nsmallest(5, "PRODUCTIVIDAD"),
                        column_config={"PRODUCTIVIDAD": _cfg_top},
                        hide_index=True,
                        use_container_width=True,
                    )
            st.divider()

        # Rankings operadores
        st.markdown("### 🏅 Ranking de Productividad del Personal")
        prod_op = df_pivot.groupby("OPERADOR")["EFICIENCIA_DIARIA"].mean().reset_index()
        prod_op.columns = ["OPERADOR", "PRODUCTIVIDAD"]
        prod_op["BONO_PTS"] = prod_op["PRODUCTIVIDAD"].apply(calcular_bono)
        cfg_prod_bar = st.column_config.ProgressColumn(
            "PRODUCTIVIDAD", format="%.1f%%", min_value=0, max_value=100
        )
        cfg_bono_num = st.column_config.NumberColumn("BONO (pts)", format="%.1f")
        c_best, c_worst = st.columns(2)
        with c_best:
            st.success("⭐ TOP 5 — EQUIPOS ESTRELLA")
            st.dataframe(
                prod_op.nlargest(5, "PRODUCTIVIDAD")[
                    ["OPERADOR", "PRODUCTIVIDAD", "BONO_PTS"]
                ],
                column_config={
                    "OPERADOR": "NOMBRE",
                    "PRODUCTIVIDAD": cfg_prod_bar,
                    "BONO_PTS": cfg_bono_num,
                },
                hide_index=True,
                use_container_width=True,
            )
        with c_worst:
            st.error("🚨 TOP 5 — FOCOS ROJOS (Mayor Rezago)")
            st.dataframe(
                prod_op.nsmallest(5, "PRODUCTIVIDAD")[
                    ["OPERADOR", "PRODUCTIVIDAD", "BONO_PTS"]
                ],
                column_config={
                    "OPERADOR": "NOMBRE",
                    "PRODUCTIVIDAD": cfg_prod_bar,
                    "BONO_PTS": cfg_bono_num,
                },
                hide_index=True,
                use_container_width=True,
            )
        st.divider()

    with tab_alertas:
        st.markdown("## 🚨 Panel de Alertas Operativas")
        st.caption("Detecta patrones críticos que el promedio oculta.")

        if df_pivot.empty:
            st.info("Sin datos suficientes para calcular alertas.")
        else:
            # ── Coeficiente de Variación ──────────────────────────────
            st.subheader("📉 Operadores con Rendimiento Errático")
            st.caption("CV alto = operador impredecible. El promedio solo no alcanza.")
            with st.expander("📖 ¿Qué mide el CV y cómo leer esta tabla?"):
                st.markdown("""
**El Coeficiente de Variación (CV) mide qué tan inconsistente es el rendimiento día a día.**
Fórmula: `CV = (Desviación Estándar ÷ Promedio) × 100`

**Ejemplo real:**

| Operador | Lun | Mar | Mié | Jue | Vie | Promedio | CV | Situación |
|----------|-----|-----|-----|-----|-----|----------|----|-----------|
| Juan | 88% | 85% | 90% | 87% | 86% | **87%** | **2%** | ✅ Confiable y predecible |
| Pedro | 40% | 95% | 100% | 55% | 95% | **77%** | **30%** | ⚠️ El promedio engaña — es errático |

**Semáforo de estabilidad:**

| Color | CV | Qué significa | Acción recomendada |
|-------|-----|---------------|-------------------|
| 🟢 Consistente | < 15% | Operador predecible — planea con confianza | Mantener |
| 🟡 Inestable | 15–30% | Variaciones notables — investigar qué días falla | Monitorear |
| 🔴 Muy errático | > 30% | El promedio no lo representa — puede ser ausentismo, rotación o desmotivación | Intervenir |

> 💡 Un operador con 77% de promedio y CV 30% es **más riesgoso** para la planta que uno con 80% y CV 5%.
                """)
            _cv = (
                df_pivot.groupby("OPERADOR")["EFICIENCIA_DIARIA"]
                .agg(["mean", "std"])
                .rename(columns={"mean": "media", "std": "desv"})
                .fillna(0)
                .assign(
                    CV=lambda x: (x["desv"] / x["media"].replace(0, 1) * 100).round(1)
                )
                .reset_index()
                .sort_values("CV", ascending=False)
            )
            _cv["ESTABILIDAD"] = _cv["CV"].apply(
                lambda v: (
                    "🔴 Muy errático"
                    if v > 30
                    else ("🟡 Inestable" if v > 15 else "🟢 Consistente")
                )
            )
            st.dataframe(
                _cv[["OPERADOR", "media", "CV", "ESTABILIDAD"]].rename(
                    columns={"media": "PRODUCTIVIDAD PROM."}
                ),
                column_config={
                    "PRODUCTIVIDAD PROM.": st.column_config.ProgressColumn(
                        "PRODUCTIVIDAD PROM.",
                        format="%.1f%%",
                        min_value=0,
                        max_value=100,
                        help="Eficiencia promedio del operador en el periodo seleccionado.",
                    ),
                    "CV": st.column_config.NumberColumn(
                        "VARIACIÓN (CV %)",
                        format="%.1f %%",
                        help="Coeficiente de Variación: qué tan inconsistente es el rendimiento. Menor = más predecible. >30% = comportamiento errático.",
                    ),
                    "ESTABILIDAD": st.column_config.TextColumn(
                        "ESTABILIDAD",
                        help="🟢 Consistente (CV < 15%) | 🟡 Inestable (CV 15–30%) | 🔴 Muy errático (CV > 30%)",
                    ),
                },
                hide_index=True,
                use_container_width=True,
            )
            _erraticos = _cv[_cv["CV"] > 30]
            if not _erraticos.empty:
                st.warning(
                    f"⚠️ {len(_erraticos)} operador(es) con CV > 30%: **{', '.join(_erraticos['OPERADOR'].tolist())}**. Revisa su bitácora en Análisis Profundo."
                )

            st.divider()

            # ── Rachas rojas consecutivas ─────────────────────────────
            st.subheader("🔴 Rachas de Días Consecutivos en Zona Crítica (< 70%)")
            st.caption("Identifica quién requiere intervención inmediata esta semana.")
            with st.expander("📖 ¿Cómo leer esta tabla?"):
                st.markdown("""
**Una racha es una secuencia de días seguidos donde el operador estuvo por debajo del 70%.**
El promedio puede esconder una crisis activa — esta tabla la detecta.

**Guía de columnas:**

| Columna | Qué mide |
|---------|----------|
| **PROD. PROM.** | Eficiencia promedio en el periodo — para contexto |
| **RACHA MÁX.** | El mayor bloque de días consecutivos en zona crítica (< 70%) durante todo el periodo |
| **DÍAS CRÍTICOS RECIENTES** | De los últimos 5 días registrados, cuántos tuvieron eficiencia < 70% |
| **ACCIÓN** | Prioridad recomendada de intervención |

**Semáforo de acción:**

| Prioridad | Condición | Qué hacer |
|-----------|-----------|-----------|
| 🚨 **URGENTE** | 3+ días críticos en los últimos 5 | Conversación con el operador hoy mismo |
| ⚠️ **SEGUIMIENTO** | 1–2 días críticos recientes | Monitorear de cerca esta semana |
| 📋 **HISTORIAL** | Tuvo rachas en el pasado, sin impacto reciente | Tener en cuenta, no urgente |

> 💡 Solo aparecen operadores con **rachas de 2+ días consecutivos**. Sin racha = no aparece en la tabla.
                """)
            _alertas_racha = []
            for _op_r, _df_r in df_pivot.groupby("OPERADOR"):
                _df_r = _df_r.sort_values("FECHA_DT").copy()
                _df_r["_ROJO"] = _df_r["EFICIENCIA_DIARIA"] < 70
                _df_r["_GRP"] = (~_df_r["_ROJO"]).cumsum()
                _rachas = _df_r[_df_r["_ROJO"]].groupby("_GRP").size()
                _racha_max = int(_rachas.max()) if not _rachas.empty else 0
                _racha_activa = int(_df_r.tail(5)["_ROJO"].sum())
                _prom_r = round(_df_r["EFICIENCIA_DIARIA"].mean(), 1)
                if _racha_max >= 2:
                    _alertas_racha.append(
                        {
                            "OPERADOR": _op_r,
                            "PROD. PROM.": _prom_r,
                            "RACHA MÁX. (días)": _racha_max,
                            "DÍAS CRÍTICOS RECIENTES": _racha_activa,
                            "ACCIÓN": (
                                "🚨 URGENTE"
                                if _racha_activa >= 3
                                else (
                                    "⚠️ SEGUIMIENTO"
                                    if _racha_activa >= 1
                                    else "📋 HISTORIAL"
                                )
                            ),
                        }
                    )
            if _alertas_racha:
                _df_racha = pd.DataFrame(_alertas_racha).sort_values(
                    "RACHA MÁX. (días)", ascending=False
                )
                st.dataframe(
                    _df_racha,
                    column_config={
                        "PROD. PROM.": st.column_config.ProgressColumn(
                            "PROD. PROM.",
                            format="%.1f%%",
                            min_value=0,
                            max_value=100,
                            help="Eficiencia promedio del operador en el periodo — solo para contexto.",
                        ),
                        "RACHA MÁX. (días)": st.column_config.NumberColumn(
                            "RACHA MÁX. (días)",
                            help="Mayor número de días CONSECUTIVOS con eficiencia < 70% registrados en el periodo.",
                        ),
                        "DÍAS CRÍTICOS RECIENTES": st.column_config.NumberColumn(
                            "DÍAS CRÍTICOS RECIENTES",
                            help="De los últimos 5 días con registro, cuántos tuvieron eficiencia < 70%.",
                        ),
                        "ACCIÓN": st.column_config.TextColumn(
                            "ACCIÓN",
                            help="🚨 URGENTE: 3+ días críticos recientes | ⚠️ SEGUIMIENTO: 1–2 días | 📋 HISTORIAL: racha pasada sin impacto reciente",
                        ),
                    },
                    hide_index=True,
                    use_container_width=True,
                )
                _urgentes = _df_racha[_df_racha["ACCIÓN"] == "🚨 URGENTE"]
                if not _urgentes.empty:
                    st.error(
                        f"🚨 Requieren atención HOY: **{', '.join(_urgentes['OPERADOR'].tolist())}**"
                    )
            else:
                st.success(
                    "✅ Sin rachas de 2+ días críticos consecutivos en este periodo."
                )

            st.divider()

            # ── Mapa de calor operador × día de semana ────────────────
            st.subheader("🗓️ Mapa de Calor — Productividad por Día de Semana")
            st.caption("¿Quién rinde sistemáticamente menos un día específico?")
            with st.expander("📖 ¿Cómo leer el mapa de calor?"):
                st.markdown("""
**Cada celda = promedio de eficiencia de ese operador en ese día de la semana (todos los lunes del periodo, por ejemplo).**

| Elemento | Qué representa |
|----------|----------------|
| **Filas** | Cada operador del área |
| **Columnas** | Días de la semana (promedio de todos los días de ese tipo en el periodo) |
| **Número en celda** | Eficiencia promedio (%) de ese operador en ese día |

**Colores:**
- 🟢 Verde → ≥ 90.1% — rendimiento óptimo ese día
- 🟡 Amarillo → 70–90% — zona de mejora
- 🔴 Rojo → < 70% — zona crítica

**Patrones que debes buscar:**

| Patrón | Lo que indica | Acción |
|--------|--------------|--------|
| **Columna entera roja** | El DÍA tiene un problema de proceso o turno | Revisar arranque, material, supervisión de ese día |
| **Fila entera roja** | El OPERADOR tiene bajo rendimiento toda la semana | Intervención individual urgente |
| **Celda aislada roja** | Evento puntual de ese operador en ese día | Revisar bitácora de esas fechas específicas |
| **Varios de la misma área en rojo el mismo día** | Problema de proceso, no de persona | Escalar a ingeniería o producción |
                """)
            _heat_df = df_pivot.copy()
            _heat_df["DIA"] = _heat_df["FECHA_DT"].dt.dayofweek.map(DIAS_ES)
            _orden_dias = [d for d in SEMANA_NSG if d in _heat_df["DIA"].unique()]
            _heat_pivot = (
                _heat_df.groupby(["OPERADOR", "DIA"])["EFICIENCIA_DIARIA"]
                .mean()
                .unstack(fill_value=0)
                .reindex(columns=_orden_dias, fill_value=0)
            )
            if not _heat_pivot.empty:
                _fig_heat = go.Figure(
                    go.Heatmap(
                        z=_heat_pivot.values,
                        x=_heat_pivot.columns.tolist(),
                        y=_heat_pivot.index.tolist(),
                        colorscale=[
                            [0.0, "#e74c3c"],
                            [0.35, "#e74c3c"],
                            [0.35, "#f39c12"],
                            [0.45, "#f39c12"],
                            [0.45, "#27ae60"],
                            [1.0, "#27ae60"],
                        ],
                        zmin=0,
                        zmax=100,
                        text=[[f"{v:.0f}%" for v in row] for row in _heat_pivot.values],
                        texttemplate="%{text}",
                        textfont=dict(size=11, color="white"),
                        hovertemplate="<b>%{y}</b><br>%{x}: %{z:.1f}%<extra></extra>",
                    )
                )
                _fig_heat.update_layout(
                    height=max(300, len(_heat_pivot) * 35 + 80),
                    margin=dict(t=20, b=20, l=10, r=10),
                    xaxis=dict(side="top"),
                    yaxis=dict(autorange="reversed"),
                )
                st.plotly_chart(_fig_heat, use_container_width=True)
                st.caption("🟢 ≥ 90.1%  |  🟡 70 – 90%  |  🔴 < 70%")
            else:
                st.info("Datos insuficientes para el mapa de calor.")

    with tab_profundo:

        # ── FASE 2: Matriz de Polivalencia ────────────────────────────────
        if col_actividad and not df_historico.empty:
            st.subheader("🔄 Matriz de Polivalencia")
            st.caption(
                "Historial completo · Colaboradores con captura en los últimos 30 días · Identifica quién puede cubrir a quién."
            )
            with st.expander("📖 ¿Cómo leer esta matriz?"):
                st.markdown("""
**Cada celda = eficiencia mediana de ese operador en esa actividad — calculada sobre su historial completo (operadores activos en los últimos 30 días).**
Al pasar el cursor sobre una celda se muestra la mediana y el número de registros que la respaldan.

| Color | Eficiencia | Significado | Acción |
|-------|-----------|-------------|--------|
| 🟢 Verde | ≥ 90.1% | Actividad dominada | Puede fungir como mentor o cobertura |
| 🟡 Amarillo | 70–90% | En desarrollo | Requiere supervisión periódica |
| 🔴 Rojo | < 70% | No domina la actividad | Capacitación prioritaria |
| ⬜ Blanco | Sin dato | No tiene registros de esta actividad en su historial | — |

> 💡 Para planear coberturas de ausencia: busca celdas verdes en la columna de la actividad crítica.
                """)

            _incl_inactivos = st.checkbox(
                "Incluir ex-colaboradores (posible reingreso)",
                value=False,
                key="poli_incl_inactivos",
                help="Muestra también operadores sin registros en los últimos 30 días. Útil para evaluar reingresos con datos objetivos de desempeño histórico.",
            )
            _df_fuente_poli = (
                df[df["FECHA_DT"].dt.weekday < 5].copy()
                if _incl_inactivos
                else df_historico
            )
            _df_fuente_poli = _df_fuente_poli[
                _df_fuente_poli[col_nombre]
                .astype(str)
                .str.upper()
                .str.strip()
                .isin({"", "NAN", "NONE", "N/A"})
                == False
            ].copy()

            _poli_agg_grp = (
                _df_fuente_poli.assign(
                    _OP=_df_fuente_poli[col_nombre].astype(str).str.upper().str.strip(),
                    _ACT=_df_fuente_poli[col_actividad]
                    .astype(str)
                    .str.strip()
                    .str.title(),
                )
                .groupby(["_OP", "_ACT"])["EFIC_NUM"]
                .agg(["median", "count", "std"])
                .reset_index()
            )
            _poli_agg_grp.columns = ["OPERADOR", "ACTIVIDAD", "EFICIENCIA", "N", "STD"]
            _poli_agg_grp["CV"] = (
                _poli_agg_grp["STD"]
                / _poli_agg_grp["EFICIENCIA"].replace(0, pd.NA)
                * 100
            ).fillna(0)
            _poli_agg_grp["CONF"] = _poli_agg_grp.apply(
                lambda r: (
                    "🟢"
                    if r["N"] >= 10 and r["CV"] < 25
                    else ("🟡" if r["N"] >= 5 or r["CV"] < 50 else "🔴")
                ),
                axis=1,
            )
            _poli_raw = _poli_agg_grp[["OPERADOR", "ACTIVIDAD", "EFICIENCIA"]].copy()

            # Último registro por operador (para identificar inactivos)
            _ultimo_reg = (
                _df_fuente_poli.assign(
                    _OP=_df_fuente_poli[col_nombre].astype(str).str.upper().str.strip()
                )
                .groupby("_OP")["FECHA_DT"]
                .max()
                .reset_index()
                .rename(columns={"_OP": "OPERADOR", "FECHA_DT": "ÚLTIMO REGISTRO"})
            )

            _n_acts = _poli_raw["ACTIVIDAD"].nunique()
            _n_ops = _poli_raw["OPERADOR"].nunique()
            _lbl_inc = " (incl. ex-colaboradores)" if _incl_inactivos else ""
            st.caption(
                f"Historial completo{_lbl_inc} · {_n_ops} operadores · {_n_acts} actividades detectadas"
            )

            _poli_pivot = _poli_raw.pivot(
                index="OPERADOR", columns="ACTIVIDAD", values="EFICIENCIA"
            )
            _poli_pivot_n = _poli_agg_grp.pivot(
                index="OPERADOR", columns="ACTIVIDAD", values="N"
            ).reindex_like(_poli_pivot)

            if not _poli_pivot.empty:
                _z_vals = _poli_pivot.values
                _z_text = [
                    [f"{v:.0f}%" if not pd.isna(v) else "—" for v in row]
                    for row in _z_vals
                ]
                _fig_poli = go.Figure(
                    go.Heatmap(
                        z=_z_vals,
                        x=_poli_pivot.columns.tolist(),
                        y=_poli_pivot.index.tolist(),
                        colorscale=[
                            [0.0, "#e74c3c"],
                            [0.7, "#e74c3c"],
                            [0.7, "#f39c12"],
                            [0.901, "#f39c12"],
                            [0.901, "#27ae60"],
                            [1.0, "#27ae60"],
                        ],
                        zmin=0,
                        zmax=100,
                        text=_z_text,
                        texttemplate="%{text}",
                        textfont=dict(size=10, color="white"),
                        customdata=_poli_pivot_n.fillna(0).values,
                        hovertemplate="<b>%{y}</b><br>%{x}: %{z:.1f}% (mediana)<br>Registros: %{customdata:.0f}<extra></extra>",
                        colorbar=dict(
                            title="Efic. %",
                            tickvals=[0, 70, 90, 100],
                            ticktext=["0%", "70% mín", "90.1% meta", "100%"],
                        ),
                    )
                )
                _fig_poli.update_layout(
                    height=max(350, len(_poli_pivot) * 32 + 120),
                    margin=dict(t=20, b=20, l=10, r=10),
                    xaxis=dict(side="top", tickangle=-35),
                    yaxis=dict(autorange="reversed"),
                )
                st.plotly_chart(_fig_poli, use_container_width=True)

                # Índice de versatilidad
                _total_acts = len(_poli_pivot.columns)
                _dominio = (_poli_pivot >= 90.1).sum(axis=1).reset_index()
                _dominio.columns = ["OPERADOR", "DOMINADAS"]
                _dominio["COBERTURA"] = _dominio["DOMINADAS"].apply(
                    lambda v: f"{v} de {_total_acts}"
                )
                _dominio["VERSATILIDAD"] = _dominio["DOMINADAS"].apply(
                    lambda v: (
                        "🟢 Alta"
                        if v >= _total_acts * 0.6
                        else "🟡 Media" if v >= _total_acts * 0.3 else "🔴 Baja"
                    )
                )
                st.markdown("##### Índice de Versatilidad por Operador")
                st.dataframe(
                    _dominio.sort_values("DOMINADAS", ascending=False)[
                        ["OPERADOR", "COBERTURA", "VERSATILIDAD"]
                    ],
                    column_config={
                        "COBERTURA": st.column_config.TextColumn(
                            "ACTIVIDADES DOMINADAS",
                            help="Cuántas actividades tiene con eficiencia ≥ 90.1% del total del periodo.",
                        ),
                        "VERSATILIDAD": st.column_config.TextColumn(
                            "VERSATILIDAD",
                            help="Alta: domina ≥ 60% de actividades | Media: 30–60% | Baja: < 30%",
                        ),
                    },
                    hide_index=True,
                    use_container_width=True,
                )

                st.markdown("##### 🎓 Matriz de Habilidades — Calificación Final")
                with st.expander(
                    "📖 ¿Qué significa cada nivel y qué acción corresponde?"
                ):
                    st.markdown("""
| Nivel | Criterio | Propuesta de acción |
|-------|----------|---------------------|
| 🥇 **Experto** | Eficiencia mediana ≥ 90.1% **y** domina ≥ 60% de actividades | Mantener ritmo · Asignar como mentor en actividades dominadas · Evaluar para actividades de mayor complejidad |
| 🥈 **Competente** | Eficiencia mediana ≥ 70% **y** domina ≥ 30% de actividades | Plan para dominar actividades en zona amarilla · Mentoría con operadores expertos · Ampliar rotación gradual |
| 🥉 **En Desarrollo** | Eficiencia mediana ≥ 60% | Capacitación enfocada en actividades críticas · Supervisión periódica · Reducir rotación hasta consolidar base |
| 🔴 **Básico** | Eficiencia mediana < 60% | Diagnóstico individual urgente · Plan de capacitación estructurado · Evaluar reubicación a actividades más simples |

> 💡 La eficiencia es la **mediana** sobre el historial completo del operador — resistente a días atípicos y outliers de captura.
                    """)

                _efic_global = (
                    _poli_agg_grp.groupby("OPERADOR")["EFICIENCIA"]
                    .median()
                    .reset_index()
                    .rename(columns={"EFICIENCIA": "_EFIC_G"})
                )
                _matriz_h = _dominio[["OPERADOR", "DOMINADAS"]].merge(
                    _efic_global, on="OPERADOR"
                )
                _matriz_h["% DOMINADAS"] = (
                    _matriz_h["DOMINADAS"] / _total_acts * 100
                ).round(1)
                _matriz_h["EFIC MEDIANA GLOBAL"] = _matriz_h["_EFIC_G"].round(1)

                def _nivel_hab(efic, dom_pct):
                    if efic >= 90.1 and dom_pct >= 60:
                        return "🥇 Experto"
                    if efic >= 70 and dom_pct >= 30:
                        return "🥈 Competente"
                    if efic >= 60:
                        return "🥉 En Desarrollo"
                    return "🔴 Básico"

                def _propuesta_hab(nivel):
                    return {
                        "🥇 Experto": "Mantener ritmo · Asignar como mentor · Evaluar actividades de mayor complejidad",
                        "🥈 Competente": "Dominar actividades en zona amarilla · Mentoría con expertos · Ampliar rotación gradual",
                        "🥉 En Desarrollo": "Capacitación en actividades críticas · Supervisión periódica · Reducir rotación",
                        "🔴 Básico": "Diagnóstico individual urgente · Plan de capacitación estructurado",
                    }.get(nivel, "")

                _matriz_h["NIVEL"] = _matriz_h.apply(
                    lambda r: _nivel_hab(r["_EFIC_G"], r["% DOMINADAS"]), axis=1
                )
                _matriz_h["PROPUESTA DE ACCIÓN"] = _matriz_h["NIVEL"].apply(
                    _propuesta_hab
                )
                # Estatus activo/inactivo
                _matriz_h["ESTATUS"] = _matriz_h["OPERADOR"].apply(
                    lambda op: "🟢 Activo" if op in _ops_activos else "⚫ Inactivo"
                )
                # Último registro
                _matriz_h = _matriz_h.merge(_ultimo_reg, on="OPERADOR", how="left")
                _matriz_h["ÚLTIMO REGISTRO"] = (
                    pd.to_datetime(_matriz_h["ÚLTIMO REGISTRO"], errors="coerce")
                    .dt.strftime("%d/%m/%Y")
                    .fillna("—")
                )
                # Ordenar: activos primero, luego por nivel y eficiencia
                _ord_niv = {
                    "🥇 Experto": 0,
                    "🥈 Competente": 1,
                    "🥉 En Desarrollo": 2,
                    "🔴 Básico": 3,
                }
                _ord_est = {"🟢 Activo": 0, "⚫ Inactivo": 1}
                _matriz_h = (
                    _matriz_h.assign(
                        _ORD=_matriz_h["NIVEL"].map(_ord_niv),
                        _EST=_matriz_h["ESTATUS"].map(_ord_est),
                    )
                    .sort_values(
                        ["_EST", "_ORD", "EFIC MEDIANA GLOBAL"],
                        ascending=[True, True, False],
                    )
                    .drop(columns=["_ORD", "_EST", "_EFIC_G"])
                )
                _cols_matriz = [
                    "OPERADOR",
                    "ESTATUS",
                    "EFIC MEDIANA GLOBAL",
                    "DOMINADAS",
                    "% DOMINADAS",
                    "NIVEL",
                    "ÚLTIMO REGISTRO",
                    "PROPUESTA DE ACCIÓN",
                ]
                st.dataframe(
                    _matriz_h[_cols_matriz],
                    column_config={
                        "ESTATUS": st.column_config.TextColumn(
                            "ESTATUS",
                            help="🟢 Activo = registros en últimos 30 días · ⚫ Inactivo = posible reingreso",
                        ),
                        "EFIC MEDIANA GLOBAL": st.column_config.ProgressColumn(
                            "EFICIENCIA GLOBAL (mediana)",
                            format="%.1f%%",
                            min_value=0,
                            max_value=100,
                            help="Mediana de la eficiencia del operador sobre todas sus actividades en el historial.",
                        ),
                        "DOMINADAS": st.column_config.NumberColumn(
                            f"DOMINADAS (de {_total_acts})",
                            format="%d",
                            help="Número de actividades con eficiencia mediana ≥ 90.1%.",
                        ),
                        "% DOMINADAS": st.column_config.NumberColumn(
                            "% ACTIVIDADES DOMINADAS", format="%.1f %%"
                        ),
                        "NIVEL": st.column_config.TextColumn("NIVEL"),
                        "ÚLTIMO REGISTRO": st.column_config.TextColumn(
                            "ÚLTIMO REGISTRO",
                            help="Fecha del último registro en la hoja REGISTRO.",
                        ),
                        "PROPUESTA DE ACCIÓN": st.column_config.TextColumn(
                            "PROPUESTA DE ACCIÓN", width="large"
                        ),
                    },
                    hide_index=True,
                    use_container_width=True,
                )

                def _excel_habilidades():
                    from openpyxl.styles import PatternFill, Font

                    _buf_h = io.BytesIO()
                    with pd.ExcelWriter(_buf_h, engine="openpyxl") as _wr_h:
                        _poli_xl = _poli_pivot.copy().reset_index()
                        _poli_xl.columns.name = None
                        _poli_xl.to_excel(
                            _wr_h, sheet_name="MATRIZ DE HABILIDADES", index=False
                        )
                        _matriz_h[_cols_matriz].to_excel(
                            _wr_h, sheet_name="CALIFICACIÓN FINAL", index=False
                        )
                        _ws = _wr_h.book["MATRIZ DE HABILIDADES"]
                        _g = PatternFill("solid", fgColor="1D6F42")
                        _y = PatternFill("solid", fgColor="9C5A00")
                        _r = PatternFill("solid", fgColor="8B0000")
                        _wf = Font(color="FFFFFF", bold=True)
                        for _row in _ws.iter_rows(min_row=2, min_col=2):
                            for _cell in _row:
                                try:
                                    _v = float(_cell.value)
                                    _cell.fill = (
                                        _g if _v >= 90.1 else (_y if _v >= 70 else _r)
                                    )
                                    _cell.font = _wf
                                except (TypeError, ValueError):
                                    pass
                    return _buf_h.getvalue()

                st.download_button(
                    label="📥 Descargar Matriz de Habilidades en Excel",
                    data=_excel_habilidades(),
                    file_name=f"Matriz_Habilidades_{ahora_local().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_habilidades",
                )

                # ── Polivalencia por Área ─────────────────────────────────
                st.markdown("---")
                st.markdown("##### 🏭 Polivalencia por Área")
                st.caption(
                    "Versatilidad medida solo dentro del área asignada — denominador justo: "
                    "N actividades propias del área, no todas las de la planta."
                )
                if col_area:
                    _areas_reg = sorted(
                        _df_fuente_poli[col_area]
                        .astype(str)
                        .str.strip()
                        .str.upper()
                        .replace({"": pd.NA, "NAN": pd.NA, "NONE": pd.NA, "N/A": pd.NA})
                        .dropna()
                        .unique()
                        .tolist()
                    )
                    if _areas_reg:
                        _sel_area_poli = st.selectbox(
                            "Área a analizar:", _areas_reg, key="poli_area_sel"
                        )
                        # Filtrar REGISTRO directamente por área — sin cruce con BDD
                        _df_area_data = _df_fuente_poli[
                            _df_fuente_poli[col_area]
                            .astype(str)
                            .str.strip()
                            .str.upper()
                            == _sel_area_poli
                        ].copy()
                        _poli_area = (
                            _df_area_data.assign(
                                _OP=_df_area_data[col_nombre]
                                .astype(str)
                                .str.upper()
                                .str.strip(),
                                _ACT=_df_area_data[col_actividad]
                                .astype(str)
                                .str.strip()
                                .str.title(),
                            )
                            .groupby(["_OP", "_ACT"])["EFIC_NUM"]
                            .agg(["median", "count", "std"])
                            .reset_index()
                        )
                        _poli_area.columns = [
                            "OPERADOR",
                            "ACTIVIDAD",
                            "EFICIENCIA",
                            "N",
                            "STD",
                        ]
                        _poli_area = _poli_area[
                            ~_poli_area["OPERADOR"].isin({"", "NAN", "NONE", "N/A"})
                        ].copy()

                        if _poli_area.empty:
                            st.info(
                                f"Sin datos históricos en el área **{_sel_area_poli.title()}**."
                            )
                        else:
                            _poli_pivot_area = _poli_area.pivot(
                                index="OPERADOR",
                                columns="ACTIVIDAD",
                                values="EFICIENCIA",
                            )
                            _poli_n_area = _poli_area.pivot(
                                index="OPERADOR", columns="ACTIVIDAD", values="N"
                            ).reindex_like(_poli_pivot_area)
                            _total_acts_area = len(_poli_pivot_area.columns)
                            _n_ops_area = len(_poli_pivot_area.index)
                            st.caption(
                                f"{_n_ops_area} operadores · {_total_acts_area} actividades propias "
                                f"del área **{_sel_area_poli.title()}**"
                            )

                            _za = _poli_pivot_area.values
                            _zt = [
                                [f"{v:.0f}%" if not pd.isna(v) else "—" for v in row]
                                for row in _za
                            ]
                            _fig_pa = go.Figure(
                                go.Heatmap(
                                    z=_za,
                                    x=_poli_pivot_area.columns.tolist(),
                                    y=_poli_pivot_area.index.tolist(),
                                    colorscale=[
                                        [0.0, "#e74c3c"],
                                        [0.7, "#e74c3c"],
                                        [0.7, "#f39c12"],
                                        [0.901, "#f39c12"],
                                        [0.901, "#27ae60"],
                                        [1.0, "#27ae60"],
                                    ],
                                    zmin=0,
                                    zmax=100,
                                    text=_zt,
                                    texttemplate="%{text}",
                                    textfont=dict(size=11, color="white"),
                                    customdata=_poli_n_area.fillna(0).values,
                                    hovertemplate="<b>%{y}</b><br>%{x}: %{z:.1f}% (mediana)<br>Registros: %{customdata:.0f}<extra></extra>",
                                    colorbar=dict(
                                        title="Efic. %",
                                        tickvals=[0, 70, 90, 100],
                                        ticktext=[
                                            "0%",
                                            "70% mín",
                                            "90.1% meta",
                                            "100%",
                                        ],
                                    ),
                                )
                            )
                            _fig_pa.update_layout(
                                height=max(280, _n_ops_area * 32 + 100),
                                margin=dict(t=20, b=20, l=10, r=10),
                                xaxis=dict(side="top", tickangle=-35),
                                yaxis=dict(autorange="reversed"),
                            )
                            st.plotly_chart(_fig_pa, use_container_width=True)

                            # Calificación dentro del área
                            _dom_a = (
                                (_poli_pivot_area >= 90.1).sum(axis=1).reset_index()
                            )
                            _dom_a.columns = ["OPERADOR", "DOMINADAS"]
                            _dom_a["% DOMINADAS"] = (
                                _dom_a["DOMINADAS"] / _total_acts_area * 100
                            ).round(1)
                            _efic_a = (
                                _poli_area.groupby("OPERADOR")["EFICIENCIA"]
                                .median()
                                .reset_index()
                                .rename(columns={"EFICIENCIA": "_EG"})
                            )
                            _mh_a = _dom_a.merge(_efic_a, on="OPERADOR")
                            _mh_a["EFIC MEDIANA ÁREA"] = _mh_a["_EG"].round(1)
                            _mh_a["ESTATUS"] = _mh_a["OPERADOR"].apply(
                                lambda op: (
                                    "🟢 Activo" if op in _ops_activos else "⚫ Inactivo"
                                )
                            )
                            _mh_a["NIVEL"] = _mh_a.apply(
                                lambda r: _nivel_hab(r["_EG"], r["% DOMINADAS"]), axis=1
                            )
                            _mh_a["PROPUESTA DE ACCIÓN"] = _mh_a["NIVEL"].apply(
                                _propuesta_hab
                            )
                            _mh_a = (
                                _mh_a.assign(
                                    _ORD=_mh_a["NIVEL"].map(_ord_niv),
                                    _EST=_mh_a["ESTATUS"].map(
                                        {"🟢 Activo": 0, "⚫ Inactivo": 1}
                                    ),
                                )
                                .sort_values(
                                    ["_EST", "_ORD", "EFIC MEDIANA ÁREA"],
                                    ascending=[True, True, False],
                                )
                                .drop(columns=["_ORD", "_EST", "_EG"])
                            )
                            st.markdown(
                                f"###### Calificación en {_sel_area_poli.title()} "
                                f"— {_total_acts_area} actividades del área"
                            )
                            st.dataframe(
                                _mh_a[
                                    [
                                        "OPERADOR",
                                        "ESTATUS",
                                        "EFIC MEDIANA ÁREA",
                                        "DOMINADAS",
                                        "% DOMINADAS",
                                        "NIVEL",
                                        "PROPUESTA DE ACCIÓN",
                                    ]
                                ],
                                column_config={
                                    "ESTATUS": st.column_config.TextColumn(
                                        "ESTATUS",
                                        help="🟢 Activo = registros en últimos 30 días · ⚫ Inactivo = posible reingreso",
                                    ),
                                    "EFIC MEDIANA ÁREA": st.column_config.ProgressColumn(
                                        f"EFICIENCIA EN {_sel_area_poli.title()}",
                                        format="%.1f%%",
                                        min_value=0,
                                        max_value=100,
                                        help="Mediana de eficiencia del operador solo en actividades de esta área.",
                                    ),
                                    "DOMINADAS": st.column_config.NumberColumn(
                                        f"DOMINADAS (de {_total_acts_area})",
                                        format="%d",
                                        help="Actividades del área con eficiencia mediana ≥ 90.1%.",
                                    ),
                                    "% DOMINADAS": st.column_config.NumberColumn(
                                        "% ACTIVIDADES DEL ÁREA", format="%.1f %%"
                                    ),
                                    "NIVEL": st.column_config.TextColumn("NIVEL"),
                                    "PROPUESTA DE ACCIÓN": st.column_config.TextColumn(
                                        "PROPUESTA DE ACCIÓN", width="large"
                                    ),
                                },
                                hide_index=True,
                                use_container_width=True,
                            )
                    else:
                        st.info(
                            "No se encontraron áreas registradas en la hoja REGISTRO."
                        )
                else:
                    st.info(
                        "Se requiere la columna ÁREA en REGISTRO para habilitar la vista por área."
                    )

            st.divider()

            # ── FASE 2: Diagnóstico Proceso vs. Persona ───────────────────
            st.subheader("🔬 ¿Problema de Proceso o de Persona?")
            st.caption(
                "Cuando muchos operadores fallan en la misma actividad, el problema no es la gente."
            )
            with st.expander("📖 ¿Cómo leer este diagnóstico?"):
                st.markdown("""
**Para cada actividad se calcula qué % de registros tuvieron eficiencia < 70%.**

| Diagnóstico | Condición | Qué indica | Acción recomendada |
|-------------|-----------|------------|-------------------|
| 🔴 Cuello sistémico | > 60% en rojo | Problema de proceso, no de persona | Revisar método, herramental o tiempo estándar |
| 🟡 Problema mixto | 30–60% en rojo | Proceso difícil + operadores que necesitan refuerzo | Capacitación grupal + mejora de condiciones |
| 🟢 Caso individual | < 30% en rojo | Actividad viable — algunos operadores necesitan apoyo | Mentoría individual |

> 💡 Columna **REGISTROS**: si tiene menos de 5 registros, el dato puede no ser representativo.
                """)

            _proc_df = (
                df_historico[[col_actividad, "EFIC_NUM"]]
                .copy()
                .assign(
                    _ACT=df_historico[col_actividad]
                    .astype(str)
                    .str.strip()
                    .str.title(),
                    _ROJO=lambda x: x["EFIC_NUM"] < 70,
                )
                .groupby("_ACT")
                .agg(
                    REGISTROS=("EFIC_NUM", "count"),
                    EN_ROJO=("_ROJO", "sum"),
                    EFIC_PROM=("EFIC_NUM", "mean"),
                )
                .reset_index()
            )
            _proc_df.columns = ["ACTIVIDAD", "REGISTROS", "EN_ROJO", "EFIC. PROM. (%)"]
            _proc_df["% EN ROJO"] = (
                _proc_df["EN_ROJO"] / _proc_df["REGISTROS"] * 100
            ).round(1)
            _proc_df["EFIC. PROM. (%)"] = _proc_df["EFIC. PROM. (%)"].round(1)
            _proc_df["DIAGNÓSTICO"] = _proc_df["% EN ROJO"].apply(
                lambda v: (
                    "🔴 Cuello sistémico"
                    if v > 60
                    else ("🟡 Problema mixto" if v > 30 else "🟢 Caso individual")
                )
            )
            _proc_df["ACCIÓN"] = _proc_df["% EN ROJO"].apply(
                lambda v: (
                    "Revisar método / herramental / estándar"
                    if v > 60
                    else (
                        "Capacitación grupal + mejora de condiciones"
                        if v > 30
                        else "Mentoría individual"
                    )
                )
            )
            _proc_df = _proc_df.sort_values("% EN ROJO", ascending=False)

            st.dataframe(
                _proc_df[
                    [
                        "ACTIVIDAD",
                        "EFIC. PROM. (%)",
                        "% EN ROJO",
                        "DIAGNÓSTICO",
                        "ACCIÓN",
                        "REGISTROS",
                    ]
                ],
                column_config={
                    "EFIC. PROM. (%)": st.column_config.ProgressColumn(
                        "EFIC. PROM. (%)",
                        format="%.1f%%",
                        min_value=0,
                        max_value=100,
                        help="Eficiencia promedio de todos los registros de esta actividad.",
                    ),
                    "% EN ROJO": st.column_config.ProgressColumn(
                        "% EN ROJO",
                        format="%.1f%%",
                        min_value=0,
                        max_value=100,
                        help="> 60% = problema sistémico de proceso | 30–60% = mixto | < 30% = individual.",
                    ),
                    "DIAGNÓSTICO": st.column_config.TextColumn(
                        "DIAGNÓSTICO",
                        help="Clasificación automática según el % de registros en zona crítica.",
                    ),
                    "ACCIÓN": st.column_config.TextColumn(
                        "ACCIÓN RECOMENDADA",
                        help="Siguiente paso sugerido según el diagnóstico.",
                    ),
                    "REGISTROS": st.column_config.NumberColumn(
                        "REGISTROS",
                        format="%d",
                        help="Cantidad de registros de esta actividad. Menos de 5 = dato poco representativo.",
                    ),
                },
                hide_index=True,
                use_container_width=True,
            )

            _cuellos = _proc_df[_proc_df["% EN ROJO"] > 60]
            if not _cuellos.empty:
                st.error(
                    f"🔴 **{len(_cuellos)} actividad(es) con cuello sistémico:** "
                    f"{', '.join(_cuellos['ACTIVIDAD'].tolist())} — escalar a ingeniería de proceso."
                )

            st.divider()

        elif not col_actividad:
            st.info(
                "ℹ️ No se detectó columna de ACTIVIDAD en la hoja REGISTRO. "
                "La Fase 2 (Polivalencia y Diagnóstico) requiere esa columna."
            )
            st.divider()

        # ── FASE 3A: PxH Real vs. Estándar ───────────────────────────────
        if col_tiempo and col_capxh and col_cantidad and not df_historico.empty:
            st.subheader("⚡ PxH Real vs. Estándar")
            st.caption(
                "Historial completo (todos los operadores, incl. ex-activos) · Agrupado por Pieza + Subproceso para calibrar el estándar."
            )
            with st.expander("📖 ¿Cómo leer este análisis?"):
                st.markdown("""
**PxH Real** = CANTIDAD producida ÷ TIEMPO USADO (en horas)
**PxH Estándar** = CAP PXH del registro — específico para esa Pieza + Subproceso

El análisis se hace por **Pieza + Subproceso** porque el mismo subproceso tiene diferente estándar según la pieza.

**Vista por Pieza+Subproceso:** ¿el estándar de esa combinación está bien calibrado?
Usa el historial completo de **todos los operadores** (incluyendo ex-activos) para tener la mayor evidencia posible. La tabla aparece ordenada de peor a mejor (los estándares más cuestionables arriba). Usa el slider para filtrar por mínimo de registros y eliminar el ruido estadístico.

| Diagnóstico | Condición | Qué indica |
|-------------|-----------|------------|
| 🟢 Viable | > 85% de registros alcanzan el CAP PXH | Estándar bien calibrado |
| 🟡 Ajustado | 50 – 85% lo alcanzan | Alcanzable pero exigente |
| 🔴 Irreal | < 50% lo alcanzan | CAP PXH posiblemente sobreestimado |

**Vista por Operador:** muestra solo operadores activos en los últimos 30 días. Cada registro se compara contra *su propio* CAP PXH (brecha normalizada) y se calcula la **mediana** sobre su historial completo. Así operadores que trabajan piezas difíciles no quedan en desventaja.

> 💡 Brecha normalizada del 100% = el operador alcanza exactamente su estándar en mediana.
                """)

            _df3 = df[df["FECHA_DT"].dt.weekday < 5].copy()
            _df3["_TIEMPO_H"] = (
                pd.to_timedelta(
                    _df3[col_tiempo].astype(str), errors="coerce"
                ).dt.total_seconds()
                / 3600
            )
            _df3["_CANT_NUM"] = convertir_serie_numerica(_df3[col_cantidad]).fillna(0)
            _df3["_CAP_NUM"] = convertir_serie_numerica(_df3[col_capxh]).fillna(0)
            _df3["_ACT"] = (
                _df3[col_actividad].astype(str).str.strip().str.title()
                if col_actividad
                else "—"
            )
            _df3["_PZ"] = _df3[col_pieza].astype(str).str.strip() if col_pieza else "—"
            _df3["_OP"] = _df3[col_nombre].astype(str).str.upper().str.strip()

            _mask3 = (
                (_df3["_TIEMPO_H"] > 0)
                & (_df3["_CANT_NUM"] > 0)
                & (_df3["_CAP_NUM"] > 0)
            )
            _df3 = _df3[_mask3].copy()
            _df3["_PXH_REAL"] = _df3["_CANT_NUM"] / _df3["_TIEMPO_H"]

            # Filtrar outliers: PxH_Real > 2.5× estándar → sospechoso de error de captura
            _n_antes = len(_df3)
            _df3 = _df3[_df3["_PXH_REAL"] <= _df3["_CAP_NUM"] * 2.5].copy()
            _n_filtrados = _n_antes - len(_df3)

            _df3["_ALCANZA"] = _df3["_PXH_REAL"] >= _df3["_CAP_NUM"]
            _df3["_BRECHA_REG"] = (_df3["_PXH_REAL"] / _df3["_CAP_NUM"]).clip(upper=2.0)

            if _df3.empty:
                st.info("No hay registros válidos para calcular PxH.")
            else:
                if _n_filtrados > 0:
                    st.caption(
                        f"ℹ️ {_n_filtrados} registro(s) excluido(s) por superar 2.5× su CAP PXH "
                        f"(posibles errores de captura)."
                    )

                # ── Por Pieza + Subproceso ────────────────────────────────
                st.markdown("##### Por Pieza + Subproceso")
                st.caption(
                    "Mediana del PxH real (resistente a outliers) y CV como indicador de confiabilidad del dato."
                )
                _pxh_ps = (
                    _df3.groupby(["_PZ", "_ACT"])
                    .agg(
                        PXH_MED=("_PXH_REAL", "median"),
                        PXH_STD=("_PXH_REAL", "std"),
                        CAP_STD=("_CAP_NUM", "median"),
                        REGISTROS=("_PXH_REAL", "count"),
                        PCT_OK=("_ALCANZA", "mean"),
                    )
                    .reset_index()
                )
                _pxh_ps.columns = [
                    "PIEZA",
                    "SUBPROCESO",
                    "PXH MED",
                    "PXH STD",
                    "CAP PXH STD",
                    "REGISTROS",
                    "PCT_OK",
                ]
                _pxh_ps["PXH MED"] = _pxh_ps["PXH MED"].round(1)
                _pxh_ps["CAP PXH STD"] = _pxh_ps["CAP PXH STD"].round(1)
                _pxh_ps["CV PxH (%)"] = (
                    _pxh_ps["PXH STD"].fillna(0)
                    / _pxh_ps["PXH MED"].replace(0, 1)
                    * 100
                ).round(1)
                _pxh_ps["BRECHA (%)"] = (
                    _pxh_ps["PXH MED"] / _pxh_ps["CAP PXH STD"] * 100
                ).round(1)
                _pxh_ps["% ALCANZA"] = (_pxh_ps["PCT_OK"] * 100).round(1)
                _pxh_ps["CONFIANZA"] = _pxh_ps.apply(
                    lambda r: (
                        "⚠️ Muestra insuf."
                        if r["REGISTROS"] < 5
                        else (
                            "🟢 Confiable"
                            if r["CV PxH (%)"] < 25
                            else (
                                "🟡 Dispersión media"
                                if r["CV PxH (%)"] < 50
                                else "🔴 Alta dispersión"
                            )
                        )
                    ),
                    axis=1,
                )
                _pxh_ps["DIAGNÓSTICO"] = _pxh_ps.apply(
                    lambda r: (
                        "— Muestra insuf."
                        if r["REGISTROS"] < 5
                        else (
                            "🟢 Viable"
                            if r["% ALCANZA"] >= 85
                            else "🟡 Ajustado" if r["% ALCANZA"] >= 50 else "🔴 Irreal"
                        )
                    ),
                    axis=1,
                )
                _pxh_ps["_SUFF"] = (_pxh_ps["REGISTROS"] >= 5).astype(int)
                _pxh_ps = _pxh_ps.sort_values(
                    ["_SUFF", "% ALCANZA", "CV PxH (%)"], ascending=[False, True, False]
                ).drop(columns=["_SUFF"])

                _min_reg = st.slider(
                    "Mínimo de registros:",
                    min_value=1,
                    max_value=20,
                    value=5,
                    step=1,
                    key="pxh_min_reg",
                    help="Oculta combinaciones Pieza+Subproceso con menos registros de los seleccionados.",
                )
                _pxh_display = _pxh_ps[_pxh_ps["REGISTROS"] >= _min_reg]

                st.dataframe(
                    _pxh_display[
                        [
                            "PIEZA",
                            "SUBPROCESO",
                            "CAP PXH STD",
                            "PXH MED",
                            "BRECHA (%)",
                            "CV PxH (%)",
                            "% ALCANZA",
                            "CONFIANZA",
                            "DIAGNÓSTICO",
                            "REGISTROS",
                        ]
                    ],
                    column_config={
                        "CAP PXH STD": st.column_config.NumberColumn(
                            "ESTÁNDAR (pzs/h)",
                            format="%.1f",
                            help="CAP PXH — estándar específico para esta Pieza+Subproceso.",
                        ),
                        "PXH MED": st.column_config.NumberColumn(
                            "MEDIANA PxH (pzs/h)",
                            format="%.1f",
                            help="Mediana de la velocidad real — resistente a outliers y datos de rotación.",
                        ),
                        "BRECHA (%)": st.column_config.ProgressColumn(
                            "REAL vs ESTÁNDAR (%)",
                            format="%.1f%%",
                            min_value=0,
                            max_value=150,
                            help="Mediana PxH ÷ CAP PXH × 100. >100% = supera el estándar en mediana.",
                        ),
                        "CV PxH (%)": st.column_config.NumberColumn(
                            "CV PxH (%)",
                            format="%.1f %%",
                            help="Coeficiente de Variación de la velocidad real. "
                            "<25% = datos homogéneos · 25–50% = dispersión media · >50% = datos poco confiables.",
                        ),
                        "% ALCANZA": st.column_config.ProgressColumn(
                            "% REGISTROS QUE ALCANZAN",
                            format="%.1f%%",
                            min_value=0,
                            max_value=100,
                            help="% de registros (limpios) donde PxH Real ≥ CAP PXH.",
                        ),
                        "CONFIANZA": st.column_config.TextColumn(
                            "CONFIANZA DEL DATO",
                            help="Basada en CV: 🟢 <25% · 🟡 25–50% · 🔴 >50% · ⚠️ muestra < 5 registros",
                        ),
                        "DIAGNÓSTICO": st.column_config.TextColumn(
                            "DIAGNÓSTICO",
                            help="🟢 >85% alcanzan | 🟡 50–85% | 🔴 <50% → revisar CAP PXH con ingeniería.",
                        ),
                        "REGISTROS": st.column_config.NumberColumn(
                            "REGISTROS",
                            format="%d",
                            help="Cantidad de registros limpios (excluyendo outliers >2.5× estándar).",
                        ),
                    },
                    hide_index=True,
                    use_container_width=True,
                )

                # Solo alertar sobre combinaciones con datos suficientes (≥ 5 registros)
                _irreales = _pxh_ps[
                    (_pxh_ps["% ALCANZA"] < 50) & (_pxh_ps["REGISTROS"] >= 5)
                ]
                if not _irreales.empty:
                    _sub_afect = (
                        _irreales.groupby("SUBPROCESO")["REGISTROS"]
                        .sum()
                        .sort_values(ascending=False)
                        .head(5)
                    )
                    _resumen_subs = " · ".join(
                        f"**{s}** ({n} reg.)" for s, n in _sub_afect.items()
                    )
                    st.warning(
                        f"⚠️ **{len(_irreales)} combinación(es) con ≥ 5 registros donde < 50% alcanza el CAP PXH.** "
                        f"Subprocesos más afectados: {_resumen_subs}. "
                        f"Revisar con ingeniería — ver tabla ordenada de peor a mejor."
                    )

                st.divider()

                # ── Por Operador (mediana de brecha normalizada + CV) ─────
                st.markdown("##### Por Operador")
                st.caption(
                    "Operadores activos en los últimos 30 días · Brecha calculada sobre su historial completo. "
                    "100% = alcanza exactamente su estándar en mediana. CV indica consistencia."
                )
                _pxh_op = (
                    _df3[_df3["_OP"].isin(_ops_activos)]
                    .groupby("_OP")
                    .agg(
                        BRECHA_MED=("_BRECHA_REG", "median"),
                        BRECHA_STD=("_BRECHA_REG", "std"),
                        REGISTROS=("_BRECHA_REG", "count"),
                        PCT_OK=("_ALCANZA", "mean"),
                    )
                    .reset_index()
                )
                _pxh_op.columns = [
                    "OPERADOR",
                    "BRECHA_MED",
                    "BRECHA_STD",
                    "REGISTROS",
                    "PCT_OK",
                ]
                _pxh_op["BRECHA vs ESTÁNDAR (%)"] = (_pxh_op["BRECHA_MED"] * 100).round(
                    1
                )
                _pxh_op["CV Brecha (%)"] = (
                    _pxh_op["BRECHA_STD"].fillna(0)
                    / _pxh_op["BRECHA_MED"].replace(0, 1)
                    * 100
                ).round(1)
                _pxh_op["% REGISTROS OK"] = (_pxh_op["PCT_OK"] * 100).round(1)
                _pxh_op["ESTATUS"] = _pxh_op["BRECHA vs ESTÁNDAR (%)"].apply(
                    lambda v: (
                        "🟢 Supera estándar"
                        if v >= 100
                        else ("🟡 En zona" if v >= 80 else "🔴 Por debajo")
                    )
                )
                _pxh_op = _pxh_op.sort_values("BRECHA vs ESTÁNDAR (%)", ascending=False)

                st.dataframe(
                    _pxh_op[
                        [
                            "OPERADOR",
                            "BRECHA vs ESTÁNDAR (%)",
                            "CV Brecha (%)",
                            "% REGISTROS OK",
                            "ESTATUS",
                            "REGISTROS",
                        ]
                    ],
                    column_config={
                        "BRECHA vs ESTÁNDAR (%)": st.column_config.ProgressColumn(
                            "BRECHA vs ESTÁNDAR (%)",
                            format="%.1f%%",
                            min_value=0,
                            max_value=150,
                            help="Mediana de (PxH Real ÷ CAP PXH) × 100. "
                            "100% = alcanza su estándar en mediana. Normalizado por pieza+subproceso.",
                        ),
                        "CV Brecha (%)": st.column_config.NumberColumn(
                            "CV Brecha (%)",
                            format="%.1f %%",
                            help="Coeficiente de Variación de su brecha normalizada. "
                            "<25% = velocidad consistente · >50% = velocidad muy errática.",
                        ),
                        "% REGISTROS OK": st.column_config.ProgressColumn(
                            "% REGISTROS QUE ALCANZAN",
                            format="%.1f%%",
                            min_value=0,
                            max_value=100,
                            help="Porcentaje de registros limpios donde igualó o superó su CAP PXH.",
                        ),
                        "ESTATUS": st.column_config.TextColumn(
                            "ESTATUS", help="🟢 Brecha ≥ 100% | 🟡 80–100% | 🔴 < 80%"
                        ),
                        "REGISTROS": st.column_config.NumberColumn(
                            "REGISTROS",
                            format="%d",
                            help="Registros limpios (excluidos outliers >2.5× estándar).",
                        ),
                    },
                    hide_index=True,
                    use_container_width=True,
                )

            st.divider()

        elif not (col_tiempo and col_capxh):
            _falt3 = []
            if not col_tiempo:
                _falt3.append("TIEMPO USADO")
            if not col_capxh:
                _falt3.append("CAP PXH")
            st.info(
                f"ℹ️ Fase 3A requiere columnas en REGISTRO: **{', '.join(_falt3)}**."
            )

        # ── FASE 3B: Análisis de Jornada ──────────────────────────────────
        if col_hora_ini and col_hora_fin and not df_historico.empty:
            st.subheader("🕐 Análisis de Jornada")
            st.caption(
                "Hora de arranque, hora de cierre y horas activas promedio por colaborador."
            )
            with st.expander("📖 ¿Cómo leer este análisis?"):
                st.markdown("""
**Cada fila resume el patrón de jornada de un colaborador en el historial de los últimos 30 días.**

| Columna | Qué mide |
|---------|----------|
| **HORA INICIO PROM.** | Hora promedio a la que empieza sus registros — mide puntualidad |
| **HORA FIN PROM.** | Hora promedio de su último registro del día |
| **HORAS ACTIVAS PROM.** | Promedio de horas entre inicio y fin de jornada |
| **DÍAS CON REGISTRO** | Días en el historial donde tuvo al menos un registro |

> 💡 La tabla está ordenada por hora de inicio: los que arrancan más tarde aparecen primero.
> Un arranque tardío consistente puede indicar ausentismo parcial, llegadas tarde o captura de datos retrasada.
                """)

            def _parsear_hora_jornada(serie):
                _s = (
                    serie.astype(str)
                    .str.strip()
                    .str.replace("a.m.", "AM", regex=False)
                    .str.replace("p.m.", "PM", regex=False)
                )
                return pd.to_datetime(_s, format="%I:%M:%S %p", errors="coerce")

            def _dec_a_str(h):
                if pd.isna(h):
                    return "—"
                _hh, _mm = int(h), int(round((h - int(h)) * 60))
                _suf = "AM" if _hh < 12 else "PM"
                _hh12 = _hh % 12 or 12
                return f"{_hh12}:{_mm:02d} {_suf}"

            _df3b = df_historico.copy()
            _df3b["_INI_DT"] = _parsear_hora_jornada(_df3b[col_hora_ini])
            _df3b["_FIN_DT"] = _parsear_hora_jornada(_df3b[col_hora_fin])
            _df3b = _df3b[_df3b["_INI_DT"].notna() & _df3b["_FIN_DT"].notna()].copy()

            if _df3b.empty:
                st.info(
                    "No se pudieron parsear los datos de HORA INICIO / HORA FINAL. "
                    "Verifica que el formato sea HH:MM:SS a.m./p.m."
                )
            else:
                _df3b["_INI_H"] = (
                    _df3b["_INI_DT"].dt.hour + _df3b["_INI_DT"].dt.minute / 60
                )
                _df3b["_FIN_H"] = (
                    _df3b["_FIN_DT"].dt.hour + _df3b["_FIN_DT"].dt.minute / 60
                )
                _df3b["_HORAS"] = (_df3b["_FIN_H"] - _df3b["_INI_H"]).clip(
                    lower=0, upper=16
                )
                _df3b["_OP"] = _df3b[col_nombre].astype(str).str.upper().str.strip()

                # Resumen diario por operador: inicio más temprano y fin más tarde del día
                _diario3b = (
                    _df3b.groupby(["_OP", "FECHA_DT"])
                    .agg(
                        INI_DIA=("_INI_H", "min"),
                        FIN_DIA=("_FIN_H", "max"),
                    )
                    .reset_index()
                )
                _diario3b["HORAS_DIA"] = (
                    _diario3b["FIN_DIA"] - _diario3b["INI_DIA"]
                ).clip(lower=0)

                # Resumen por operador
                _jornada = (
                    _diario3b.groupby("_OP")
                    .agg(
                        INI_PROM=("INI_DIA", "mean"),
                        FIN_PROM=("FIN_DIA", "mean"),
                        HORAS_PROM=("HORAS_DIA", "mean"),
                        DIAS=("FECHA_DT", "count"),
                    )
                    .reset_index()
                )
                _jornada.columns = [
                    "OPERADOR",
                    "INI_DEC",
                    "FIN_DEC",
                    "HORAS_PROM",
                    "DÍAS CON REGISTRO",
                ]
                _jornada["HORA INICIO PROM."] = _jornada["INI_DEC"].apply(_dec_a_str)
                _jornada["HORA FIN PROM."] = _jornada["FIN_DEC"].apply(_dec_a_str)
                _jornada["HORAS ACTIVAS PROM."] = _jornada["HORAS_PROM"].round(1)
                _jornada = _jornada.sort_values("INI_DEC", ascending=False)

                st.dataframe(
                    _jornada[
                        [
                            "OPERADOR",
                            "HORA INICIO PROM.",
                            "HORA FIN PROM.",
                            "HORAS ACTIVAS PROM.",
                            "DÍAS CON REGISTRO",
                        ]
                    ],
                    column_config={
                        "HORA INICIO PROM.": st.column_config.TextColumn(
                            "HORA INICIO PROM.",
                            help="Hora promedio del primer registro del día. Ordenado de más tarde a más temprano.",
                        ),
                        "HORA FIN PROM.": st.column_config.TextColumn(
                            "HORA FIN PROM.",
                            help="Hora promedio del último registro del día.",
                        ),
                        "HORAS ACTIVAS PROM.": st.column_config.NumberColumn(
                            "HORAS ACTIVAS PROM.",
                            format="%.1f h",
                            help="Promedio de horas entre el primer y último registro del día.",
                        ),
                        "DÍAS CON REGISTRO": st.column_config.NumberColumn(
                            "DÍAS CON REGISTRO",
                            format="%d días",
                            help="Días en el historial donde el colaborador tuvo al menos un registro.",
                        ),
                    },
                    hide_index=True,
                    use_container_width=True,
                )

                # Gráfico de distribución de horas de inicio
                _dist = _diario3b["INI_DIA"].dropna()
                if len(_dist) > 0:
                    st.markdown("##### Distribución de Horas de Arranque")
                    st.caption(
                        "Frecuencia de registros según la hora de inicio — concentración ideal alrededor de las 8:00 AM."
                    )
                    _bins = [i * 0.5 for i in range(14, 25)]  # 7:00 a 12:00
                    _labels = [_dec_a_str(b) for b in _bins[:-1]]
                    _hist_vals, _ = np.histogram(_dist, bins=_bins)
                    _fig3b = go.Figure(
                        go.Bar(
                            x=_labels,
                            y=_hist_vals,
                            marker_color=[
                                (
                                    "#27ae60"
                                    if b <= 8.25
                                    else ("#f39c12" if b <= 9.0 else "#e74c3c")
                                )
                                for b in _bins[:-1]
                            ],
                            text=_hist_vals,
                            textposition="auto",
                            textfont=dict(color="white", weight="bold"),
                        )
                    )
                    _fig3b.update_layout(
                        height=280,
                        xaxis=dict(title="Hora de Inicio", showgrid=False),
                        yaxis=dict(title="Cantidad de Registros", showgrid=False),
                        margin=dict(t=20, b=10, l=10, r=10),
                    )
                    st.plotly_chart(_fig3b, use_container_width=True)
                    st.caption(
                        "🟢 Antes de las 8:15  |  🟡 8:15 – 9:00  |  🔴 Después de las 9:00"
                    )

            st.divider()

        elif not (col_hora_ini and col_hora_fin):
            _falt3b = []
            if not col_hora_ini:
                _falt3b.append("HORA INICIO")
            if not col_hora_fin:
                _falt3b.append("HORA FINAL")
            st.info(
                f"ℹ️ Fase 3B requiere columnas en REGISTRO: **{', '.join(_falt3b)}**."
            )

    with tab_colaboradores:
        # ── Consulta individual ───────────────────────────────────────────
        st.subheader("🔍 Consulta Individual de Operador")
        lista_ops = sorted(df_pivot["OPERADOR"].unique().tolist())
        if not lista_ops:
            st.info("No hay operadores disponibles en el periodo seleccionado.")
        else:
            op_sel = st.selectbox(
                "Selecciona un operador:", lista_ops, key="prod_op_sel"
            )
            if op_sel:
                df_op = (
                    df_pivot[df_pivot["OPERADOR"] == op_sel]
                    .sort_values("FECHA_DT")
                    .copy()
                )
                prom_op = df_op["EFICIENCIA_DIARIA"].mean()
                bono_op = calcular_bono(prom_op)
                dias_trabajados = int((df_op["EFICIENCIA_DIARIA"] > 0).sum())
                dias_activos_tot = len(df_op)
                corona = " 🔥" if round(prom_op, 1) >= 100.0 else ""
                sem_color, sem_emoji, sem_texto = _semaforo_prod(prom_op)

                st.markdown(
                    f"""
                <div style='background:white; border-left:8px solid {sem_color}; padding:22px 28px; border-radius:12px; margin-bottom:20px; box-shadow:0 4px 14px rgba(0,0,0,0.07);'>
                  <h3 style='margin:0; color:#2c3e50;'>{sem_emoji} {op_sel}{corona}</h3>
                  <div style='color:{sem_color}; font-weight:700; font-size:15px; margin-top:4px;'>{sem_texto}</div>
                  <div style='display:flex; gap:40px; margin-top:18px; flex-wrap:wrap;'>
                    <div><div style='color:#7f8c8d; font-size:11px; text-transform:uppercase; letter-spacing:1px;'>PRODUCTIVIDAD PROMEDIO</div>
                         <div style='font-size:36px; font-weight:900; color:{sem_color};'>{prom_op:.1f}%</div></div>
                    <div><div style='color:#7f8c8d; font-size:11px; text-transform:uppercase; letter-spacing:1px;'>PUNTOS DE BONO</div>
                         <div style='font-size:36px; font-weight:900; color:#e67e22;'>{bono_op:.1f} pts</div></div>
                    <div><div style='color:#7f8c8d; font-size:11px; text-transform:uppercase; letter-spacing:1px;'>DÍAS CON REGISTRO</div>
                         <div style='font-size:36px; font-weight:900; color:#2c3e50;'>{dias_trabajados} / {dias_activos_tot}</div></div>
                  </div>
                </div>
                """,
                    unsafe_allow_html=True,
                )

                # Velocidad de recuperación
                if len(df_op) >= 4:
                    _x_vel = list(range(len(df_op)))
                    _y_vel = df_op["EFICIENCIA_DIARIA"].values
                    _slope = float(np.polyfit(_x_vel, _y_vel, 1)[0])
                    if abs(_slope) >= 1.5:
                        _t_emoji = "↗️" if _slope > 0 else "↘️"
                        _t_txt = (
                            f"Mejorando {_slope:+.1f}%/día"
                            if _slope > 0
                            else f"Deterioro {_slope:+.1f}%/día"
                        )
                        _t_color = "#27ae60" if _slope > 0 else "#e74c3c"
                    else:
                        _t_emoji, _t_txt, _t_color = (
                            "➡️",
                            "Tendencia estable en el periodo",
                            "#7f8c8d",
                        )
                    st.markdown(
                        f"<div style='background:{_t_color}15; border-left:4px solid {_t_color}; "
                        f"padding:8px 14px; border-radius:6px; margin-bottom:16px;'>"
                        f"<span style='color:{_t_color}; font-weight:700;'>{_t_emoji} {_t_txt}</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )

                # Bitácora diaria
                st.markdown("#### 📅 Bitácora Diaria del Periodo")
                df_op["FECHA_STR"] = df_op["FECHA_DT"].dt.strftime("%d/%m/%Y")
                df_op["DIA_SEM"] = df_op["FECHA_DT"].dt.dayofweek.map(DIAS_ES)
                df_op["SEMÁFORO"] = df_op["EFICIENCIA_DIARIA"].apply(
                    lambda v: (
                        "🟢 Óptimo"
                        if v >= 90.1
                        else ("🟡 Mejora" if v >= 70 else "🔴 Crítico")
                    )
                )

                if col_actividad:
                    df_act_op = (
                        df_periodo[
                            df_periodo[col_nombre].str.upper().str.strip() == op_sel
                        ]
                        .assign(_DT=lambda x: x["FECHA_DT"].dt.normalize())
                        .groupby("_DT")[col_actividad]
                        .apply(lambda x: " / ".join(x.dropna().unique()))
                        .reset_index()
                    )
                    df_act_op.columns = ["FECHA_DT_NRM", "ACTIVIDADES"]
                    df_op["FECHA_DT_NRM"] = df_op["FECHA_DT"].dt.normalize()
                    df_op = pd.merge(
                        df_op,
                        df_act_op,
                        left_on="FECHA_DT_NRM",
                        right_on="FECHA_DT_NRM",
                        how="left",
                    )
                    df_op["ACTIVIDADES"] = df_op["ACTIVIDADES"].fillna("—")
                    cols_show = [
                        "FECHA_STR",
                        "DIA_SEM",
                        "ACTIVIDADES",
                        "EFICIENCIA_DIARIA",
                        "SEMÁFORO",
                    ]
                    col_cfg = {
                        "FECHA_STR": "FECHA",
                        "DIA_SEM": "DÍA",
                        "ACTIVIDADES": "ACTIVIDADES",
                        "EFICIENCIA_DIARIA": st.column_config.ProgressColumn(
                            "EFICIENCIA", format="%.1f%%", min_value=0, max_value=100
                        ),
                        "SEMÁFORO": "ESTATUS",
                    }
                else:
                    cols_show = [
                        "FECHA_STR",
                        "DIA_SEM",
                        "EFICIENCIA_DIARIA",
                        "SEMÁFORO",
                    ]
                    col_cfg = {
                        "FECHA_STR": "FECHA",
                        "DIA_SEM": "DÍA",
                        "EFICIENCIA_DIARIA": st.column_config.ProgressColumn(
                            "EFICIENCIA", format="%.1f%%", min_value=0, max_value=100
                        ),
                        "SEMÁFORO": "ESTATUS",
                    }

                st.dataframe(
                    df_op[cols_show],
                    column_config=col_cfg,
                    hide_index=True,
                    use_container_width=True,
                )

                # Semáforo por actividad + recomendaciones
                df_raw_op = (
                    df_periodo[df_periodo[col_nombre].str.upper().str.strip() == op_sel]
                    .sort_values("FECHA_DT")
                    .copy()
                )
                if not df_raw_op.empty:
                    df_raw_op["_FECHA"] = df_raw_op["FECHA_DT"].dt.strftime("%d/%m/%Y")
                    df_raw_op["_DIA"] = df_raw_op["FECHA_DT"].dt.dayofweek.map(DIAS_ES)
                    df_raw_op["_ESTATUS"] = df_raw_op["EFIC_NUM"].apply(
                        lambda v: (
                            "🟢 Óptimo"
                            if v >= 90.1
                            else ("🟡 Mejora" if v >= 70 else "🔴 Crítico")
                        )
                    )

                if col_actividad and not df_raw_op.empty:
                    st.markdown("#### 🎯 Diagnóstico por Actividad")
                    act_op = (
                        df_raw_op.groupby(col_actividad)["EFIC_NUM"]
                        .mean()
                        .reset_index()
                        .sort_values("EFIC_NUM", ascending=False)
                    )
                    act_op.columns = ["ACTIVIDAD", "PROD"]
                    destacan = act_op[act_op["PROD"] >= 90.1]
                    mejorar = act_op[(act_op["PROD"] >= 70) & (act_op["PROD"] < 90.1)]
                    criticas = act_op[act_op["PROD"] < 70]

                    def _recom(nivel, actividad):
                        if nivel == "verde":
                            return f"Actividad dominada. <b>{actividad}</b> puede ser área de mentoría — considera asignarlo como apoyo a compañeros."
                        if nivel == "amarillo":
                            return f"Rendimiento aceptable en <b>{actividad}</b>. Recomienda refuerzo de técnica y supervisión periódica."
                        return f"Capacitación prioritaria en <b>{actividad}</b>. Se recomienda acompañamiento diario y práctica supervisada."

                    for _titulo, _df_cat, _nivel, _color, _emoji in [
                        (
                            "Actividades donde DESTACA",
                            destacan,
                            "verde",
                            "#27ae60",
                            "🟢",
                        ),
                        (
                            "Actividades que necesita TRABAJAR",
                            mejorar,
                            "amarillo",
                            "#f39c12",
                            "🟡",
                        ),
                        (
                            "Actividades en ESTADO CRÍTICO",
                            criticas,
                            "rojo",
                            "#e74c3c",
                            "🔴",
                        ),
                    ]:
                        if _df_cat.empty:
                            continue
                        st.markdown(f"**{_emoji} {_titulo}**")
                        for _, _row in _df_cat.iterrows():
                            _act = _row["ACTIVIDAD"]
                            _efic = _row["PROD"]
                            _ancho = min(_efic, 100)
                            st.markdown(
                                f"""
                            <div style='background:white; padding:12px 15px; border-radius:8px; border-left:5px solid {_color}; margin-bottom:8px; box-shadow:0 1px 4px rgba(0,0,0,0.07);'>
                                <div style='display:flex; justify-content:space-between; align-items:center;'>
                                    <span style='font-weight:600; font-size:14px;'>{_act}</span>
                                    <span style='color:{_color}; font-weight:800; font-size:16px;'>{_efic:.1f}%</span>
                                </div>
                                <div style='background:#eee; height:6px; border-radius:3px; margin:6px 0;'>
                                    <div style='background:{_color}; width:{_ancho}%; height:100%; border-radius:3px;'></div>
                                </div>
                                <div style='font-size:12px; color:#555; margin-top:4px;'>💡 {_recom(_nivel, _act)}</div>
                            </div>
                            """,
                                unsafe_allow_html=True,
                            )

                # Expander desglose completo
                with st.expander(
                    f"📋 Desglose completo de actividades — {op_sel}  ({len(df_raw_op)} registros individuales)",
                    expanded=False,
                ):
                    if col_actividad and not df_raw_op.empty:
                        _src = ["_FECHA", "_DIA"]
                        _dest = ["FECHA", "DÍA"]
                        if col_pieza:
                            _src.append(col_pieza)
                            _dest.append("PIEZA")
                        _src.append(col_actividad)
                        _dest.append("ACTIVIDAD")
                        if col_cantidad:
                            _src.append(col_cantidad)
                            _dest.append("CANTIDAD")
                        _src += ["EFIC_NUM", "_ESTATUS"]
                        _dest += ["EFICIENCIA", "ESTATUS"]
                        df_desglose = df_raw_op[_src].copy()
                        df_desglose.columns = _dest
                        _cfg_des = {
                            "EFICIENCIA": st.column_config.ProgressColumn(
                                "EFICIENCIA (%)",
                                format="%.1f%%",
                                min_value=0,
                                max_value=100,
                            )
                        }
                        if col_cantidad:
                            _cfg_des["CANTIDAD"] = st.column_config.NumberColumn(
                                "CANTIDAD HECHA", format="%d pzs"
                            )
                        st.caption(
                            "Cada fila es un registro individual — con pieza y cantidad el operador puede validar de inmediato si los datos son correctos."
                        )
                        st.dataframe(
                            df_desglose,
                            column_config=_cfg_des,
                            hide_index=True,
                            use_container_width=True,
                        )
                    elif df_raw_op.empty:
                        st.info(
                            "No hay registros individuales para este operador en el periodo."
                        )
                    else:
                        st.info(
                            "No se detectó columna de ACTIVIDAD en la hoja REGISTRO."
                        )

                # Excel del operador
                st.divider()
                st.markdown("##### 📥 Exportar reporte individual")

                def _generar_excel():
                    df_bita_xl = df_op[
                        ["FECHA_STR", "DIA_SEM", "EFICIENCIA_DIARIA", "SEMÁFORO"]
                    ].copy()
                    df_bita_xl.columns = ["FECHA", "DÍA", "EFICIENCIA (%)", "ESTATUS"]
                    df_bita_xl["EFICIENCIA (%)"] = df_bita_xl["EFICIENCIA (%)"].round(1)
                    df_resumen_xl = pd.DataFrame(
                        {
                            "CONCEPTO": [
                                "OPERADOR",
                                "PERIODO INICIO",
                                "PERIODO FIN",
                                "PRODUCTIVIDAD PROMEDIO",
                                "PUNTOS DE BONO  (Acuerdo NSG-RH-AC-002)",
                                "DÍAS CON REGISTRO",
                                "DÍAS ACTIVOS EN PLANTA",
                                "PORCENTAJE DE ASISTENCIA",
                                "SEMÁFORO DE RENDIMIENTO",
                            ],
                            "VALOR": [
                                op_sel,
                                f_ini.strftime("%d/%m/%Y"),
                                f_fin.strftime("%d/%m/%Y"),
                                f"{prom_op:.1f}%",
                                f"{bono_op:.1f} / 40 pts",
                                dias_trabajados,
                                dias_activos_tot,
                                f"{(dias_trabajados / dias_activos_tot * 100) if dias_activos_tot > 0 else 0:.1f}%",
                                sem_texto,
                            ],
                        }
                    )
                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        df_resumen_xl.to_excel(
                            writer, sheet_name="RESUMEN", index=False
                        )
                        df_bita_xl.to_excel(
                            writer, sheet_name="BITÁCORA DIARIA", index=False
                        )
                        if (
                            col_actividad
                            and not df_raw_op.empty
                            and "_FECHA" in df_raw_op.columns
                        ):
                            _src_xl = ["_FECHA", "_DIA"]
                            _dest_xl = ["FECHA", "DÍA"]
                            if col_pieza:
                                _src_xl.append(col_pieza)
                                _dest_xl.append("PIEZA")
                            _src_xl.append(col_actividad)
                            _dest_xl.append("ACTIVIDAD")
                            if col_cantidad:
                                _src_xl.append(col_cantidad)
                                _dest_xl.append("CANTIDAD")
                            _src_xl += ["EFIC_NUM", "_ESTATUS"]
                            _dest_xl += ["EFICIENCIA (%)", "ESTATUS"]
                            df_act_xl = df_raw_op[_src_xl].copy()
                            df_act_xl.columns = _dest_xl
                            df_act_xl["EFICIENCIA (%)"] = df_act_xl[
                                "EFICIENCIA (%)"
                            ].round(1)
                            df_act_xl.to_excel(
                                writer, sheet_name="DETALLE ACTIVIDADES", index=False
                            )
                        if col_actividad and not df_raw_op.empty:
                            _act_diag = (
                                df_raw_op.groupby(col_actividad)["EFIC_NUM"]
                                .mean()
                                .reset_index()
                                .sort_values("EFIC_NUM", ascending=False)
                            )
                            _act_diag.columns = ["ACTIVIDAD", "EFICIENCIA (%)"]
                            _act_diag["EFICIENCIA (%)"] = _act_diag[
                                "EFICIENCIA (%)"
                            ].round(1)
                            _act_diag["CLASIFICACION"] = _act_diag[
                                "EFICIENCIA (%)"
                            ].apply(
                                lambda v: (
                                    "DESTACA"
                                    if v >= 90.1
                                    else ("NECESITA TRABAJAR" if v >= 70 else "CRITICO")
                                )
                            )
                            _act_diag["RECOMENDACION"] = _act_diag.apply(
                                lambda r: (
                                    f"Actividad dominada. Candidato a apoyar/mentorear en {r['ACTIVIDAD']}."
                                    if r["EFICIENCIA (%)"] >= 90.1
                                    else (
                                        f"Refuerzo tecnico y supervision periodica en {r['ACTIVIDAD']}."
                                        if r["EFICIENCIA (%)"] >= 70
                                        else f"Capacitacion prioritaria. Acompanamiento diario en {r['ACTIVIDAD']}."
                                    )
                                ),
                                axis=1,
                            )
                            _act_diag.to_excel(
                                writer,
                                sheet_name="DIAGNOSTICO ACTIVIDADES",
                                index=False,
                            )
                    buf.seek(0)
                    return buf

                nombre_xl = f"Productividad_{op_sel.replace(' ', '_')}_{f_ini.strftime('%Y%m%d')}_{f_fin.strftime('%Y%m%d')}.xlsx"
                try:
                    st.download_button(
                        label="📥 Descargar Reporte Excel",
                        data=_generar_excel(),
                        file_name=nombre_xl,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except ImportError:
                    st.warning("Instala openpyxl: `pip install openpyxl`")
                except Exception as _e_xl:
                    st.error(f"No se pudo generar el Excel: {_e_xl}")

    with tab_lideres:
        # Panel de Líderes
        try:
            st.divider()
            st.markdown("## 👔 Panel de Líderes de Área")
            st.caption(
                "Productividad calculada sobre días trabajados — las faltas no impactan al líder (asistencia es responsabilidad del colaborador). Mismo algoritmo de bono de 40 pts — Acuerdo NSG-RH-AC-002."
            )
            if not col_area:
                st.info("No se detectó columna de ÁREA en la hoja REGISTRO.")
            else:
                _lideres_resumen_xl = []
                _lideres_equipos_xl = {}
                for _lider, _areas in LIDERES_AREAS.items():
                    try:
                        _areas_norm = {normalizar_clave(a) for a in _areas}
                        _areas_str = " + ".join(_areas)
                        _df_area = df_periodo[
                            df_periodo[col_area]
                            .apply(normalizar_clave)
                            .isin(_areas_norm)
                        ].copy()
                        _df_area_ant = (
                            df_anterior[
                                df_anterior[col_area]
                                .apply(normalizar_clave)
                                .isin(_areas_norm)
                            ].copy()
                            if not df_anterior.empty
                            else pd.DataFrame()
                        )
                        if _df_area.empty:
                            st.warning(
                                f"Sin registros para **{_lider}** (Área: {_areas_str}) en este periodo."
                            )
                            continue
                        _piv = _construir_pivot_rrhh(_df_area, col_nombre, "EFIC_NUM")
                        # Excluir días con cero (faltas) — el líder no es responsable de asistencias
                        _piv_sf = _piv[_piv["EFICIENCIA_DIARIA"] > 0]
                        _prod_lider = (
                            _piv_sf.groupby("OPERADOR")["EFICIENCIA_DIARIA"]
                            .mean()
                            .mean()
                            if not _piv_sf.empty
                            else 0.0
                        )
                        _bono_lider = calcular_bono(_prod_lider)
                        _n_ops = _piv["OPERADOR"].nunique()
                        _prod_ant_l = None
                        if not _df_area_ant.empty:
                            _piv_ant = _construir_pivot_rrhh(
                                _df_area_ant, col_nombre, "EFIC_NUM"
                            )
                            if not _piv_ant.empty:
                                _piv_ant_sf = _piv_ant[
                                    _piv_ant["EFICIENCIA_DIARIA"] > 0
                                ]
                                _prod_ant_l = (
                                    _piv_ant_sf.groupby("OPERADOR")["EFICIENCIA_DIARIA"]
                                    .mean()
                                    .mean()
                                    if not _piv_ant_sf.empty
                                    else None
                                )
                        if _prod_ant_l is not None:
                            _delta_l = _prod_lider - _prod_ant_l
                            _delta_l_txt = f"{'▲' if _delta_l >= 0 else '▼'} {abs(_delta_l):.1f}% vs periodo ant."
                            _delta_l_color = "#27ae60" if _delta_l >= 0 else "#e74c3c"
                            _delta_l_xl = f"{_delta_l:+.1f}%"
                        else:
                            _delta_l_txt, _delta_l_color, _delta_l_xl = (
                                "Sin periodo anterior",
                                "#95a5a6",
                                "N/A",
                            )
                        _sl_color, _sl_emoji, _sl_texto = _semaforo_prod(_prod_lider)
                        _corona_l = " 🔥" if round(_prod_lider, 1) >= 100.0 else ""
                        _lideres_resumen_xl.append(
                            {
                                "LÍDER": _lider,
                                "ÁREA(S)": _areas_str,
                                "PRODUCTIVIDAD (%)": round(_prod_lider, 1),
                                "BONO (pts)": round(_bono_lider, 1),
                                "OPERADORES": _n_ops,
                                "CLASIFICACION": (
                                    "OPTIMO"
                                    if _prod_lider >= 90.1
                                    else ("MEJORA" if _prod_lider >= 70 else "CRITICO")
                                ),
                                "DELTA VS PERIODO ANT.": _delta_l_xl,
                                "PERIODO": f"{f_ini.strftime('%d/%m/%Y')} — {f_fin.strftime('%d/%m/%Y')}",
                            }
                        )
                        _equipo_base = (
                            _piv_sf.groupby("OPERADOR")["EFICIENCIA_DIARIA"]
                            .mean()
                            .reset_index()
                            .sort_values("EFICIENCIA_DIARIA", ascending=False)
                        )
                        _equipo_base["BONO (pts)"] = (
                            _equipo_base["EFICIENCIA_DIARIA"]
                            .apply(calcular_bono)
                            .round(1)
                        )
                        _equipo_base["EFIC (%)"] = _equipo_base[
                            "EFICIENCIA_DIARIA"
                        ].round(1)
                        _equipo_base["CLASIFICACION"] = _equipo_base[
                            "EFICIENCIA_DIARIA"
                        ].apply(
                            lambda v: (
                                "OPTIMO"
                                if v >= 90.1
                                else ("MEJORA" if v >= 70 else "CRITICO")
                            )
                        )
                        _sheet_name = re.sub(r"[\\/*?\[\]:]", "-", _areas_str)[:28]
                        _lideres_equipos_xl[_sheet_name] = _equipo_base[
                            ["OPERADOR", "EFIC (%)", "BONO (pts)", "CLASIFICACION"]
                        ]
                        st.markdown(
                            f"""
                        <div style='background:white; border-left:8px solid {_sl_color}; padding:22px 28px; border-radius:12px; margin-bottom:12px; box-shadow:0 4px 14px rgba(0,0,0,0.07);'>
                            <div style='display:flex; justify-content:space-between; align-items:flex-start; flex-wrap:wrap; gap:10px;'>
                                <div>
                                    <h4 style='margin:0; color:#2c3e50;'>{_sl_emoji} {_lider}{_corona_l}</h4>
                                    <div style='color:#7f8c8d; font-size:12px; margin-top:3px;'>ÁREA BAJO SU GESTIÓN: <b>{_areas_str}</b></div>
                                    <div style='color:{_sl_color}; font-weight:700; font-size:14px; margin-top:5px;'>{_sl_texto}</div>
                                </div>
                                <div style='text-align:right;'>
                                    <div style='font-size:42px; font-weight:900; color:{_sl_color}; line-height:1;'>{_prod_lider:.1f}%</div>
                                    <div style='color:{_delta_l_color}; font-weight:700; font-size:13px; margin-top:4px;'>{_delta_l_txt}</div>
                                </div>
                            </div>
                            <div style='display:flex; gap:40px; margin-top:18px; flex-wrap:wrap;'>
                                <div><div style='color:#7f8c8d; font-size:11px; text-transform:uppercase; letter-spacing:1px;'>PUNTOS DE BONO</div>
                                     <div style='font-size:28px; font-weight:900; color:#e67e22;'>{_bono_lider:.1f} / 40 pts</div></div>
                                <div><div style='color:#7f8c8d; font-size:11px; text-transform:uppercase; letter-spacing:1px;'>OPERADORES EN ÁREA</div>
                                     <div style='font-size:28px; font-weight:900; color:#2c3e50;'>{_n_ops}</div></div>
                            </div>
                        </div>
                        """,
                            unsafe_allow_html=True,
                        )
                        with st.expander(
                            f"👥 Ranking del equipo de {_lider}  ({_n_ops} operadores)",
                            expanded=False,
                        ):
                            _equipo_show = _equipo_base.copy()
                            _equipo_show["SEMÁFORO"] = _equipo_show[
                                "EFICIENCIA_DIARIA"
                            ].apply(
                                lambda v: (
                                    "🟢 Óptimo"
                                    if v >= 90.1
                                    else ("🟡 Mejora" if v >= 70 else "🔴 Crítico")
                                )
                            )
                            st.dataframe(
                                _equipo_show[
                                    [
                                        "OPERADOR",
                                        "EFICIENCIA_DIARIA",
                                        "BONO (pts)",
                                        "SEMÁFORO",
                                    ]
                                ],
                                column_config={
                                    "OPERADOR": "NOMBRE",
                                    "EFICIENCIA_DIARIA": st.column_config.ProgressColumn(
                                        "PRODUCTIVIDAD",
                                        format="%.1f%%",
                                        min_value=0,
                                        max_value=100,
                                    ),
                                    "BONO (pts)": st.column_config.NumberColumn(
                                        "BONO (pts)", format="%.1f"
                                    ),
                                    "SEMÁFORO": "ESTATUS",
                                },
                                hide_index=True,
                                use_container_width=True,
                            )
                            if len(_piv["FECHA_DT"].unique()) > 1:
                                _tend_area = (
                                    _piv.groupby("FECHA_DT")["EFICIENCIA_DIARIA"]
                                    .mean()
                                    .reset_index()
                                )
                                _tend_area["_VIS"] = _tend_area[
                                    "EFICIENCIA_DIARIA"
                                ].clip(upper=100)
                                _fig_ldr = go.Figure()
                                _fig_ldr.add_trace(
                                    go.Scatter(
                                        x=_tend_area["FECHA_DT"],
                                        y=_tend_area["_VIS"],
                                        mode="lines+markers+text",
                                        text=[
                                            f"{v:.0f}%"
                                            for v in _tend_area["EFICIENCIA_DIARIA"]
                                        ],
                                        textposition="top center",
                                        textfont=dict(size=11, color="black"),
                                        line=dict(color="#2c3e50", width=2),
                                        marker=dict(
                                            size=10,
                                            color=[
                                                (
                                                    "#27ae60"
                                                    if v >= 90.1
                                                    else (
                                                        "#f39c12"
                                                        if v >= 70
                                                        else "#e74c3c"
                                                    )
                                                )
                                                for v in _tend_area["EFICIENCIA_DIARIA"]
                                            ],
                                            line=dict(width=2, color="white"),
                                        ),
                                    )
                                )
                                _fig_ldr.add_hline(
                                    y=90.1,
                                    line_dash="dash",
                                    line_color="#27ae60",
                                    line_width=1,
                                    annotation_text="Meta 90.1%",
                                    annotation_font=dict(size=10, color="#27ae60"),
                                )
                                _fig_ldr.update_layout(
                                    height=250,
                                    yaxis=dict(
                                        range=[0, 115], showgrid=False, title=""
                                    ),
                                    xaxis=dict(showgrid=False),
                                    margin=dict(t=15, b=10, l=10, r=10),
                                    title=dict(
                                        text="Tendencia diaria del área",
                                        font=dict(size=13),
                                    ),
                                )
                                st.plotly_chart(_fig_ldr, use_container_width=True)
                    except Exception as _err_lider:
                        st.error(f"Error al procesar datos de {_lider}: {_err_lider}")

                if _lideres_resumen_xl:
                    st.divider()
                    st.markdown("##### 📥 Exportar reporte de líderes")

                    def _generar_excel_lideres():
                        _buf = io.BytesIO()
                        with pd.ExcelWriter(_buf, engine="openpyxl") as _wr:
                            pd.DataFrame(_lideres_resumen_xl).to_excel(
                                _wr, sheet_name="RESUMEN LIDERES", index=False
                            )
                            for _sn, _df_eq in _lideres_equipos_xl.items():
                                _df_eq.to_excel(_wr, sheet_name=_sn, index=False)
                        _buf.seek(0)
                        return _buf

                    _nombre_xl_ldr = f"Lideres_NSG_{f_ini.strftime('%Y%m%d')}_{f_fin.strftime('%Y%m%d')}.xlsx"
                    try:
                        st.download_button(
                            label="📥 Descargar Reporte de Líderes (Excel)",
                            data=_generar_excel_lideres(),
                            file_name=_nombre_xl_ldr,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                    except ImportError:
                        st.warning("Instala openpyxl: `pip install openpyxl`")
                    except Exception as _e_xl_ldr:
                        st.error(f"No se pudo generar el Excel de líderes: {_e_xl_ldr}")

        except Exception as _err_panel_ldr:
            st.error(f"Error en el Panel de Líderes: {_err_panel_ldr}")

        # ── Comparativo entre líderes ─────────────────────────────────
        if col_area and not df_periodo.empty:
            st.divider()
            st.subheader("📊 Comparativo entre Líderes")
            st.caption(
                "Desempeño consolidado de cada líder vs. la meta óptima del área."
            )
            _comp_rows = []
            for _ldr, _ldr_areas in LIDERES_AREAS.items():
                _ldr_areas_norm = {normalizar_clave(a) for a in _ldr_areas}
                _df_ldr = df_periodo[
                    df_periodo[col_area].apply(normalizar_clave).isin(_ldr_areas_norm)
                ].copy()
                if _df_ldr.empty:
                    continue
                _piv_ldr = _construir_pivot_rrhh(_df_ldr, col_nombre, "EFIC_NUM")
                if _piv_ldr.empty:
                    continue
                _prod_ldr = round(_piv_ldr["EFICIENCIA_DIARIA"].mean(), 1)
                _bono_ldr = round(calcular_bono(_prod_ldr), 1)
                _ops_ldr = _piv_ldr["OPERADOR"].nunique()
                _sem_col, _sem_emo, _sem_txt = _semaforo_prod(_prod_ldr)
                _comp_rows.append(
                    {
                        "LÍDER": _ldr,
                        "ÁREA": " + ".join(_ldr_areas),
                        "PRODUCTIVIDAD (%)": _prod_ldr,
                        "BONO (pts)": _bono_ldr,
                        "OPERADORES": _ops_ldr,
                        "SEMÁFORO": f"{_sem_emo} {_sem_txt}",
                    }
                )
            if _comp_rows:
                _df_comp = pd.DataFrame(_comp_rows).sort_values(
                    "PRODUCTIVIDAD (%)", ascending=False
                )
                st.dataframe(
                    _df_comp,
                    column_config={
                        "PRODUCTIVIDAD (%)": st.column_config.ProgressColumn(
                            "PRODUCTIVIDAD (%)",
                            format="%.1f%%",
                            min_value=0,
                            max_value=100,
                            help="Productividad promedio del equipo del líder en el periodo.",
                        ),
                        "BONO (pts)": st.column_config.NumberColumn(
                            "BONO (pts)",
                            format="%.1f pts",
                            help="Puntos de bono calculados sobre el promedio del equipo — Acuerdo NSG-RH-AC-002.",
                        ),
                        "OPERADORES": st.column_config.NumberColumn(
                            "OPERADORES",
                            format="%d",
                            help="Número de operadores únicos con registro en el periodo.",
                        ),
                        "SEMÁFORO": st.column_config.TextColumn(
                            "ESTATUS",
                            help="🟢 Óptimo ≥ 90.1% | 🟡 Mejora 70–90% | 🔴 Crítico < 70%",
                        ),
                    },
                    hide_index=True,
                    use_container_width=True,
                )

    # ── PLANEACIÓN DE TURNO ───────────────────────────────────────────────
    with tab_planeacion:
        st.markdown("## 📋 Planeación de Turno")
        st.caption("Herramientas operativas para el titular de producción.")

        def _abrev_op(nombre_completo):
            """'JUAN CARLOS PÉREZ RUIZ' → 'Juan P.' para evitar ambigüedad en nombres repetidos."""
            _p = nombre_completo.split()
            if len(_p) >= 2:
                return f"{_p[0].title()} {_p[1][0].upper()}."
            return _p[0].title() if _p else nombre_completo

        # ── Buscador de Cobertura ─────────────────────────────────────────
        st.subheader("🆘 Buscador de Cobertura")
        st.caption("¿Quién puede cubrir una posición ausente en el turno de hoy?")

        if df_historico_all.empty or not col_actividad:
            st.info(
                "Se necesitan datos históricos y la columna ACTIVIDAD para usar esta herramienta."
            )
        else:
            _acts_cob = sorted(
                df_historico_all[col_actividad]
                .astype(str)
                .str.strip()
                .str.title()
                .unique()
                .tolist()
            )
            _piezas_cob = (
                ["TODAS LAS PIEZAS"]
                + sorted(
                    df_historico_all[col_pieza]
                    .astype(str)
                    .str.strip()
                    .unique()
                    .tolist()
                )
                if col_pieza
                else ["TODAS LAS PIEZAS"]
            )
            _cc1, _cc2, _cc3 = st.columns([2, 2, 1])
            with _cc1:
                _act_cob = st.selectbox(
                    "Actividad a cubrir:", _acts_cob, key="cob_actividad"
                )
            with _cc2:
                _pieza_cob = st.selectbox(
                    "Pieza (opcional):", _piezas_cob, key="cob_pieza"
                )
            with _cc3:
                _top_n = st.number_input(
                    "Top N:",
                    min_value=1,
                    max_value=10,
                    value=3,
                    step=1,
                    key="cob_top_n",
                )

            _df_cob = df_historico_all.assign(
                _ACT=df_historico_all[col_actividad]
                .astype(str)
                .str.strip()
                .str.title(),
                _OP=df_historico_all[col_nombre].astype(str).str.upper().str.strip(),
            )
            _df_cob = _df_cob[_df_cob["_ACT"] == _act_cob]
            if col_pieza and _pieza_cob != "TODAS LAS PIEZAS":
                _df_cob = _df_cob[
                    _df_cob[col_pieza].astype(str).str.strip() == _pieza_cob
                ]

            if _df_cob.empty:
                _txt_pz = (
                    f" en **{_pieza_cob}**" if _pieza_cob != "TODAS LAS PIEZAS" else ""
                )
                st.warning(
                    f"Sin registros históricos de **{_act_cob}**{_txt_pz} en colaboradores activos."
                )
            else:
                _grp_cob = (
                    _df_cob.groupby("_OP")
                    .agg(
                        EFIC_PROM=("EFIC_NUM", "median"),
                        ULTIMA_VEZ=("FECHA_DT", "max"),
                        REGISTROS=("EFIC_NUM", "count"),
                    )
                    .reset_index()
                    .sort_values("EFIC_PROM", ascending=False)
                    .head(int(_top_n))
                )
                _medallas = ["🥇", "🥈", "🥉"] + [f"#{i+4}" for i in range(7)]
                for _i, (_, _row) in enumerate(_grp_cob.iterrows()):
                    _sc = (
                        "#27ae60"
                        if _row["EFIC_PROM"] >= 90.1
                        else ("#f39c12" if _row["EFIC_PROM"] >= 70 else "#e74c3c")
                    )
                    _dias = (ahora_local().date() - _row["ULTIMA_VEZ"].date()).days
                    _ulti = "Hoy" if _dias == 0 else f"Hace {_dias} día(s)"
                    st.markdown(
                        f"<div style='background:white; border-left:6px solid {_sc}; padding:14px 20px; "
                        f"border-radius:10px; margin-bottom:10px; box-shadow:0 2px 8px rgba(0,0,0,0.06); "
                        f"display:flex; justify-content:space-between; align-items:center;'>"
                        f"<div><span style='font-size:20px;'>{_medallas[_i]}</span>"
                        f"<span style='font-weight:700; font-size:16px; margin-left:8px;'>{_row['_OP']}</span>"
                        f"<span style='color:#95a5a6; font-size:12px; margin-left:12px;'>"
                        f"{_ulti} · {int(_row['REGISTROS'])} registros</span></div>"
                        f"<div style='color:{_sc}; font-weight:900; font-size:24px;'>{_row['EFIC_PROM']:.1f}%</div></div>",
                        unsafe_allow_html=True,
                    )

        st.divider()

        # ── Alineación Sugerida del Día ───────────────────────────────────
        st.subheader("📅 Alineación Sugerida del Día")
        st.caption(
            "Programa de hoy → subprocesos de la BDD → top operadores por mediana de eficiencia histórica · área · pieza · subproceso."
        )

        _req = [k for k in ["fecha", "pieza", "area"] if not col_prog.get(k)]
        _req_bdd = [k for k in ["pieza", "proceso", "subproceso"] if not col_bdd.get(k)]
        if _req:
            st.info(f"Columnas faltantes en PROGRAMA: {', '.join(_req)}.")
        elif _req_bdd:
            st.info(f"Columnas faltantes en BDD: {', '.join(_req_bdd)}.")
        elif df_programa.empty or df_bdd.empty:
            st.info("El programa de producción o la BDD están vacíos.")
        elif not col_actividad or not col_pieza or df_historico_all.empty:
            st.info(
                "Se necesitan datos históricos con columnas PIEZA y ACTIVIDAD para generar la alineación."
            )
        else:
            _hoy_alin = ahora_local().date().strftime("%d/%m/%Y")
            _prog_hoy = df_programa[df_programa[col_prog["fecha"]] == _hoy_alin].copy()

            if _prog_hoy.empty:
                st.info(f"No hay piezas programadas para hoy ({_hoy_alin}).")
            else:
                if col_prog.get("total"):
                    _prog_hoy["_CANT"] = convertir_serie_numerica(
                        _prog_hoy[col_prog["total"]]
                    ).fillna(0)
                else:
                    _prog_hoy["_CANT"] = 0
                _prog_hoy = _prog_hoy[_prog_hoy["_CANT"] > 0].copy()

                # Lookup: pieza+actividad → top operadores (incluye líderes)
                _hist = df_historico_all.assign(
                    _PZ=df_historico_all[col_pieza].astype(str).str.strip(),
                    _ACT=df_historico_all[col_actividad]
                    .astype(str)
                    .str.strip()
                    .str.title(),
                    _OP=df_historico_all[col_nombre]
                    .astype(str)
                    .str.upper()
                    .str.strip(),
                )

                _alin_rows = []
                for _, _rp in _prog_hoy.iterrows():
                    _pz = str(_rp[col_prog["pieza"]]).strip()
                    _ar = str(_rp[col_prog["area"]]).strip()
                    _qt = int(_rp["_CANT"])

                    # Subprocesos de esta pieza+área desde la BDD
                    _subs = (
                        df_bdd[
                            (df_bdd[col_bdd["pieza"]].astype(str).str.strip() == _pz)
                            & (
                                df_bdd[col_bdd["proceso"]].astype(str).str.strip()
                                == _ar
                            )
                        ][col_bdd["subproceso"]]
                        .astype(str)
                        .str.strip()
                        .str.title()
                        .unique()
                    )

                    if len(_subs) == 0:
                        # Sin subprocesos en BDD: fila genérica por pieza
                        _df_gen = (
                            _hist[_hist["_PZ"] == _pz]
                            .groupby("_OP")["EFIC_NUM"]
                            .median()
                            .reset_index()
                            .sort_values("EFIC_NUM", ascending=False)
                            .head(3)
                        )
                        _mds = ["🥇", "🥈", "🥉"]
                        _sugs = (
                            "   ".join(
                                f"{_mds[_j]} {_abrev_op(r['_OP'])} {r['EFIC_NUM']:.0f}%"
                                for _j, (_, r) in enumerate(_df_gen.iterrows())
                            )
                            if not _df_gen.empty
                            else "Sin historial"
                        )
                        _alin_rows.append(
                            {
                                "ÁREA": _ar,
                                "PIEZA": _pz,
                                "CANTIDAD": _qt,
                                "SUBPROCESO": "—",
                                "TOP OPERADORES": _sugs,
                            }
                        )
                    else:
                        for _sub in _subs:
                            _df_sub = (
                                _hist[(_hist["_PZ"] == _pz) & (_hist["_ACT"] == _sub)]
                                .groupby("_OP")["EFIC_NUM"]
                                .median()
                                .reset_index()
                                .sort_values("EFIC_NUM", ascending=False)
                                .head(3)
                            )
                            _mds = ["🥇", "🥈", "🥉"]
                            _sugs = (
                                "   ".join(
                                    f"{_mds[_j]} {_abrev_op(r['_OP'])} {r['EFIC_NUM']:.0f}%"
                                    for _j, (_, r) in enumerate(_df_sub.iterrows())
                                )
                                if not _df_sub.empty
                                else "Sin historial"
                            )
                            _alin_rows.append(
                                {
                                    "ÁREA": _ar,
                                    "PIEZA": _pz,
                                    "CANTIDAD": _qt,
                                    "SUBPROCESO": _sub,
                                    "TOP OPERADORES": _sugs,
                                }
                            )

                if _alin_rows:
                    _df_alin = pd.DataFrame(_alin_rows).sort_values(
                        ["ÁREA", "PIEZA", "SUBPROCESO"]
                    )
                    st.dataframe(
                        _df_alin,
                        column_config={
                            "CANTIDAD": st.column_config.NumberColumn(
                                "CANTIDAD",
                                format="%d pzs",
                                help="Piezas programadas para hoy.",
                            ),
                            "SUBPROCESO": st.column_config.TextColumn(
                                "SUBPROCESO",
                                help="Operación requerida según la BDD para esta pieza y área.",
                            ),
                            "TOP OPERADORES": st.column_config.TextColumn(
                                "TOP OPERADORES (historial 30d)",
                                help="Operadores activos (últimos 30 días) rankeados por eficiencia mediana histórica en esa pieza + subproceso. La asignación final corresponde al líder.",
                            ),
                        },
                        hide_index=True,
                        use_container_width=True,
                    )
                    st.caption(
                        "💡 Basado en el programa de hoy, subprocesos de la BDD e historial de los últimos 30 días. La asignación final corresponde al líder de área."
                    )
                else:
                    st.info(
                        "No se generaron filas de alineación — verifica que las piezas del programa coincidan con la BDD."
                    )


# ============================================================
# CAPTURA — COMPONENTES NUEVOS
# ============================================================


def obtener_areas_pendientes_corte(
    df_programa,
    col_prog,
    df_auditorias,
    df_bdd,
    col_bdd,
    fecha_sel,
    corte_sel,
    todas_areas,
    area_completada,
):
    pendientes = []
    for _ar in todas_areas:
        if _ar == area_completada:
            continue
        _df_plan_ar = obtener_plan_del_dia(df_programa, col_prog, fecha_sel, _ar)
        if _df_plan_ar.empty:
            continue
        _df_aud_ar, _col_aud_ar = obtener_auditorias_hoy(df_auditorias, fecha_sel, _ar)
        _pend_ar = obtener_piezas_pendientes(
            _df_plan_ar,
            col_prog,
            df_bdd,
            col_bdd,
            _df_aud_ar,
            _col_aud_ar,
            _ar,
            corte_sel,
        )
        _caps_ar = (
            _df_aud_ar[_df_aud_ar[_col_aud_ar["corte"]] == corte_sel]
            if not _df_aud_ar.empty and _col_aud_ar.get("corte")
            else pd.DataFrame()
        )
        if _caps_ar.empty or len(_pend_ar) > 0:
            pendientes.append(_ar)
    return pendientes


def render_progreso_por_corte(
    df_aud_hoy, col_aud, df_plan_dia, col_prog, df_bdd, col_bdd, area_sel
):
    if df_plan_dia.empty:
        return
    if validar_columnas(col_prog, ["pieza", "total"]) or validar_columnas(
        col_bdd, ["pieza", "proceso", "subproceso"]
    ):
        return

    # Base pieza × subproceso con programado — misma lógica que calcular_resumen
    df_sub = df_bdd[df_bdd[col_bdd["proceso"]] == area_sel][
        [col_bdd["pieza"], col_bdd["subproceso"]]
    ].copy()
    piezas_plan = df_plan_dia[col_prog["pieza"]].unique()
    df_base = df_sub[df_sub[col_bdd["pieza"]].isin(piezas_plan)].copy()
    if df_base.empty:
        return
    df_base = pd.merge(
        df_base,
        df_plan_dia[[col_prog["pieza"], col_prog["total"]]],
        left_on=col_bdd["pieza"],
        right_on=col_prog["pieza"],
        how="left",
    )
    df_base[col_prog["total"]] = convertir_serie_numerica(
        df_base[col_prog["total"]]
    ).fillna(0)
    df_base = df_base[df_base[col_prog["total"]] > 0].copy()
    if df_base.empty:
        return

    # Función auxiliar: acumulado de eficiencia hasta un conjunto de cortes
    def _acum_pct(cortes_incluidos):
        if (
            df_aud_hoy.empty
            or not col_aud.get("corte")
            or not col_aud.get("real")
            or not col_aud.get("pieza")
            or not col_aud.get("subproceso")
        ):
            return 0.0
        _df_f = df_aud_hoy[df_aud_hoy[col_aud["corte"]].isin(cortes_incluidos)].copy()
        if _df_f.empty:
            return 0.0
        _df_f[col_aud["real"]] = pd.to_numeric(
            _df_f[col_aud["real"]], errors="coerce"
        ).fillna(0)
        _df_max = (
            _df_f.groupby([col_aud["pieza"], col_aud["subproceso"]])[col_aud["real"]]
            .max()
            .reset_index()
        )
        _df_m = pd.merge(
            df_base,
            _df_max,
            left_on=[col_bdd["pieza"], col_bdd["subproceso"]],
            right_on=[col_aud["pieza"], col_aud["subproceso"]],
            how="left",
        ).fillna({col_aud["real"]: 0})
        _real_s = pd.to_numeric(_df_m[col_aud["real"]], errors="coerce").fillna(0)
        _prog_s = convertir_serie_numerica(_df_m[col_prog["total"]]).fillna(0)
        _mask = _prog_s > 0
        if not _mask.any():
            return 0.0
        return (_real_s[_mask] / _prog_s[_mask] * 100).mean()

    # Incremento por bloque: cada barra es la aportación de ese corte al total del día
    _cortes_orden = list(CORTES_DICT.keys())
    _meta_bloque = round(100 / len(_cortes_orden), 1)  # 33.3%
    _datos = []
    _acum_ant = 0.0
    for i, _ck in enumerate(_cortes_orden):
        _acum_act = _acum_pct(_cortes_orden[: i + 1])
        _incremento = round(max(0.0, _acum_act - _acum_ant), 1)
        _datos.append(
            {
                "CORTE": _ck,
                "AVANCE": _incremento,
                "CAPTURADO": _acum_act > 0,
            }
        )
        _acum_ant = _acum_act

    if not any(d["AVANCE"] > 0 for d in _datos):
        return

    _df_g = pd.DataFrame(_datos)
    st.markdown("#### ⏱️ Avance por Corte del Turno")
    _fig = go.Figure()
    for _, _row in _df_g.iterrows():
        _val = _row["AVANCE"]
        _cap = _row["CAPTURADO"]
        if not _cap:
            _color = "#bdc3c7"
        elif _val >= _meta_bloque * 0.8:
            _color = "#27ae60"
        elif _val >= _meta_bloque * 0.6:
            _color = "#f39c12"
        else:
            _color = "#e74c3c"
        _texto = f"{_val}%" if _cap else "Sin captura"
        _fig.add_trace(
            go.Bar(
                name=_row["CORTE"],
                x=[_val],
                y=[_row["CORTE"]],
                orientation="h",
                marker_color=_color,
                text=[_texto],
                textposition="auto",
                textfont=dict(color="white", size=13, weight="bold"),
                showlegend=False,
            )
        )
    _fig.add_vline(
        x=_meta_bloque,
        line_dash="dash",
        line_color="#27ae60",
        line_width=2,
        annotation_text=f"Meta por bloque ({_meta_bloque}%)",
        annotation_font=dict(color="#27ae60", size=11),
    )
    _rango_max = max(d["AVANCE"] for d in _datos) * 1.4
    _fig.update_layout(
        height=170,
        showlegend=False,
        xaxis=dict(
            range=[0, max(_rango_max, _meta_bloque * 2)],
            title="% de aportación al programa del día",
            showgrid=False,
        ),
        yaxis=dict(showgrid=False),
        margin=dict(t=5, b=10, l=10, r=10),
    )
    st.plotly_chart(_fig, use_container_width=True)


def render_checklist_cobertura(
    df_plan_dia, col_prog, df_aud_hoy, col_aud, df_bdd, col_bdd, area_sel
):
    if df_plan_dia.empty or not col_prog.get("pieza"):
        return
    st.markdown("#### ✅ Cobertura de Auditoría — Pieza × Corte")
    _hora_actual = ahora_local().hour
    _hora_map = {"11:00 AM (3h)": 11, "14:00 PM (6h)": 14, "17:00 PM (9h)": 17}
    _piezas = df_plan_dia[col_prog["pieza"]].unique()
    _filas = []
    for _pieza in _piezas:
        _fila = {"PIEZA": _pieza}
        _subs_esperados = set()
        if not validar_columnas(col_bdd, ["pieza", "proceso", "subproceso"]):
            _subs_esperados = set(
                df_bdd[
                    (df_bdd[col_bdd["pieza"]] == _pieza)
                    & (df_bdd[col_bdd["proceso"]] == area_sel)
                ][col_bdd["subproceso"]].unique()
            )
        for _ck in CORTES_DICT.keys():
            _hora_c = _hora_map.get(_ck, 0)
            _regs = pd.DataFrame()
            if not df_aud_hoy.empty and col_aud.get("corte") and col_aud.get("pieza"):
                _regs = df_aud_hoy[
                    (df_aud_hoy[col_aud["corte"]] == _ck)
                    & (df_aud_hoy[col_aud["pieza"]] == _pieza)
                ]
            if not _regs.empty:
                if _subs_esperados and col_aud.get("subproceso"):
                    _subs_cap = set(_regs[col_aud["subproceso"]].tolist())
                    _fila[_ck] = (
                        "✅ Completo"
                        if _subs_esperados.issubset(_subs_cap)
                        else "⚠️ Parcial"
                    )
                else:
                    _fila[_ck] = "✅ Capturado"
            elif _hora_actual >= _hora_c:
                _fila[_ck] = "⏳ Pendiente"
            else:
                _fila[_ck] = "🕐 Por llegar"
        _filas.append(_fila)
    _df_check = pd.DataFrame(_filas)
    st.dataframe(_df_check, hide_index=True, use_container_width=True)


# ============================================================
# ADMIN
# ============================================================


def _leer_usuarios():
    libro = conectar_libro()
    if not libro:
        return []
    try:
        return _ejecutar_con_reintentos(
            lambda: libro.worksheet("USUARIOS").get_all_records()
        )
    except Exception as _exc:
        st.warning(f"⚠️ No se pudo leer la lista de usuarios: {_exc}")
        return []


def _guardar_usuarios(lista):
    libro = conectar_libro()
    if not libro:
        return False
    try:
        hoja = libro.worksheet("USUARIOS")
        _datos = [["USUARIO", "HASH", "ROL"]] + [
            [_u["USUARIO"], _u["HASH"], _u["ROL"]] for _u in lista
        ]
        _ejecutar_con_reintentos(lambda: hoja.clear())
        _ejecutar_con_reintentos(lambda: hoja.update("A1", _datos))
        return True
    except Exception as _exc:
        st.warning(f"⚠️ Error al guardar usuarios: {_exc}")
        return False


def render_admin():
    st.markdown("#### ⚙️ Gestión de Usuarios")
    _lista = _leer_usuarios()
    if _lista is None or (not _lista and not isinstance(_lista, list)):
        st.error("No se pudo conectar con la hoja USUARIOS.")
        return

    st.markdown("**Usuarios activos**")
    for _i, _u in enumerate(_lista):
        _ca, _cb, _cc = st.columns([3, 2, 1])
        with _ca:
            st.markdown(f"👤 **{_u['USUARIO']}**")
        with _cb:
            st.caption(_u["ROL"])
        with _cc:
            if _u["USUARIO"] != st.session_state.usuario:
                if st.button("🗑️", key=f"del_u_{_i}", help=f"Eliminar {_u['USUARIO']}"):
                    _lista.pop(_i)
                    if _guardar_usuarios(_lista):
                        st.success(f"Usuario '{_u['USUARIO']}' eliminado.")
                    else:
                        st.error("Error al guardar cambios.")
                    st.rerun()
            else:
                st.caption("(tú)")

    st.divider()
    st.markdown("**Agregar usuario**")
    with st.form("form_nuevo_usuario"):
        _nu = st.text_input("Nombre de usuario")
        _np = st.text_input("Contraseña", type="password")
        _nr = st.selectbox("Rol", ["auditor", "produccion", "admin"])
        if st.form_submit_button("Agregar", type="primary"):
            if not _nu.strip() or not _np:
                st.error("Nombre y contraseña son obligatorios.")
            elif any(_u["USUARIO"].lower() == _nu.strip().lower() for _u in _lista):
                st.error(f"El usuario '{_nu}' ya existe.")
            else:
                _lista.append(
                    {
                        "USUARIO": _nu.strip(),
                        "HASH": hashlib.sha256(_np.encode()).hexdigest(),
                        "ROL": _nr,
                    }
                )
                if _guardar_usuarios(_lista):
                    st.success(f"Usuario '{_nu}' agregado con rol '{_nr}'.")
                    st.rerun()
                else:
                    st.error("Error al guardar en Google Sheets.")


# ============================================================
# LOGIN
# ============================================================


def render_login():
    _, _lc2, _ = st.columns([1, 2, 1])
    with _lc2:
        if os.path.exists(LOGO_FILENAME):
            st.image(LOGO_FILENAME, width=130)
        st.markdown(
            "<h2 style='text-align:center;color:#8B1A1A;margin-top:8px;'>Pulso NSG</h2>"
            "<p style='text-align:center;color:#888;margin-top:-10px;'>Auditoría de Turno · Planta</p>",
            unsafe_allow_html=True,
        )
        st.divider()
        _usuario = st.text_input("Usuario", placeholder="Tu nombre de usuario")
        _password = st.text_input(
            "Contraseña", type="password", placeholder="Tu contraseña"
        )
        if st.button("Ingresar", type="primary", use_container_width=True):
            _lista = _leer_usuarios()
            if not _lista:
                st.error("Sin conexión con la base de usuarios. Intenta de nuevo.")
                return
            _hash = hashlib.sha256(_password.encode()).hexdigest()
            _match = next(
                (
                    u
                    for u in _lista
                    if u["USUARIO"].lower() == _usuario.lower() and u["HASH"] == _hash
                ),
                None,
            )
            if _match:
                st.session_state.usuario = _match["USUARIO"]
                st.session_state.rol = _match["ROL"]
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos.")


# ============================================================
# MAIN
# ============================================================


def main():
    st.set_page_config(
        layout="wide",
        page_title="Pulso NSG",
        page_icon="🛡️",
        initial_sidebar_state="collapsed",
    )
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;700&display=swap');
        html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
        .stApp { background-color: #F8F9FA; }
        .metric-card {
            background: white; padding: 20px; border-radius: 12px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.05);
            border-left: 5px solid #8B1A1A; text-align: center;
        }
        .capture-container {
            background: white; padding: 30px; border-radius: 15px;
            border: 1px solid #E0E0E0; margin-top: 20px;
            box-shadow: 0 10px 25px rgba(0,0,0,0.03);
        }
        .stButton>button {
            width: 100%; background: linear-gradient(135deg, #8B1A1A 0%, #6B1414 100%) !important;
            color: white !important; border-radius: 10px !important;
            padding: 18px !important; font-weight: 700 !important; border: none !important;
        }
        [data-testid="stSidebar"], [data-testid="collapsedControl"] { display: none !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    if "usuario" not in st.session_state:
        st.session_state.usuario = None
    if "rol" not in st.session_state:
        st.session_state.rol = None

    if not st.session_state.usuario:
        render_login()
        st.stop()

    if "form_id" not in st.session_state:
        st.session_state.form_id = 0
    if "guardando" not in st.session_state:
        st.session_state.guardando = False
    if "last_ops" not in st.session_state:
        st.session_state.last_ops = 1
    if "area_corte_completada" not in st.session_state:
        st.session_state.area_corte_completada = None
    if "nav_area" not in st.session_state:
        st.session_state.nav_area = None
    if "nav_corte" not in st.session_state:
        st.session_state.nav_corte = None

    df_programa, col_prog = preparar_dataframe("PROGRAMA", 1)
    df_bdd_raw, col_bdd = preparar_dataframe("BDD", 0)
    df_auditorias, _ = preparar_dataframe("AUDITAR", 0)
    df_bdd = filtrar_bdd_activa(df_bdd_raw, col_bdd)

    # ── Header / pleca de la app ─────────────────────────────────────────
    _logo_col, _title_col, _user_col = st.columns([1, 5, 2])
    with _logo_col:
        if os.path.exists(LOGO_FILENAME):
            st.image(LOGO_FILENAME, width=80)
    with _title_col:
        st.markdown(
            """
            <div style="
                background: linear-gradient(90deg, #6B1414 0%, #8B1A1A 60%, #6B1414 100%);
                border-radius: 10px;
                border-bottom: 3px solid #C0C0C0;
                padding: 10px 20px;
                margin-bottom: 4px;
            ">
                <div style="color:#C0C0C0; font-size:26px; font-weight:900; letter-spacing:2px; line-height:1.1;">
                    Pulso NSG
                </div>
                <div style="color:#e8e8e8; font-size:13px; letter-spacing:1px; opacity:0.85;">
                    Auditoría de Turno &nbsp;·&nbsp; Planta
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with _user_col:
        _uh1, _uh2 = st.columns([3, 2])
        with _uh1:
            st.markdown(
                f"<div style='color:#666;font-size:12px;margin-top:10px;text-align:right;'>"
                f"👤 <b>{st.session_state.usuario}</b><br>"
                f"<span style='color:#aaa;font-size:11px;'>{st.session_state.rol}</span></div>",
                unsafe_allow_html=True,
            )
        with _uh2:
            st.markdown("<div style='margin-top:6px;'></div>", unsafe_allow_html=True)
            if st.button("↩ Salir", key="btn_logout"):
                st.session_state.usuario = None
                st.session_state.rol = None
                st.rerun()

    _rol = st.session_state.get("rol", "produccion")
    if _rol == "admin":
        tab_captura, tab_dashboard, tab_productividad, tab_admin = st.tabs(
            [
                "📦 CAPTURA Y AUDITORIA",
                "📊 VER RESULTADOS",
                "📈 PRODUCTIVIDAD",
                "⚙️ ADMINISTRACIÓN",
            ]
        )
    elif _rol != "produccion":
        tab_captura, tab_dashboard, tab_productividad = st.tabs(
            ["📦 CAPTURA Y AUDITORIA", "📊 VER RESULTADOS", "📈 PRODUCTIVIDAD"]
        )
    else:
        tab_dashboard, tab_productividad = st.tabs(
            ["📊 VER RESULTADOS", "📈 PRODUCTIVIDAD"]
        )

    if _rol != "produccion":
        with tab_captura:
            # ── Columnas: Captura (izq 55%) | Referencias (der 45%) ─────────
            _col_form, _col_ref = st.columns([11, 9])

            # ── Paso 1: FECHA y ÁREA en columna de captura ───────────────────
            with _col_form:
                _fc1, _fc2 = st.columns(2)
                with _fc1:
                    _fecha_dt = st.date_input(
                        "📅 FECHA", ahora_local().date(), key="cap_fecha"
                    )
                    fecha_sel = _fecha_dt.strftime("%d/%m/%Y")
                with _fc2:
                    _lista_areas = obtener_areas_con_programa(
                        df_programa, col_prog, fecha_sel
                    )
                    _nav_a = st.session_state.get("nav_area")
                    if _nav_a and _nav_a in _lista_areas:
                        del st.session_state["nav_area"]
                        st.session_state["cap_area"] = _nav_a
                    area_sel = st.selectbox("🏭 ÁREA", _lista_areas, key="cap_area")

            # ── Datos: se calculan con fecha y área antes de renderizar más ──
            df_plan_dia = obtener_plan_del_dia(
                df_programa, col_prog, fecha_sel, area_sel
            )
            df_aud_hoy, col_aud = obtener_auditorias_hoy(
                df_auditorias, fecha_sel, area_sel
            )
            avance_global, df_resumen_final = calcular_resumen(
                df_plan_dia,
                col_prog,
                df_bdd,
                col_bdd,
                df_auditorias,
                fecha_sel,
                area_sel,
            )

            _sin_programa = df_plan_dia.empty
            _cortes_info = {}
            for _ck, _ch in CORTES_DICT.items():
                if _sin_programa:
                    # Sin programa: auto-completo, no requiere auditoría
                    _cortes_info[_ck] = {
                        "horas": _ch,
                        "pendientes": 0,
                        "completo": True,
                        "n_cap": 0,
                    }
                    continue
                _pend_c = obtener_piezas_pendientes(
                    df_plan_dia,
                    col_prog,
                    df_bdd,
                    col_bdd,
                    df_aud_hoy,
                    col_aud,
                    area_sel,
                    _ck,
                )
                _caps_c = (
                    df_aud_hoy[df_aud_hoy[col_aud["corte"]] == _ck]
                    if not df_aud_hoy.empty and col_aud.get("corte")
                    else pd.DataFrame()
                )
                _completo = not _caps_c.empty and len(_pend_c) == 0
                _cortes_info[_ck] = {
                    "horas": _ch,
                    "pendientes": len(_pend_c),
                    "completo": _completo,
                    "n_cap": len(_caps_c),
                }
            # Secuencial: solo el primer corte incompleto disponible
            _primer_pendiente = next(
                (k for k, v in _cortes_info.items() if not v["completo"]), None
            )
            _cortes_disp = [_primer_pendiente] if _primer_pendiente else []

            # ── Detección de área/corte recién completada ─────────────────────
            _flag_ac = st.session_state.area_corte_completada
            _mostrar_interstitial = False
            _interstitial_area = None
            _interstitial_corte = None
            if _flag_ac:
                _prev_area, _prev_corte = _flag_ac
                if (
                    _prev_area == area_sel
                    and _prev_corte in _cortes_info
                    and _cortes_info[_prev_corte]["completo"]
                ):
                    _mostrar_interstitial = True
                    _interstitial_area = _prev_area
                    _interstitial_corte = _prev_corte
                    # NO limpiar aquí — se limpia solo en el handler del botón
                else:
                    # Las condiciones ya no aplican (cambió área o el corte dejó de estar completo)
                    st.session_state.area_corte_completada = None

            # ── Paso 2: CORTE + formulario en columna de captura ─────────────
            with _col_form:
                if _mostrar_interstitial:
                    # ── INTERSTITIAL: área completada ─────────────────────────
                    corte_sel = _interstitial_corte
                    horas_acum = CORTES_DICT.get(corte_sel, 0)

                    st.markdown(
                        f"<div style='background:#27ae60;color:white;padding:14px 20px;"
                        f"border-radius:10px;font-weight:bold;font-size:18px;margin-bottom:16px;'>"
                        f"✅ {_interstitial_area} — {_interstitial_corte} AUDITADO</div>",
                        unsafe_allow_html=True,
                    )

                    _df_comp = (
                        df_aud_hoy[df_aud_hoy[col_aud["corte"]] == _interstitial_corte]
                        if not df_aud_hoy.empty and col_aud.get("corte")
                        else pd.DataFrame()
                    )
                    if not _df_comp.empty:
                        st.markdown(
                            f"**Registros capturados — {_interstitial_corte}:**"
                        )
                        _col_hora_comp = encontrar_columna(
                            _df_comp, ["HORA", "HORA REGISTRO"]
                        )
                        _comp_cols = [
                            c
                            for c in [
                                col_aud.get("pieza"),
                                col_aud.get("subproceso"),
                                col_aud.get("real"),
                                _col_hora_comp,
                            ]
                            if c
                        ]
                        _df_comp_show = _df_comp[_comp_cols].copy()
                        _df_comp_show.columns = (
                            ["PIEZA", "SUBPROCESO", "REAL", "HORA"]
                        )[: len(_comp_cols)]
                        st.dataframe(
                            _df_comp_show, hide_index=True, use_container_width=True
                        )

                    st.divider()

                    _todas_areas = obtener_lista_areas(df_programa, col_prog)
                    _areas_pendientes = obtener_areas_pendientes_corte(
                        df_programa,
                        col_prog,
                        df_auditorias,
                        df_bdd,
                        col_bdd,
                        fecha_sel,
                        _interstitial_corte,
                        _todas_areas,
                        _interstitial_area,
                    )

                    if _areas_pendientes:
                        st.markdown(f"**Áreas pendientes para {_interstitial_corte}:**")
                        _btn_cols = st.columns(len(_areas_pendientes))
                        for _bi, _ar_pend in enumerate(_areas_pendientes):
                            with _btn_cols[_bi]:
                                if st.button(
                                    f"→ {_ar_pend}",
                                    key=f"btn_sig_area_{_ar_pend}",
                                    use_container_width=True,
                                    type="primary",
                                ):
                                    st.session_state.area_corte_completada = None
                                    st.session_state.nav_area = _ar_pend
                                    st.session_state.nav_corte = _interstitial_corte
                                    st.rerun()
                    else:
                        st.markdown(
                            f"<div style='background:#1a5276;color:white;padding:14px 20px;"
                            f"border-radius:10px;font-weight:bold;font-size:16px;margin-top:8px;'>"
                            f"🏁 Corte {_interstitial_corte} — todas las áreas han sido auditadas</div>",
                            unsafe_allow_html=True,
                        )
                        if st.button(
                            "📊 Ver análisis del turno",
                            use_container_width=True,
                            key="btn_ver_analisis_turno",
                        ):
                            st.session_state.area_corte_completada = None
                            st.rerun()

                elif _cortes_disp:
                    # ── FORMULARIO ACTIVO ─────────────────────────────────────
                    corte_sel = _cortes_disp[0]
                    st.session_state["cap_corte"] = corte_sel
                    horas_acum = CORTES_DICT[corte_sel]
                    _es_ultimo = corte_sel == list(CORTES_DICT.keys())[-1]
                    _msg_corte = (
                        f"⏱️ Corte a capturar: **{corte_sel}** — último corte del turno."
                        if _es_ultimo
                        else f"⏱️ Corte a capturar: **{corte_sel}** — completa este corte para desbloquear el siguiente."
                    )
                    st.info(_msg_corte, icon=None)

                    # Estado del corte: piezas pendientes con nombres
                    _info_corte = _cortes_info.get(corte_sel, {})
                    _n_pend_corte = _info_corte.get("pendientes", 0)
                    _n_cap_corte = _info_corte.get("n_cap", 0)
                    _total_piezas = (
                        len(df_plan_dia[col_prog["pieza"]].unique())
                        if not df_plan_dia.empty and col_prog.get("pieza")
                        else 0
                    )
                    if _n_pend_corte > 0:
                        _pend_nombres = obtener_piezas_pendientes(
                            df_plan_dia,
                            col_prog,
                            df_bdd,
                            col_bdd,
                            df_aud_hoy,
                            col_aud,
                            area_sel,
                            corte_sel,
                        )
                        _auditadas = _total_piezas - len(_pend_nombres)
                        st.warning(
                            f"**Corte {corte_sel}: {_auditadas}/{_total_piezas} piezas auditadas.**  \n"
                            f"Pendientes: {', '.join(_pend_nombres)}.  \n"
                            f"Captura cada pieza aunque la cantidad real sea 0.",
                            icon="⚠️",
                        )

                    st.markdown(
                        f"<div style='background:#8B1A1A;color:white;padding:10px 18px;"
                        f"border-radius:8px;font-weight:bold;font-size:15px;"
                        f"margin:8px 0 14px 0;letter-spacing:0.5px;'>"
                        f"📝 CAPTURA DE AUDITORÍA &nbsp;·&nbsp; {corte_sel}</div>",
                        unsafe_allow_html=True,
                    )

                    piezas_validas = (
                        df_plan_dia[col_prog["pieza"]].unique()
                        if not df_plan_dia.empty and col_prog.get("pieza")
                        else []
                    )
                    piezas_pendientes = obtener_piezas_pendientes(
                        df_plan_dia,
                        col_prog,
                        df_bdd,
                        col_bdd,
                        df_aud_hoy,
                        col_aud,
                        area_sel,
                        corte_sel,
                    )
                    lista_desplegable = (
                        piezas_pendientes if piezas_pendientes else list(piezas_validas)
                    )
                    f_id = st.session_state.form_id

                    if not lista_desplegable:
                        st.warning("No hay piezas disponibles para auditar.")
                        p_sel = None
                        df_s = pd.DataFrame()
                        sub_list = []
                    else:
                        _n_total = len(piezas_validas)
                        _n_pend = len(piezas_pendientes)
                        p_sel = st.selectbox(
                            f"PIEZA — {_n_pend} de {_n_total} pendiente(s)",
                            lista_desplegable,
                            key=f"p_{f_id}",
                            help="Piezas pendientes de auditar en este corte. Si el proceso estuvo detenido, captura cantidad real en 0 para registrar el paro y completar la auditoría del corte.",
                        )
                        if not validar_columnas(
                            col_bdd, ["pieza", "proceso", "subproceso"]
                        ):
                            df_s = df_bdd[
                                (df_bdd[col_bdd["pieza"]] == p_sel)
                                & (df_bdd[col_bdd["proceso"]] == area_sel)
                            ].copy()
                        else:
                            df_s = pd.DataFrame()
                            st.error("No se encontraron columnas clave en la hoja BDD.")
                        reps = []
                        if (
                            p_sel
                            and not df_aud_hoy.empty
                            and col_aud.get("corte")
                            and col_aud.get("pieza")
                        ):
                            reps = df_aud_hoy[
                                (df_aud_hoy[col_aud["corte"]] == corte_sel)
                                & (df_aud_hoy[col_aud["pieza"]] == p_sel)
                            ][col_aud["subproceso"]].tolist()
                        sub_list = (
                            [
                                s
                                for s in df_s[col_bdd["subproceso"]].unique()
                                if s not in reps
                            ]
                            if not df_s.empty
                            else []
                        )
                        if reps:
                            st.caption(
                                f"Ya capturados en {corte_sel}: {', '.join(reps)}"
                            )

                    s_sel = st.selectbox(
                        "SUB-PROCESO",
                        sub_list if sub_list else [PIEZA_TERMINADA],
                        key=f"s_{f_id}",
                        help="Operación a auditar.",
                    )
                    _ff1, _ff2 = st.columns(2)
                    with _ff1:
                        ops = st.number_input(
                            "OPERADORES",
                            min_value=1,
                            step=1,
                            value=st.session_state.last_ops,
                            key=f"ops_{f_id}",
                            help="Número de personas trabajando en este subproceso durante el corte.",
                        )
                        real = st.number_input(
                            "CANTIDAD REAL",
                            min_value=0,
                            step=1,
                            key=f"r_{f_id}",
                            help="Piezas buenas producidas hasta este corte. No incluir rechazos. Si el proceso estuvo detenido, captura 0 — es obligatorio registrar aunque no haya producción.",
                        )
                    with _ff2:
                        mins = st.number_input(
                            "MIN. PARO",
                            min_value=0,
                            step=1,
                            key=f"m_{f_id}",
                            help="Minutos de paro no productivo (fallas, esperas, descansos). "
                            "Se descuentan del tiempo disponible para calcular la meta.",
                        )
                        mot = st.selectbox(
                            "MOTIVO PARO",
                            MOTIVOS_PARO,
                            key=f"mot_{f_id}",
                            help="Causa principal del paro. Si no hubo paro, deja 'SIN PARO'.",
                        )
                    notas = st.text_input(
                        "NOTAS",
                        key=f"n_{f_id}",
                        help="Observaciones adicionales. Obligatorio si la cantidad supera 1.5× la meta "
                        "o el motivo de paro es 'OTRO (ESPECIFICAR EN NOTAS)'.",
                    )

                    if s_sel and s_sel != PIEZA_TERMINADA:
                        try:
                            if (
                                not df_s.empty
                                and s_sel in df_s[col_bdd["subproceso"]].values
                            ):
                                if not col_bdd.get("pzxh"):
                                    st.warning(
                                        "No se encontró la columna CAP PXH en la BDD."
                                    )
                                else:
                                    pz_h_val = df_s[
                                        df_s[col_bdd["subproceso"]] == s_sel
                                    ][col_bdd["pzxh"]].iloc[0]
                                    pz_h = float(pz_h_val) if pz_h_val else 0
                                    meta = int(
                                        (pz_h * max(0, horas_acum - (mins / 60))) * ops
                                    )
                                    efic = (
                                        round(real / meta * 100, 1) if meta > 0 else 0
                                    )
                                    _ec = (
                                        "#27ae60"
                                        if efic >= 80
                                        else "#f39c12" if efic >= 70 else "#e74c3c"
                                    )
                                    st.markdown(
                                        f"<div style='background:#f8f9fa;border-left:4px solid {_ec};"
                                        f"padding:10px 14px;border-radius:6px;margin:8px 0;'>"
                                        f"<b>Meta: {meta} pzs</b> &nbsp;·&nbsp; "
                                        f"<span style='color:{_ec};font-weight:bold;'>"
                                        f"Eficiencia actual: {efic}%</span></div>",
                                        unsafe_allow_html=True,
                                    )
                                    _bloqueo_real = (
                                        meta > 0
                                        and real > meta * 1.5
                                        and not notas.strip()
                                    )
                                    _bloqueo_mot = (
                                        mot == "OTRO (ESPECIFICAR EN NOTAS)"
                                        and not notas.strip()
                                    )
                                    if _bloqueo_real:
                                        st.warning(
                                            f"⚠️ Cantidad real ({real}) supera 1.5× la meta ({meta}). "
                                            f"Agrega una nota explicando el motivo antes de guardar."
                                        )
                                    if _bloqueo_mot:
                                        st.warning(
                                            "Captura una nota cuando el motivo sea OTRO."
                                        )
                                    if not _bloqueo_real and not _bloqueo_mot:
                                        if st.button(
                                            "💾 GUARDAR REGISTRO",
                                            type="primary",
                                            use_container_width=True,
                                            key=f"btn_{f_id}",
                                            disabled=st.session_state.guardando,
                                        ):
                                            st.session_state.guardando = True
                                            with st.spinner("Guardando..."):
                                                exito = guardar_registro(
                                                    fecha_sel,
                                                    area_sel,
                                                    corte_sel,
                                                    p_sel,
                                                    s_sel,
                                                    real,
                                                    meta,
                                                    ops,
                                                    mot,
                                                    mins,
                                                    notas,
                                                    st.session_state.get("usuario", ""),
                                                )
                                            st.session_state.guardando = False
                                            if exito:
                                                st.toast("✅ Guardado exitoso")
                                                invalidar_cache_hoja("AUDITAR")
                                                st.session_state.last_ops = ops
                                                st.session_state.area_corte_completada = (
                                                    area_sel,
                                                    corte_sel,
                                                )
                                                st.session_state.form_id += 1
                                                time.sleep(1.5)
                                                st.rerun()
                            else:
                                st.warning(
                                    "No hay datos de capacidad para este subproceso."
                                )
                        except ValueError:
                            st.error(
                                "La capacidad PZ X H contiene un valor no numérico."
                            )
                        except Exception as exc:
                            st.error(f"Error al calcular o guardar: {exc}")
                    else:
                        st.success("✅ Pieza completada.")

                    # ── Mini-historial del corte ──────────────────────────────
                    st.divider()
                    _df_mini_hist = (
                        df_aud_hoy[df_aud_hoy[col_aud["corte"]] == corte_sel]
                        if not df_aud_hoy.empty and col_aud.get("corte")
                        else pd.DataFrame()
                    )
                    if not _df_mini_hist.empty:
                        st.markdown(f"##### 🗒️ Ya guardados — {corte_sel}")
                        _col_hora_mini = encontrar_columna(
                            _df_mini_hist, ["HORA", "HORA REGISTRO"]
                        )
                        _mini_cols = [
                            c
                            for c in [
                                col_aud.get("pieza"),
                                col_aud.get("subproceso"),
                                col_aud.get("real"),
                                _col_hora_mini,
                            ]
                            if c
                        ]
                        _df_mini_show = _df_mini_hist[_mini_cols].copy()
                        _df_mini_show.columns = (
                            ["PIEZA", "SUBPROCESO", "REAL", "HORA"]
                        )[: len(_mini_cols)]
                        st.dataframe(
                            _df_mini_show, hide_index=True, use_container_width=True
                        )
                    else:
                        st.caption(f"Sin capturas para {corte_sel} todavía.")

                else:
                    # ── TURNO COMPLETADO ──────────────────────────────────────
                    corte_sel = list(CORTES_DICT.keys())[0]
                    horas_acum = CORTES_DICT[corte_sel]
                    st.markdown(
                        "<div style='background:#27ae60;color:white;padding:12px 18px;"
                        "border-radius:8px;font-weight:bold;font-size:16px;margin:8px 0;'>"
                        "✅ TURNO AUDITADO — TODOS LOS CORTES COMPLETADOS</div>",
                        unsafe_allow_html=True,
                    )

            # ── Columna de referencias ────────────────────────────────────────
            with _col_ref:
                st.markdown(
                    f"<div style='background:#8B1A1A;color:white;padding:8px 16px;"
                    f"border-radius:8px;font-weight:bold;font-size:14px;margin-bottom:10px;'>"
                    f"SISTEMA <span style='color:#C0C0C0;'>NSG</span> AUDITORÍA "
                    f"&nbsp;·&nbsp; {area_sel}</div>",
                    unsafe_allow_html=True,
                )
                render_kpis(avance_global, df_resumen_final)
                with st.expander("ℹ️ ¿Cómo leer estos números?", expanded=False):
                    st.markdown(
                        "**Eficiencia Global** — promedio de (Real ÷ Programado) por cada pieza "
                        "y subproceso auditado. Meta mínima: **80%**.\n\n"
                        "**Meta Turno** — total de piezas programadas en el día para esta área.\n\n"
                        "**Real Turno** — piezas registradas hasta el momento en todos los cortes.\n\n"
                        "**Cortes (badges)** — 🟢 completado · 🟠 pendiente · ⚫ sin captura aún.\n\n"
                        "**Avance por Corte** — aportación de cada corte al total del día. "
                        "Cada bloque vale idealmente 33.3%. Verde = ≥ 26.6% (ritmo para cumplir el 80%)."
                    )
                st.divider()

                # Badges de estado de cortes
                _badges = []
                for _ck, _cv in _cortes_info.items():
                    if _cv["completo"]:
                        _badges.append(
                            f"<span style='background:#27ae60;color:white;padding:5px 12px;"
                            f"border-radius:20px;margin:2px;font-size:11px;font-weight:bold;"
                            f"display:inline-block;'>✅ {_ck} ({_cv['n_cap']})</span>"
                        )
                    elif _cortes_disp and _ck == corte_sel:
                        _badges.append(
                            f"<span style='background:#e67e22;color:white;padding:5px 12px;"
                            f"border-radius:20px;margin:2px;font-size:11px;font-weight:bold;"
                            f"display:inline-block;'>⚡ {_ck} — {_cv['pendientes']} pend.</span>"
                        )
                    else:
                        _badges.append(
                            f"<span style='background:#95a5a6;color:white;padding:5px 12px;"
                            f"border-radius:20px;margin:2px;font-size:11px;"
                            f"display:inline-block;'>🕐 {_ck}</span>"
                        )
                st.markdown(
                    "<div style='line-height:2.2;'>" + " ".join(_badges) + "</div>",
                    unsafe_allow_html=True,
                )

                # Aviso próximo
                _hora_corte_map = {
                    "11:00 AM (3h)": 11,
                    "14:00 PM (6h)": 14,
                    "17:00 PM (9h)": 17,
                }
                if ahora_local().hour < _hora_corte_map.get(corte_sel, 0):
                    st.info(
                        f"ℹ️ El corte **{corte_sel}** aún no ha llegado. "
                        f"Puedes capturar anticipado."
                    )

                st.divider()

                # Programa del día
                with st.expander("📋 Programa del Día", expanded=True):
                    if not df_plan_dia.empty:
                        _prog_show = df_plan_dia[
                            [col_prog["pieza"], col_prog["total"]]
                        ].copy()
                        _prog_show.columns = ["PIEZA", "PROGRAMADO"]
                        _prog_show["PROGRAMADO"] = (
                            pd.to_numeric(_prog_show["PROGRAMADO"], errors="coerce")
                            .fillna(0)
                            .astype(int)
                        )
                        if not df_resumen_final.empty:
                            _agg_prog = (
                                df_resumen_final.groupby("PIEZA")["% REAL"]
                                .mean()
                                .round(1)
                                .reset_index()
                            )
                            _agg_prog.columns = ["PIEZA", "%"]
                            _agg_prog["%"] = _agg_prog["%"].clip(0, 100)
                            _prog_show = _prog_show.merge(
                                _agg_prog[["PIEZA", "%"]], on="PIEZA", how="left"
                            )
                            _prog_show["%"] = _prog_show["%"].fillna(0.0)
                        st.dataframe(
                            _prog_show,
                            column_config={
                                "PROGRAMADO": st.column_config.NumberColumn(
                                    "PROG", format="%d"
                                ),
                                "%": st.column_config.ProgressColumn(
                                    "AVANCE %",
                                    format="%.0f%%",
                                    min_value=0,
                                    max_value=100,
                                ),
                            },
                            hide_index=True,
                            use_container_width=True,
                            height=250,
                        )
                    else:
                        st.info("No hay piezas programadas.")

                # Checklist de cobertura
                render_checklist_cobertura(
                    df_plan_dia,
                    col_prog,
                    df_aud_hoy,
                    col_aud,
                    df_bdd,
                    col_bdd,
                    area_sel,
                )
                render_capacidades(
                    df_s if "df_s" in locals() else pd.DataFrame(),
                    col_bdd,
                    p_sel if "p_sel" in locals() else None,
                )

            # ── Analytics: ancho completo debajo de ambas columnas ───────────
            with st.expander("📊 Ver análisis del turno", expanded=True):
                render_progreso_por_corte(
                    df_aud_hoy,
                    col_aud,
                    df_plan_dia,
                    col_prog,
                    df_bdd,
                    col_bdd,
                    area_sel,
                )
                render_graficos(avance_global, df_resumen_final)
                st.divider()
                render_estatus_detallado(df_resumen_final)

    with tab_dashboard:
        try:
            render_dashboard_direccion(
                df_auditorias, df_programa, df_bdd, col_prog, col_bdd
            )
        except Exception as _err_dash:
            st.error(f"Error en el módulo de Resultados: {_err_dash}")
            import traceback as _tb

            with st.expander("Detalle técnico"):
                st.code(_tb.format_exc())

    with tab_productividad:
        try:
            render_productividad(df_programa, col_prog, df_bdd, col_bdd)
        except Exception as _err_prod:
            st.error(
                f"Error en el módulo de Productividad: {_err_prod}\n\n"
                "Las pestañas de Captura y Resultados siguen funcionando con normalidad."
            )
            import traceback as _tb

            with st.expander("Detalle técnico del error"):
                st.code(_tb.format_exc())

    if _rol == "admin":
        with tab_admin:
            render_admin()

    st.divider()
    with st.expander("🔑 Cambiar mi contraseña", expanded=False):
        with st.form("form_cambiar_pass"):
            _cp_actual = st.text_input("Contraseña actual", type="password")
            _cp_nueva = st.text_input("Nueva contraseña", type="password")
            _cp_conf = st.text_input("Confirmar nueva contraseña", type="password")
            if st.form_submit_button("Actualizar"):
                if not _cp_actual or not _cp_nueva or not _cp_conf:
                    st.error("Completa todos los campos.")
                elif _cp_nueva != _cp_conf:
                    st.error("La nueva contraseña y la confirmación no coinciden.")
                else:
                    _lista_u = _leer_usuarios()
                    _hash_actual = hashlib.sha256(_cp_actual.encode()).hexdigest()
                    _idx = next(
                        (
                            i
                            for i, u in enumerate(_lista_u)
                            if u["USUARIO"] == st.session_state.usuario
                            and u["HASH"] == _hash_actual
                        ),
                        None,
                    )
                    if _idx is None:
                        st.error("La contraseña actual es incorrecta.")
                    else:
                        _lista_u[_idx]["HASH"] = hashlib.sha256(
                            _cp_nueva.encode()
                        ).hexdigest()
                        if _guardar_usuarios(_lista_u):
                            st.success("Contraseña actualizada correctamente.")
                        else:
                            st.error("Error al guardar. Intenta de nuevo.")


if __name__ == "__main__":
    main()
